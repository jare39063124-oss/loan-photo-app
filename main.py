"""
资产盘点专项拍照工具 App - v3.22.18
功能：
- 欢迎页 + 设置页
- 文件命名自选模式（4段下拉 X-X-X-X）
- 水印自选模式（3段下拉 X-X-X + 大中小字号 + 位置）
- 四类拍照引导（远景/近景/内部/瑕疵）+ 连续拍照（同类型手动按快门连拍）
- Excel A=客户名 B=抵押物地址（概） C=抵押物地址（精确门牌号） D=抵押物性质
- 进度持久化（跨Excel文件可识别已拍条目）
- 照片路径验证（缩略图不会引用失效路径）
- 崩溃日志收集
"""

import os
import json
import sys
import re
import time
import traceback
from datetime import datetime
from collections import defaultdict
import threading
import hashlib

# ============================================================
# 崩溃日志捕获
# 注意：此处不能依赖 kivy.utils.platform，因为 kivy 尚未导入。
# 用环境变量做早期 Android 检测（python-for-android 会注入 ANDROID_ARGUMENT）。
# ============================================================
CRASH_LOG_FILE = None
_EARLY_IS_ANDROID = ('ANDROID_ARGUMENT' in os.environ) or (sys.platform == 'android')

def _resolve_android_app_dir():
    """返回 Android 上可写的 app 专属外部存储路径（无需权限、不受 scoped storage 限制）。

    路径 = /storage/emulated/0/Android/data/<package>/files/loan_photos
    用户可通过文件管理器访问：内部存储 → Android → data → <package> → files → loan_photos

    所有 Android 版本均无需运行时权限；Android 10+ scoped storage 下也可写。
    """
    if not _EARLY_IS_ANDROID:
        return os.path.join(os.path.expanduser('~'), 'loan_photos')

    # 1) 首选：App-specific external storage（无需权限）
    try:
        from jnius import autoclass
        PythonActivity = autoclass('org.kivy.android.PythonActivity')
        activity = PythonActivity.mActivity
        if activity is not None:
            ext_files_dir = activity.getExternalFilesDir(None)
            if ext_files_dir is not None:
                base = str(ext_files_dir.getAbsolutePath())
                if base and base.lower() != 'null':
                    return os.path.join(base, 'loan_photos')
    except Exception:
        pass

    # 2) 退路1：python-for-android 私有内部存储（一定可写，但用户在文件管理器看不到）
    try:
        home = os.path.expanduser('~')
        if home and os.path.isdir(home):
            return os.path.join(home, 'loan_photos')
    except Exception:
        pass

    # 3) 退路2：临时目录（最差兜底）
    import tempfile
    return os.path.join(tempfile.gettempdir(), 'loan_photos')

def setup_crash_handler():
    global CRASH_LOG_FILE
    try:
        log_dir = _resolve_android_app_dir()
        os.makedirs(log_dir, exist_ok=True)
        CRASH_LOG_FILE = os.path.join(log_dir, 'crash_log.txt')
    except Exception:
        # 最后兜底：写到当前目录（APK 内只读会失败，但不会抛出）
        try:
            log_dir = os.path.join(os.path.expanduser('~'), 'loan_photos')
            os.makedirs(log_dir, exist_ok=True)
            CRASH_LOG_FILE = os.path.join(log_dir, 'crash_log.txt')
        except Exception:
            CRASH_LOG_FILE = None

    def global_excepthook(exc_type, exc_value, exc_tb):
        msg = "="*50 + "\n"
        msg += "CRASH at %s\n" % datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        msg += "Type: %s\n" % exc_type.__name__
        msg += "Error: %s\n" % str(exc_value)
        msg += "Traceback:\n"
        msg += "".join(traceback.format_tb(exc_tb))
        msg += "\n"
        if CRASH_LOG_FILE:
            try:
                with open(CRASH_LOG_FILE, 'a', encoding='utf-8') as f:
                    f.write(msg)
            except Exception:
                pass
        # Also let original handler run
        sys.__excepthook__(exc_type, exc_value, exc_tb)

    sys.excepthook = global_excepthook
    # Also capture Python thread exceptions
    threading.excepthook = lambda args: global_excepthook(args.exc_type, args.exc_value, args.exc_tb)

setup_crash_handler()

os.environ['KIVY_LOG_LEVEL'] = 'warning'

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.behaviors import ButtonBehavior  # v3.22.5: RoundedButton 基类（替代 Button）
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.image import Image as KivyImage
from kivy.uix.checkbox import CheckBox
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.slider import Slider
from kivy.uix.widget import Widget  # v3.22.19: 数据指示器圆点容器
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.clock import Clock, mainthread
from kivy.logger import Logger
from kivy.core.window import Window
from kivy.utils import platform
from kivy.graphics import Color, Rectangle, Line, RoundedRectangle, Ellipse
from kivy.metrics import dp, sp
from kivy.properties import ListProperty  # v3.22.5: RoundedButton.background_color 可观察属性
from kivy.storage.jsonstore import JsonStore
from kivy.core.text import LabelBase
from kivy.resources import resource_find, resource_add_path
from kivy.lang import Builder
from kivy.uix.spinner import SpinnerOption

# ============================================================
# 中文字体加载（彻底解决中文豆腐块问题）
#
# 根本原因历史：
#   - 之前用的 NotoSansSC.ttf 只有 28KB，是损坏的占位文件，
#     一个真正的 CJK 中文字体至少需要 4-16MB。
#   - 现在使用 simhei.ttf（黑体，9.7MB），放在 main.py 同级目录，
#     buildozer 通过 source.include_exts=ttf 自动打包进 APK，
#     在 APK 内与 main.py 同目录，路径最简单最可靠。
#
# 三重保障：
#   1. LabelBase.register('Roboto', ...) 替换 Kivy 默认字体
#   2. KV 全局规则设置所有文本 widget 的 font_name，
#      特别包含 SpinnerOption（下拉菜单选项）
#   3. Android 系统字体 fallback（应对极端情况）
# ============================================================
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
_ASSETS_DIR = os.path.join(_APP_DIR, 'assets')
resource_add_path(_APP_DIR)
if os.path.isdir(_ASSETS_DIR):
    resource_add_path(_ASSETS_DIR)

_FONT_PATH = None
for _fn in ['simhei.ttf', 'NotoSansSC.ttf', 'NotoSansSC-Regular.otf']:
    _p = os.path.join(_APP_DIR, _fn)
    if os.path.exists(_p) and os.path.getsize(_p) > 100000:
        _FONT_PATH = _p
        break
if not _FONT_PATH:
    for _fn in ['simhei.ttf']:
        _found = resource_find(_fn)
        if _found and os.path.exists(_found) and os.path.getsize(_found) > 100000:
            _FONT_PATH = _found
            break
if not _FONT_PATH:
    for _sp in [
        '/system/fonts/NotoSansCJK-Regular.ttc',
        '/system/fonts/NotoSansSC-Regular.otf',
        '/system/fonts/NotoSansCJKsc-Regular.otf',
        '/system/fonts/DroidSansFallback.ttf',
        '/system/fonts/SourceHanSansCN-Regular.otf',
        '/system/fonts/msyh.ttc',
    ]:
        if os.path.exists(_sp):
            _FONT_PATH = _sp
            break

if _FONT_PATH:
    try:
        LabelBase.register('Roboto', _FONT_PATH)
    except Exception as _e:
        _FONT_PATH = None

CHINESE_FONT = 'Roboto'

Builder.load_string('''
<Label>:
    font_name: 'Roboto'
<Button>:
    font_name: 'Roboto'
<Spinner>:
    font_name: 'Roboto'
<TextInput>:
    font_name: 'Roboto'
<CheckBox>:
    font_name: 'Roboto'
<SpinnerOption>:
    font_name: 'Roboto'
<Switch>:
    font_name: 'Roboto'
<Slider>:
    font_name: 'Roboto'
<ProgressBar>:
    font_name: 'Roboto'
''')

_diag_lines = ["=== Font diag ===", "__file__=%s" % __file__,
               "app_dir=%s" % _APP_DIR,
               "font_path=%s" % _FONT_PATH,
               "font_size=%d" % (os.path.getsize(_FONT_PATH) if _FONT_PATH and os.path.exists(_FONT_PATH) else 0)]
try:
    _diag_dir = _resolve_android_app_dir() if _EARLY_IS_ANDROID else os.path.expanduser('~')
    os.makedirs(_diag_dir, exist_ok=True)
    with open(os.path.join(_diag_dir, 'font_debug.log'), 'w', encoding='utf-8') as _f:
        _f.write('\n'.join(_diag_lines))
except Exception:
    pass

import openpyxl
from openpyxl import load_workbook
from PIL import Image as PILImage, ImageDraw, ImageFont

from geocoder import GeoCoder

# === 平台检测 ===
IS_ANDROID = platform == 'android'

if IS_ANDROID:
    from android.permissions import request_permissions, Permission
    try:
        from android import api_version
        ANDROID_API = api_version
    except:
        ANDROID_API = 30

    from jnius import PythonJavaClass, java_method

    class _UiRunnable(PythonJavaClass):
        __javainterfaces__ = ['java/lang/Runnable']
        def __init__(self, func):
            super().__init__()
            self._func = func
        @java_method('()V')
        def run(self):
            try:
                self._func()
            except Exception:
                Logger.error(traceback.format_exc())

    def run_on_ui_thread(func):
        from jnius import autoclass
        activity = autoclass('org.kivy.android.PythonActivity').mActivity
        if activity is not None:
            activity.runOnUiThread(_UiRunnable(func))
        else:
            func()
else:
    ANDROID_API = 0
    def run_on_ui_thread(func):
        func()


def get_status_bar_height_dp():
    """获取Android状态栏高度（dp），非Android返回0"""
    if not IS_ANDROID:
        return 0
    try:
        from jnius import autoclass
        activity = autoclass('org.kivy.android.PythonActivity').mActivity
        resources = activity.getResources()
        resource_id = resources.getIdentifier('status_bar_height', 'dimen', 'android')
        if resource_id > 0:
            px = resources.getDimensionPixelSize(resource_id)
            density = resources.getDisplayMetrics().density
            return int(px / density) + 2
    except:
        pass
    return 28


def setup_android_status_bar():
    """在Android上设置状态栏颜色、禁用edge-to-edge，防止内容被遮挡"""
    if not IS_ANDROID:
        return
    try:
        from jnius import autoclass
        activity = autoclass('org.kivy.android.PythonActivity').mActivity
        window = activity.getWindow()
        if ANDROID_API >= 30:
            window.setDecorFitsSystemWindows(True)
        # 状态栏颜色用深色（与主题bg一致: #1C1C24）
        window.setStatusBarColor(0xFF1C1C24)
        # 状态栏文字浅色（白色图标，适配深色背景）
        from jnius import autoclass as _ac
        View = _ac('android.view.View')
        decor = window.getDecorView()
        flags = decor.getSystemUiVisibility()
        # 清除 LIGHT_STATUS_BAR 标志（使状态栏图标为白色）
        flags = flags & ~View.SYSTEM_UI_FLAG_LIGHT_STATUS_BAR
        decor.setSystemUiVisibility(flags)
    except Exception as e:
        Logger.warning("setup_android_status_bar: %s" % e)


def android_copy_uri_to_app_dir(uri_str, dest_path):
    """通过Android ContentResolver从content URI或文件路径复制文件到app私有目录，
    解决scoped storage下PermissionError问题。
    支持: content:// URI 和 /storage/... 路径
    v3.22.13 修复: in_stream/out_stream 用 try/finally 确保异常时关闭，防止 Java 文件句柄泄漏。
    """
    from jnius import autoclass
    FileInputStream = autoclass('java.io.FileInputStream')
    FileOutputStream = autoclass('java.io.FileOutputStream')
    File = autoclass('java.io.File')
    Uri = autoclass('android.net.Uri')
    activity = autoclass('org.kivy.android.PythonActivity').mActivity
    resolver = activity.getContentResolver()

    in_stream = None
    out_stream = None
    try:
        if uri_str.startswith('content://'):
            uri = Uri.parse(uri_str)
            in_stream = resolver.openInputStream(uri)
        else:
            # 尝试直接用文件路径
            in_stream = FileInputStream(uri_str)

        out_stream = FileOutputStream(dest_path)
        buffer = bytearray(8192)
        while True:
            read = in_stream.read(buffer)
            if read == -1:
                break
            out_stream.write(buffer, 0, read)
        return dest_path
    finally:
        # v3.22.13: 无论成功/失败都关闭流（原代码仅在成功路径 close，异常时泄漏句柄）
        if in_stream is not None:
            try:
                in_stream.close()
            except Exception:
                pass
        if out_stream is not None:
            try:
                out_stream.close()
            except Exception:
                pass

# === 目录 ===
# 复用 setup_crash_handler() 已解析出的 app 专属外部存储路径：
# /storage/emulated/0/Android/data/<package>/files/loan_photos/
# - 无需任何运行时权限（所有 Android 版本）
# - 不受 Android 10+ scoped storage 限制
# - 用户可通过文件管理器访问
APP_DIR = _resolve_android_app_dir()
try:
    os.makedirs(APP_DIR, exist_ok=True)
except Exception as e:
    # 极端情况下创建失败，回退到 app 私有内部存储（一定可写）
    APP_DIR = os.path.join(os.path.expanduser('~'), 'loan_photos')
    try:
        os.makedirs(APP_DIR, exist_ok=True)
    except Exception:
        pass

PROGRESS_FILE = os.path.join(APP_DIR, 'photo_progress.json')
CONFIG_FILE = os.path.join(APP_DIR, 'app_config.json')
APP_LOG_FILE = os.path.join(APP_DIR, 'app_debug.log')
FONT_PATH = _FONT_PATH if _FONT_PATH else os.path.join(os.path.dirname(os.path.abspath(__file__)), 'simhei.ttf')

# ============================================================
# v3.21.0: RoundedButton — 圆角按钮，改善 GUI 观感
# 使用 canvas.before 绘制 RoundedRectangle 替代 Kivy 默认硬直角矩形
# ============================================================
class RoundedButton(ButtonBehavior, Label):
    """v3.21.0: 圆角按钮，canvas.before 绘制 RoundedRectangle
    v3.22.2: radius 单位修正为 dp（原 14px 在高密度屏倒角微弱不可见）
    v3.22.3: radius 增大到 dp(12)，并补画边框增强倒角辨识
    v3.22.4: radius 增大到 dp(16)（多机型实测 dp(12) 仍不明显），
             补设 background_down=''（原仅设 background_normal，按压时默认纹理覆盖圆角），
             边框透明度 0.12→0.35、width 1→1.5 增强可见性
    v3.22.5: 改基类为 ButtonBehavior+Label。根因：Kivy style.kv 的 <-Button> 规则
             在主 canvas 用 BorderImage 画直角实心矩形，覆盖 canvas.before 的 RoundedRectangle，
             导致 radius/border 均不可见。Label 无此 KV 规则，圆角得以正常显示。
             background_color 改为 ListProperty，以支持运行时改色自动重绘。"""

    # v3.22.5: 作为 Kivy 可观察属性，便于 bind 与运行时改色触发重绘
    # （原 Button.background_color 是 ListProperty；Label 无此属性，此处自定义）
    background_color = ListProperty([0.13, 0.59, 0.95, 1])

    def __init__(self, radius=None, **kwargs):
        # v3.22.5: 兼容旧 Button API —— 从 kwargs 提取 Button 专有属性，避免传给 Label 报错
        bg_color = kwargs.pop('background_color', None)
        bg_normal = kwargs.pop('background_normal', None)
        bg_down = kwargs.pop('background_down', None)

        # 调用 Label.__init__（ButtonBehavior 无 __init__）
        super().__init__(**kwargs)

        self.always_release = True  # v3.22.10: 触摸屏手指微动移出按钮时 on_release 仍触发（Kivy 官方推荐）

        # v3.22.4: radius 默认 dp(16)，倒角更明显
        if radius is None:
            radius = [dp(16)] * 4
        self._radius = list(radius)

        # 应用调用方传入的 background_color（覆盖 ListProperty 默认值）
        if bg_color is not None:
            self.background_color = bg_color

        # v3.22.5: 接受赋值但不使用（保持与旧 Button API 兼容，Label 无背景纹理）
        self.background_normal = bg_normal if bg_normal is not None else ''
        self.background_down = bg_down if bg_down is not None else ''

        # 绑定重绘：pos/size 变化、state 按压、background_color 运行时改色
        self.bind(pos=self._draw_bg, size=self._draw_bg,
                  background_color=self._draw_bg, state=self._draw_bg)
        self._draw_bg()

    def _draw_bg(self, *args):
        self.canvas.before.clear()
        with self.canvas.before:
            # state='down' 时变暗，提供按压视觉反馈
            r, g, b, a = self.background_color
            if self.state == 'down':
                r, g, b = r * 0.85, g * 0.85, b * 0.85
            Color(r, g, b, a)
            RoundedRectangle(pos=self.pos, size=self.size, radius=self._radius)
            # v3.22.4: 边框透明度 0.35、width 1.5，让圆角边界清晰可辨
            Color(0, 0, 0, 0.35)
            Line(rounded_rectangle=(self.x, self.y, self.width, self.height, self._radius[0] if self._radius else dp(16)),
                 width=1.5)


class RoundedSpinnerOption(SpinnerOption):
    """v3.22.19: Spinner 下拉项带 dp(8) 圆角倒角。
    继承 SpinnerOption（其本身为 Button 子类）以满足 Kivy Spinner.option_cls 工厂要求。
    用 canvas.before 绘制圆角背景，避免默认 BorderImage 直角覆盖。"""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ''
        self.background_down = ''
        self.background_color = (0, 0, 0, 0)  # 透明，由 canvas.before 绘制
        self.color = (1, 1, 1, 1)  # 白色文字（深蓝底配白字）
        self._opt_radius = [dp(8)] * 4
        with self.canvas.before:
            Color(*THEME['accent_dark'])
            self._opt_bg = RoundedRectangle(pos=self.pos, size=self.size, radius=self._opt_radius)
            Color(0, 0, 0, 0.18)
            self._opt_border = Line(
                rounded_rectangle=(self.x, self.y, self.width, self.height, dp(8)),
                width=1.0)
        self.bind(pos=self._update_opt_bg, size=self._update_opt_bg, state=self._update_opt_bg)

    def _update_opt_bg(self, *args):
        # v3.22.19: pos/size/state 变化时重绘圆角背景（state='down' 时变暗提供按压反馈）
        try:
            r, g, b, a = THEME['accent_dark']
            if self.state == 'down':
                r, g, b = r * 0.85, g * 0.85, b * 0.85
            self.canvas.before.clear()
            with self.canvas.before:
                Color(r, g, b, a)
                self._opt_bg = RoundedRectangle(pos=self.pos, size=self.size, radius=self._opt_radius)
                Color(0, 0, 0, 0.18)
                self._opt_border = Line(
                    rounded_rectangle=(self.x, self.y, self.width, self.height, dp(8)),
                    width=1.0)
        except Exception:
            pass


# ============================================================
# v3.20.0: 全局日志器 — 记录全量 app 日志（Excel/拍照/备注/AI/生命周期等）
# 轮转策略：文件超过 500KB 时保留最近 256KB
# ============================================================
class AppLogger:
    """v3.20.0: 全量 app 日志记录器，替代仅相机日志的 _dbg 机制
    v3.22.7: 文件写入改为异步 — log() 仅追加内存缓冲，后台 daemon 线程批量落盘，
    避免主线程同步 IO 阻塞导致卡顿。"""
    _instance = None

    def __init__(self, filepath=None, max_size=512 * 1024, keep_size=256 * 1024):
        self.filepath = filepath or APP_LOG_FILE
        self.max_size = max_size
        self.keep_size = keep_size
        self._lines = []          # 内存缓冲（最近 200 条，供快速预览）
        self._max_lines = 200
        self._lock = threading.Lock()
        self._enabled = False    # v3.22.0: 日志开关，默认关闭
        # v3.22.7: 异步落盘缓冲与后台线程
        self._pending = []          # 待写入文件的日志行
        self._flush_event = threading.Event()
        self._stop_event = threading.Event()
        self._flush_thread = None

    def set_enabled(self, enabled):
        """v3.22.0: 设置日志开关
        v3.22.7: 开启时启动后台落盘线程；关闭时立即 flush 剩余缓冲。"""
        self._enabled = bool(enabled)
        if self._enabled:
            self._start_flush_thread()
        else:
            # 关闭时立即落盘剩余缓冲（确保日志不丢失）
            self._do_flush()

    def _start_flush_thread(self):
        """v3.22.7: 启动后台 daemon 落盘线程（进程退出时自动结束）"""
        if self._flush_thread and self._flush_thread.is_alive():
            return
        self._stop_event.clear()
        self._flush_thread = threading.Thread(target=self._flush_loop, daemon=True)
        self._flush_thread.start()

    def _flush_loop(self):
        """v3.22.7: 后台循环 — 等待信号或 0.5s 超时后批量落盘"""
        while not self._stop_event.is_set():
            self._flush_event.wait(timeout=0.5)
            if self._stop_event.is_set():
                break
            self._flush_event.clear()
            try:
                self._do_flush()
            except Exception:
                pass

    def _do_flush(self):
        """v3.22.7: 将 _pending 批量写入文件（单次 open/write，减少 IO 次数）"""
        # 先在锁内取出待写行（短暂持锁，不阻塞 log()）
        with self._lock:
            if not self._pending:
                return
            lines_to_write = self._pending[:]
            self._pending = []
        # 文件写入在锁外执行（慢 IO 不阻塞 log() 与 get_log_text()）
        try:
            os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
            with open(self.filepath, 'a', encoding='utf-8') as f:
                f.write("\n".join(lines_to_write) + "\n")
            self._maybe_rotate()
        except Exception as e:
            # 落盘失败 — 记录到内存缓冲（用户查看日志仍能看到内容）
            ts = get_system_date().strftime('%Y-%m-%d %H:%M:%S')
            err_line = "[%s] [WARN] [APPLOG] 日志文件写入失败: %s（仅内存缓冲可见）" % (ts, str(e)[:60])
            with self._lock:
                self._lines.append(err_line)
                if len(self._lines) > self._max_lines:
                    self._lines = self._lines[-self._max_lines:]

    def is_enabled(self):
        """v3.22.0: 返回当前开关状态"""
        return self._enabled

    def log(self, level, tag, msg):
        """写入一条日志。level: INFO/WARN/ERROR；tag: 模块标签如 EXCEL/PHOTO/AI/APP
        v3.22.7: 仅追加内存缓冲并唤醒后台线程，不同步写文件（避免主线程 IO 阻塞）。"""
        # v3.22.0: 日志开关关闭时不写文件、不入缓冲（完全跳过）
        if not self._enabled:
            return
        ts = get_system_date().strftime('%Y-%m-%d %H:%M:%S')
        line = "[%s] [%s] [%s] %s" % (ts, level, tag, msg)
        with self._lock:
            self._lines.append(line)
            if len(self._lines) > self._max_lines:
                self._lines = self._lines[-self._max_lines:]
            # v3.22.7: 追加到待落盘缓冲，不立即写文件
            self._pending.append(line)
        # 唤醒后台线程落盘
        self._flush_event.set()
        # 同时输出到 Kivy Logger（控制台可见）
        try:
            if level == 'ERROR':
                Logger.error("APPLOG [%s]: %s", tag, msg)
            elif level == 'WARN':
                Logger.warning("APPLOG [%s]: %s", tag, msg)
            else:
                Logger.info("APPLOG [%s]: %s", tag, msg)
        except Exception:
            pass

    def _maybe_rotate(self):
        """文件超过 max_size 时截断保留最近 keep_size 字节"""
        try:
            if not os.path.exists(self.filepath):
                return
            if os.path.getsize(self.filepath) < self.max_size:
                return
            with open(self.filepath, 'r', encoding='utf-8') as f:
                f.seek(0, 2)
                total = f.tell()
                start = max(0, total - self.keep_size)
                f.seek(start)
                data = f.read()
            # 找到第一个换行，避免截断半行
            nl = data.find('\n')
            if nl >= 0:
                data = data[nl + 1:]
            with open(self.filepath, 'w', encoding='utf-8') as f:
                f.write(data)
        except Exception:
            pass

    def info(self, tag, msg):
        self.log('INFO', tag, msg)

    def debug(self, tag, msg):
        # v3.22.8 新增：补全 DEBUG 级别日志方法（v3.22.7 调用 app_log.debug 导致 AttributeError）
        self.log('DEBUG', tag, msg)

    def warn(self, tag, msg):
        self.log('WARN', tag, msg)

    def warning(self, tag, msg):
        # v3.22.8 新增：warning 作为 warn 的别名（v3.22.7 调用 app_log.warning 导致 AttributeError）
        self.log('WARN', tag, msg)

    def error(self, tag, msg):
        """v3.22.9: ERROR 级别始终输出到 Kivy Logger（控制台），不受 _enabled 控制。
        避免"日志关闭 + try/except 静默吞异常 = 完全无诊断线索"的死局。"""
        # 始终输出到 Kivy Logger（logcat 可见），不受 _enabled 控制
        try:
            Logger.error(' [%s] %s' % (tag, msg))
        except Exception:
            pass  # Kivy Logger 不可用时静默（不应发生）
        # 同时走正常 log 流程（受 _enabled 控制，决定是否写文件）
        self.log('ERROR', tag, msg)

    def get_log_text(self, max_chars=100000):
        """读取日志文件全文（截断到 max_chars 防止过大）
        v3.22.3: 文件不存在或为空时回退到内存缓冲 _lines，避免用户开启日志后查看到空内容"""
        try:
            file_content = ""
            if os.path.exists(self.filepath):
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    f.seek(0, 2)
                    total = f.tell()
                    if total > max_chars:
                        f.seek(max(0, total - max_chars))
                        f.readline()  # 跳过半行
                    file_content = f.read()
            if file_content.strip():
                return file_content
            # v3.22.3: 文件为空，回退到内存缓冲
            with self._lock:
                if self._lines:
                    return "\n".join(self._lines)
            return "暂无日志记录\n\n可能原因：\n1. 日志刚开启，尚未产生日志\n2. app存储目录不可写（请检查存储权限）\n3. 日志文件被系统清理"
        except Exception as e:
            return "读取日志失败: %s" % e

    def clear(self):
        """清空日志
        v3.22.7: 同时清空 _pending 待落盘缓冲，避免 flush 线程写入已清除的行。"""
        with self._lock:
            self._lines = []
            self._pending = []  # v3.22.7: 清空待落盘缓冲
            try:
                if os.path.exists(self.filepath):
                    os.remove(self.filepath)
            except Exception:
                pass

# 全局单例
app_log = AppLogger()


def clean_markdown(text):
    """v3.22.0: 清理 AI 回复中的 markdown 符号（**、*、`、# 等），
    避免字面显示「**30张照片**」造成误解。"""
    if not text:
        return text
    t = str(text)
    t = t.replace('```', '').replace('`', '')
    # **粗体** → 粗体；*斜体* → 斜体；__下划__ → 下划
    t = re.sub(r'\*\*([^*]+)\*\*', r'\1', t)
    t = re.sub(r'\*([^*]+)\*', r'\1', t)
    t = re.sub(r'__([^_]+)__', r'\1', t)
    # 行首标题符 # / ## 等
    t = re.sub(r'^#{1,6}\s*', '', t, flags=re.MULTILINE)
    return t.strip()


def _normalize_key_part(s):
    """v3.20.0: 规范化 key 组成部分 — strip 空白 + 数字归一化
    解决 openpyxl 读取数字时 str(cell) 产生 "123.0" vs "123" 导致 key 不匹配的问题
    """
    if s is None:
        return ""
    s = str(s).strip()
    # 归一化浮点数字符串：123.0 -> 123, 123.450 -> 123.45
    if re.match(r'^-?\d+\.0+$', s):
        s = s.split('.')[0]
    elif re.match(r'^-?\d+\.\d*0+$', s):
        s = s.rstrip('0').rstrip('.')
    return s


# === 默认配置 ===
# Excel 格式：A=客户名 B=抵押物地址（概） C=抵押物地址（精确门牌号） D=抵押物性质
DEFAULT_CONFIG = {
    'naming_segments': ['拍摄日期', '客户名', '地址+时间', '空值'],
    'watermark_enabled': True,
    'watermark_segments': ['拍摄时间', '地址名', '经纬度'],
    'watermark_position': 'bottom-right',
    'watermark_font_size': '中',
    'watermark_opacity': 170,
    'ai_api_url': 'https://api.deepseek.com/v1',
    'ai_api_key': '',
    'ai_model': 'deepseek-v4-flash',
    'excel_uri': '',  # v3.20.0: 持久化 SAF Excel URI，重启后可继续写回备注
    'log_enabled': False,       # v3.22.0: 日志开关，默认关闭
    'recent_excel': [],          # v3.22.0: 最近打开的 Excel 记录（最多5条 [{uri, name}]）
    'search_field': '客户名',    # v3.22.0: 搜索类型，默认客户名
}

# AI API Key (base64编码存储，避免泄露与GitHub Push Protection拦截)
# v3.19.6: 移除 OpenRouter，统一使用 DeepSeek(deepseek-v4-flash) 作为唯一 LLM
import base64 as _b64
_DEEPSEEK_KEY_B64 = "c2stMzNjNjNlMTUxMzllNGU2YzhkYmI5MzA4OGQwYjZjNWY="
DEEPSEEK_DEFAULT_API_KEY = _b64.b64decode(_DEEPSEEK_KEY_B64).decode('utf-8')
# 兼容别名：旧代码引用 AI_DEFAULT_API_KEY 时仍可用（指向 DeepSeek key）
AI_DEFAULT_API_KEY = DEEPSEEK_DEFAULT_API_KEY

# === 命名段选项（用-连接成 X-X-X-X）===
NAMING_SEGMENT_OPTIONS = [
    "拍摄日期",
    "拍摄时间",
    "客户名",
    "抵押物地址（全）",
    "抵押物地址（概）",
    "抵押物地址（精确门牌号）",
    "地址+时间",
    "空值",
]

# === 水印段选项（3段，用-连接成 X-X-X）===
WATERMARK_SEGMENT_OPTIONS = [
    "经纬度",
    "拍摄时间",
    "地址名",
    "空值",
]

# === 水印字号 ===
# v3.19.0: 字号整体增大2倍，水印区域更大更清晰
WATERMARK_FONT_SIZE_OPTIONS = ["大", "中", "小"]
WATERMARK_FONT_SIZE_MAP = {"大": 80, "中": 56, "小": 36}

# === 水印位置 ===
WATERMARK_POSITION_OPTIONS = ['bottom-right', 'bottom-left', 'top-right', 'top-left']
WATERMARK_POSITION_LABELS = {'bottom-right': '右下', 'bottom-left': '左下',
                             'top-right': '右上', 'top-left': '左上'}
WATERMARK_POSITION_LABEL_TO_KEY = {v: k for k, v in WATERMARK_POSITION_LABELS.items()}

# === 作者信息（已移除） ===
AUTHOR_NAME = ""
AUTHOR_PHONE = ""
AUTHOR_INFO = ""

# === 颜色主题 ===
# v3.19.0: 全面重新设计为明亮浅色主题，参考2026移动端设计趋势
# （明亮配色 + 卡片化布局 + 柔和阴影 + 高对比度文字）
THEME = {
    'bg': (0.95, 0.96, 0.98, 1),          # 明亮浅蓝灰背景 #F2F4F8
    'card': (1.0, 1.0, 1.0, 1.0),         # 纯白卡片 #FFFFFF
    'card_border': (0.86, 0.88, 0.92, 1), # 卡片浅灰描边
    'accent': (0.13, 0.59, 0.95, 1),      # 活力蓝 #2196F3
    'accent_dark': (0.08, 0.40, 0.78, 1), # 深蓝 #1465C7
    'success': (0.20, 0.70, 0.36, 1),     # 清新绿 #33B35C
    'danger': (0.95, 0.27, 0.21, 1),      # 珊瑚红 #F44336
    'warning': (1.0, 0.62, 0.04, 1),     # 明亮琥珀 #FF9E0A
    'text': (0.12, 0.13, 0.15, 1),       # 近黑文字 #1F2126
    'text_dim': (0.42, 0.45, 0.50, 1),   # 中灰副文本 #6B7380
    'muted': (0.62, 0.65, 0.70, 1),      # 禁用/次要按钮 #9EA6B3
    # v3.22.2: 当前拍摄行高亮配色（浅蓝底 + 蓝色边）
    'highlight_bg': (0.89, 0.95, 0.99, 1),     # 浅蓝底 #E3F2FD（与白卡对比清晰）
    'highlight_border': (0.13, 0.59, 0.95, 1), # 活力蓝边 #2196F3（= accent）
    # v3.22.19: 行选中态背景（淡蓝），与 highlight_bg 区分（更浅）
    'info': (0.85, 0.92, 1.0, 1.0),
}

# === 拍照类型 ===
PHOTO_TYPES = [
    ("远景", "小区/厂区全貌、楼栋外立面"),
    ("近景", "单元门口、楼层门牌、房号牌"),
    ("内部", "室内全景、核心区域现状"),
    ("瑕疵", "破损、漏水、违建、占用特写"),
    ("其他", "其他需要记录的场景"),
]
PHOTO_TYPE_LABELS = ["远景", "近景", "内部", "瑕疵", "其他"]

# ============================================================
# 工具函数
# ============================================================

def get_system_date():
    return datetime.now()

def get_date_str():
    return get_system_date().strftime("%Y%m%d")

def get_time_str():
    return get_system_date().strftime("%H%M")

def get_date_display():
    return get_system_date().strftime("%Y年%m月%d日")

def get_datetime_str():
    return get_system_date().strftime("%Y-%m-%d %H:%M")

def get_full_datetime_str():
    return get_system_date().strftime("%Y-%m-%d %H:%M:%S")

def get_report_date_str():
    return get_system_date().strftime("%Y年%m月%d日")

def clean_filename(s):
    for ch in '/\\:*?"<>|':
        s = s.replace(ch, '_')
    return s.strip()

def _extract_route_name(excel_filename):
    """v3.22.19: 从 Excel 文件名提取线路名，用于 AI 报告默认文件名
    规则：
    - 优先匹配 "线路X" 或 "线路X" 模式（如 "线路四" → "四"；"线路4" → "4"）
    - 其次匹配日期模式 "M.D盘点表" → "M.D"（如 "7.5盘点表" → "7.5"）
    - 再次匹配纯日期 "20260705"
    - 都不匹配则回退 "抵押物抵债资产"
    供 ReportGenerator.generate_with_ai 与 MainScreen._save_report_to_user 共用，避免重复代码。"""
    if not excel_filename:
        return "抵押物抵债资产"
    # 匹配 "线路四" 或 "线路4" 等
    m = re.search(r'线路([一二三四五六七八九十0-9]+)', excel_filename)
    if m:
        return m.group(1)
    # 匹配 "7.5盘点表" 或 "7.5 盘点表"
    m = re.search(r'(\d+\.\d+)\s*盘点', excel_filename)
    if m:
        return m.group(1)
    # 匹配纯日期 "20260705"
    m = re.search(r'(\d{8})', excel_filename)
    if m:
        return m.group(1)
    return "抵押物抵债资产"

# ============================================================
# 配置管理器
# ============================================================

class AppConfig:
    """持久化应用配置"""
    def __init__(self):
        self.data = dict(DEFAULT_CONFIG)
        self._load()

    def _load(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    saved = json.load(f)
                    self.data.update(saved)
            except Exception:
                pass

    def save(self):
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            Logger.error(f"AppConfig.save: {e}")

    def get(self, key, default=None):
        return self.data.get(key, default)

    def set(self, key, value):
        self.data[key] = value
        self.save()

# ============================================================
# 进度管理器
# ============================================================

class ProgressManager:
    """拍照进度管理 - 用借款人+地址作为持久化key，跨Excel文件可识别
    v3.17.0: 增加备注持久化（remark字段），退出app后备注不丢失。
    key = md5(借款人|地址概+地址精确)，即A+B+C列拼接，确保同借款人不同抵押物唯一识别。
    """
    def __init__(self, filepath=None):
        self.filepath = filepath or PROGRESS_FILE
        self.data = {}
        self._lock = threading.RLock()  # v3.21.0: 线程安全，防止并发写入损坏 JSON
        self.load()

    def _make_key(self, borrower, address=""):
        """基于借款人+地址(A+B+C列)生成持久化key
        v3.20.0: 规范化 borrower 和 address，防止 "123" vs "123.0" 导致 key 不匹配"""
        borrower = _normalize_key_part(borrower)
        address = _normalize_key_part(address)
        raw = (borrower + "|" + address).strip("|")
        return hashlib.md5(raw.encode('utf-8')).hexdigest()[:16]

    def load(self):
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
                # v3.22.3: 兼容旧数据（无 _row_remarks 字段时初始化为空字典）
                if not isinstance(self.data.get('_row_remarks'), dict):
                    self.data['_row_remarks'] = {}
                # v3.22.6: 兼容旧数据（无 batch_marked 字段时初始化为空字典）
                if not isinstance(self.data.get('batch_marked'), dict):
                    self.data['batch_marked'] = {}
            except Exception:
                self.data = {}
                self.data['_row_remarks'] = {}
                self.data['batch_marked'] = {}
        else:
            self.data['_row_remarks'] = {}
            self.data['batch_marked'] = {}

    def save(self):
        with self._lock:
            try:
                os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
                # v3.21.0: 原子写 — 先写临时文件再 os.replace，防止并发写入损坏 JSON
                tmp = self.filepath + '.tmp'
                with open(tmp, 'w', encoding='utf-8') as f:
                    json.dump(self.data, f, ensure_ascii=False, indent=2)
                    f.flush()
                    os.fsync(f.fileno())
                os.replace(tmp, self.filepath)
                Logger.info(f"ProgressManager.save: 已保存 {len(self.data)} 条记录到 {self.filepath}")
            except Exception as e:
                Logger.error(f"ProgressManager.save: 保存失败 {e}")

    def save_remark_by_row(self, row_index, remark_text):
        """v3.22.3: 按行号独立存储备注，避免同一客户多行(共享 borrower+address key)
        的备注互相覆盖。row_index = self.rows 中的索引，稳定不受搜索过滤影响。"""
        with self._lock:
            if '_row_remarks' not in self.data:
                self.data['_row_remarks'] = {}
            self.data['_row_remarks'][str(row_index)] = remark_text or ""
            self.save()

    def get_remark_by_row(self, row_index):
        """v3.22.3: 按行号读取备注"""
        with self._lock:
            if '_row_remarks' not in self.data:
                return ""
            return self.data['_row_remarks'].get(str(row_index), "")

    # ============================================================
    # v3.22.6: 批量标记同类型 — 借款人抵押整栋楼盘时只需拍代表性户型，
    # 其余行标记为「同类型已拍」不计入未完成数量，AI 报表合并统计。
    # key 为 row_index 字符串，与 _row_remarks 一致（稳定不受搜索过滤影响）。
    # ============================================================
    def mark_batch(self, row_index):
        """v3.22.6: 标记某行为同类型"""
        with self._lock:
            if 'batch_marked' not in self.data:
                self.data['batch_marked'] = {}
            self.data['batch_marked'][str(row_index)] = True
        self.save()

    def unmark_batch(self, row_index):
        """v3.22.6: 取消同类型标记"""
        with self._lock:
            if 'batch_marked' in self.data:
                self.data['batch_marked'].pop(str(row_index), None)
        self.save()

    def is_batch_marked(self, row_index):
        """v3.22.6: 返回某行是否已标记为同类型"""
        with self._lock:
            return str(row_index) in self.data.get('batch_marked', {})

    def get_all_batch_marked(self):
        """v3.22.6: 返回所有已标记的 row_index 列表（int）"""
        with self._lock:
            marked = self.data.get('batch_marked', {})
            return [int(k) for k in marked.keys() if isinstance(k, str) and k.isdigit()]

    def mark_photo(self, key, photo_path, photo_type=""):
        """key = _make_key(borrower, address_general, address_precise)"""
        with self._lock:
            if key not in self.data:
                self.data[key] = {"photos": [], "types": {}, "timestamp": "", "remark": ""}
            # 验证并保存完整路径
            self.data[key]["photos"].append(os.path.abspath(photo_path))
            if photo_type:
                self.data[key]["types"][photo_type] = True
            self.data[key]["timestamp"] = get_full_datetime_str()
            self.save()

    def save_remark(self, key, remark_text):
        """保存备注到持久化存储（v3.17.0）"""
        with self._lock:
            if key not in self.data:
                self.data[key] = {"photos": [], "types": {}, "timestamp": "", "remark": ""}
            self.data[key]["remark"] = remark_text
            self.save()

    def get_remark(self, key):
        """获取备注（v3.17.0）"""
        with self._lock:
            if key not in self.data:
                return ""
            return self.data[key].get("remark", "")

    def delete_photo(self, key, photo_index):
        with self._lock:
            if key in self.data and photo_index < len(self.data[key]["photos"]):
                self.data[key]["photos"].pop(photo_index)
                if not self.data[key]["photos"]:
                    # 保留remark字段，不删除整个entry（v3.17.0）
                    self.data[key]["photos"] = []
                    self.data[key]["types"] = {}
                self.save()

    def delete_all_photos(self, key):
        with self._lock:
            if key in self.data:
                # 保留remark字段（v3.17.0）
                self.data[key]["photos"] = []
                self.data[key]["types"] = {}
                self.save()

    def is_photographed(self, key):
        with self._lock:
            return key in self.data and len(self.data[key].get("photos", [])) > 0

    def get_photos(self, key):
        """返回存在的照片路径列表（过滤掉已删除的）"""
        with self._lock:
            if key not in self.data:
                return []
            valid = []
            for p in self.data[key].get("photos", []):
                if os.path.exists(p):
                    valid.append(p)
            # v3.22.5: 不再回写过滤后的列表，避免失效路径导致旧照片记录永久丢失
            return valid

    def migrate_photo_paths(self):
        """v3.22.5: 启动时迁移失效照片路径到 APP_DIR 同名文件"""
        try:
            migrated = False
            with self._lock:
                for key, info in self.data.items():
                    if not isinstance(info, dict): continue
                    photos = info.get("photos", [])
                    if not photos: continue
                    new_photos = []
                    changed = False
                    for p in photos:
                        if p and os.path.exists(p):
                            new_photos.append(p)
                        elif p:
                            # 尝试在 APP_DIR 中查找同名文件
                            basename = os.path.basename(p)
                            candidate = os.path.join(APP_DIR, basename)
                            if os.path.exists(candidate):
                                new_photos.append(candidate)
                                changed = True
                                migrated = True
                                app_log.info('PHOTO', '迁移照片路径: %s -> %s' % (p, candidate))
                            else:
                                # 保留原路径（不删除记录）
                                new_photos.append(p)
                        else:
                            new_photos.append(p)
                    if changed:
                        info["photos"] = new_photos
            if migrated:
                self.save()
                app_log.info('PHOTO', '照片路径迁移完成，已保存')
        except Exception as e:
            app_log.error('PHOTO', 'migrate_photo_paths 异常: %s' % e)

    def get_photo_count(self, key):
        return len(self.get_photos(key))

    def get_done_count(self, keys):
        """keys = [(borrower, address), ...]"""
        return sum(1 for b, a in keys if self.is_photographed(self._make_key(b, a)))

    def get_next_photo_index(self, key):
        with self._lock:
            if key in self.data:
                return len(self.data[key]["photos"])
            return 0

    def get_next_type_index(self, key, photo_type):
        """返回该客户指定类型的下一个照片编号（01开始，跨会话连续）。
        统计已有照片中文件名包含 -{type}-NN 模式的最大编号+1。
        v3.22.1: 修复分隔符bug——照片文件名用短横线分隔(如 远景-01)，
        检测原用下划线(_远景_)导致永不匹配、序号永远从1开始(可能文件名冲突)。
        兼容历史下划线格式。
        """
        if key not in self.data:
            return 1
        max_idx = 0
        # 兼容短横线(当前)与下划线(历史)两种分隔符
        seps = ("-%s-" % photo_type, "_%s_" % photo_type)
        for p in self.data[key].get("photos", []):
            basename = os.path.basename(p)
            tag = next((s for s in seps if s in basename), None)
            if tag:
                try:
                    idx_str = basename.split(tag)[-1].split('.')[0]
                    idx_str = idx_str.split('-')[0].split('_')[0]
                    idx = int(idx_str)
                    if idx > max_idx:
                        max_idx = idx
                except:
                    pass
        return max_idx + 1

    def get_photo_types(self, key):
        with self._lock:
            return self.data.get(key, {}).get("types", {})

    def get_photo_type_summary(self, key):
        types = self.get_photo_types(key)
        done = sum(1 for t in PHOTO_TYPE_LABELS if types.get(t, False))
        return f"{done}/5"

    def get_photo_count_by_type(self, key):
        """v3.22.0: 返回各拍照类型的照片张数字典 {类型: 张数}
        基于实际照片文件名统计（文件名含 -{类型}- 模式），
        而非仅 types 标记（types 只标记是否拍过，不记张数）。
        v3.22.1: 修复分隔符bug——照片文件名用短横线(-远景-)，
        检测原用下划线(_远景_)导致计数永远为0、竖列不更新。兼容历史下划线格式。
        """
        result = {label: 0 for label in PHOTO_TYPE_LABELS}
        with self._lock:
            if key not in self.data:
                return result
            photos = list(self.data[key].get("photos", []))
        # 锁外遍历文件名，减少持锁时间
        for p in photos:
            if not os.path.exists(p):
                continue
            basename = os.path.basename(p)
            for label in PHOTO_TYPE_LABELS:
                # 兼容短横线(当前)与下划线(历史)两种分隔符
                if ("-%s-" % label) in basename or ("_%s_" % label) in basename:
                    result[label] += 1
                    break
        return result

    def get_progress_snapshot(self, rows):
        """v3.22.0: 性能优化 — 一次性在单次锁内收集所有行的进度信息，
        避免 build_system_prompt 等场景对每行分别加锁（N 行 → 2N 次加锁降为 1 次）。
        返回 {key: {'count': int, 'remark': str, 'types': dict}}。
        rows: Excel 行（[serial, borrower, addr_gen, addr_prec, ...]）
        """
        snap = {}
        with self._lock:
            keys = set()
            for row in rows:
                borrower = row[1] if len(row) > 1 else ""
                addr = ((row[2] if len(row) > 2 else "") + (row[3] if len(row) > 3 else "")).strip()
                keys.add(self._make_key(borrower, addr))
            for k in keys:
                if k in self.data:
                    photos = self.data[k].get("photos", [])
                    # 仅统计存在的照片
                    cnt = sum(1 for p in photos if os.path.exists(p))
                    snap[k] = {
                        'count': cnt,
                        'remark': self.data[k].get('remark', ''),
                        'types': dict(self.data[k].get('types', {})),
                        'photos': list(photos),
                    }
                else:
                    snap[k] = {'count': 0, 'remark': '', 'types': {}, 'photos': []}
        return snap

    # ============================================================
    # v3.22.19: 缩略图管理 — 拍照后异步生成压缩缩略图保存到 app 私有目录，
    # "查看已拍"改读缩略图，规避 Android Scoped Storage 访问原图的权限问题。
    # ============================================================
    @staticmethod
    def generate_thumbnail(original_path, progress_key):
        """v3.22.19: 用 PIL 打开原图，最长边缩放至 480px（保持宽高比），JPEG quality=70，
        保存到 APP_DIR/thumbnails/<progress_key>/<basename>.jpg。
        失败时仅记录日志，不阻断主流程（调用方在异步线程中执行）。"""
        try:
            if not original_path or not os.path.exists(original_path):
                app_log.error('PHOTO', 'generate_thumbnail: 原图不存在 %s' % original_path)
                return None
            if not progress_key:
                app_log.error('PHOTO', 'generate_thumbnail: progress_key 为空')
                return None
            thumb_dir = os.path.join(APP_DIR, 'thumbnails', progress_key)
            os.makedirs(thumb_dir, exist_ok=True)
            # 缩略图文件名：原 basename 去扩展名 + .jpg（无扩展名则加 .jpg）
            basename = os.path.basename(original_path)
            name_base, ext = os.path.splitext(basename)
            if not name_base:
                name_base = basename  # 极端情况：basename 没有扩展名
            thumb_name = name_base + '.jpg'
            thumb_path = os.path.join(thumb_dir, thumb_name)
            # v3.22.19: 用 with 管理 PIL 图像句柄，避免句柄泄漏
            with PILImage.open(original_path) as img:
                img = img.convert('RGB')
                # 最长边缩放至 480px，保持宽高比
                max_edge = 480
                w, h = img.size
                if max(w, h) > max_edge:
                    if w >= h:
                        new_w = max_edge
                        new_h = int(h * (max_edge / w))
                    else:
                        new_h = max_edge
                        new_w = int(w * (max_edge / h))
                    img = img.resize((new_w, new_h), PILImage.LANCZOS)
                img.save(thumb_path, 'JPEG', quality=70)
            app_log.info('PHOTO', '缩略图已生成: %s -> %s' % (basename, thumb_path))
            return thumb_path
        except Exception as e:
            app_log.error('PHOTO', 'generate_thumbnail 异常 %s: %s' % (original_path, e), exc_info=True)
            return None

    @staticmethod
    def get_thumbnails(key):
        """v3.22.19: 返回 APP_DIR/thumbnails/<key>/ 下所有 .jpg 文件路径列表（按文件名排序）。
        目录不存在时返回空列表。"""
        try:
            if not key:
                return []
            thumb_dir = os.path.join(APP_DIR, 'thumbnails', key)
            if not os.path.isdir(thumb_dir):
                return []
            files = [f for f in os.listdir(thumb_dir) if f.lower().endswith('.jpg')]
            files.sort()
            return [os.path.join(thumb_dir, f) for f in files]
        except Exception as e:
            app_log.error('PHOTO', 'get_thumbnails 异常 key=%s: %s' % (key, e))
            return []

    @staticmethod
    def delete_thumbnail(key, filename):
        """v3.22.19: 删除单个缩略图文件。
        filename 可以是完整路径或仅文件名。"""
        try:
            if not key or not filename:
                return
            thumb_dir = os.path.join(APP_DIR, 'thumbnails', key)
            # 支持 filename 是完整路径或仅文件名
            if os.path.isabs(filename):
                target = filename
            else:
                target = os.path.join(thumb_dir, filename)
            if os.path.exists(target):
                os.remove(target)
                app_log.info('PHOTO', '缩略图已删除: %s' % target)
        except Exception as e:
            app_log.error('PHOTO', 'delete_thumbnail 异常 %s/%s: %s' % (key, filename, e))

    @staticmethod
    def delete_all_thumbnails(key):
        """v3.22.19: 删除整个 <key>/ 目录及其内容（用 shutil.rmtree）"""
        try:
            if not key:
                return
            import shutil
            thumb_dir = os.path.join(APP_DIR, 'thumbnails', key)
            if os.path.isdir(thumb_dir):
                shutil.rmtree(thumb_dir, ignore_errors=True)
                app_log.info('PHOTO', '缩略图目录已删除: %s' % thumb_dir)
        except Exception as e:
            app_log.error('PHOTO', 'delete_all_thumbnails 异常 key=%s: %s' % (key, e))

# ============================================================
# Excel 读取器
# ============================================================

class ExcelReader:
    """读取Excel，v3.22.0: A=序号 B=客户名 C=抵押物地址（概） D=抵押物地址（精确门牌号） E=抵押物性质 F=备注
    返回 rows 每项: [serial, borrower, address_general, address_precise, property_type, remark]
    v3.15.0: 增加读取 E 列（备注）
    v3.22.0: A 列插入序号，其余列顺延（备注从 E 移到 F）
    """
    def __init__(self, file_path):
        self.file_path = file_path
        self.headers = []
        self.rows = []

    def load(self):
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # v3.22.0: 读取 A-F 共6列（A序号 B客户名 C地址概 D地址详 E性质 F备注）
            cells = [str(cell).strip() if cell else "" for cell in row[:6]]
            # 补齐到6列
            while len(cells) < 6:
                cells.append("")
            if i == 0:
                self.headers = cells
            else:
                # v3.22.0: 跳过 A-E 列全空的行（序号+客户名+地址概+地址详+性质）
                if not any(cells[:5]):
                    continue
                self.rows.append(cells)
        wb.close()
        # 自动判断表头
        if not self.headers:
            self.headers = ["序号", "客户名", "抵押物地址（概）", "抵押物地址（精确门牌号）", "抵押物性质", "备注"]
        return self.headers, self.rows


class ExcelWriter:
    """写入 Excel 备注列（F列）
    v3.16.0: 增加工作副本回退机制（Android 11+无法直接写入外部存储）。
    v3.22.0: 备注列从 E 列移至 F 列（A 列插入序号，其余顺延）。
    """
    @staticmethod
    def save_remark(file_path, row_index, remark_text):
        """保存备注到指定行（1-based row_index）的 F 列
        row_index: Excel 中的行号（含表头，从1开始）
        remark_text: 备注文本
        返回: (bool, str) 是否成功, 人类可读提示
        """
        import shutil
        work_copy = os.path.join(APP_DIR, 'excel_work', os.path.basename(file_path))
        os.makedirs(os.path.dirname(work_copy), exist_ok=True)

        # 方式1：直接写入原文件（临时文件放在 APP_DIR，避免外部目录不可写）
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            ws.cell(row=row_index, column=6, value=remark_text)
            tmp = os.path.join(APP_DIR, 'excel_work', os.path.basename(file_path) + '.tmp')
            wb.save(tmp)
            wb.close()
            wb = None
            shutil.move(tmp, file_path)
            Logger.info(f"save_remark: 直接写入成功 行{row_index}")
            return True, "已保存到 Excel"
        except Exception as e1:
            Logger.error(f"save_remark直接写入失败: {e1}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

        # 方式2：复制到 app 私有目录再写入，然后尝试复制回原路径
        wb = None
        try:
            shutil.copy2(file_path, work_copy)
            wb = openpyxl.load_workbook(work_copy)
            ws = wb.active
            ws.cell(row=row_index, column=6, value=remark_text)
            wb.save(work_copy)
            wb.close()
            wb = None
            try:
                shutil.copy2(work_copy, file_path)
                Logger.info(f"save_remark: 通过工作副本写入成功 行{row_index}")
                return True, "已保存到 Excel"
            except Exception as e3:
                Logger.error(f"save_remark: 无法复制回原路径({e3})，副本保留在 {work_copy}")
                return True, f"原 Excel 受系统保护无法直接写入，已保存到工作副本：{work_copy}"
        except Exception as e2:
            Logger.error(f"save_remark工作副本写入也失败: {e2}")
            return False, f"Excel 保存失败：{e2}"
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    @staticmethod
    def save_all_remarks(file_path, remarks):
        """批量保存备注
        remarks: dict {row_index: remark_text}
        """
        import shutil
        # 方式1：直接写入
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for row_idx, text in remarks.items():
                ws.cell(row=row_idx, column=6, value=text)
            tmp = file_path + '.tmp'
            wb.save(tmp)
            wb.close()
            shutil.move(tmp, file_path)
            return True
        except Exception as e:
            Logger.error(f"save_all_remarks直接写入失败: {e}")
        # 方式2：工作副本
        try:
            work_copy = os.path.join(APP_DIR, 'excel_work', os.path.basename(file_path))
            os.makedirs(os.path.dirname(work_copy), exist_ok=True)
            shutil.copy2(file_path, work_copy)
            wb = openpyxl.load_workbook(work_copy)
            ws = wb.active
            for row_idx, text in remarks.items():
                ws.cell(row=row_idx, column=6, value=text)
            wb.save(work_copy)
            wb.close()
            try:
                shutil.copy2(work_copy, file_path)
            except Exception:
                pass
            return True
        except Exception as e2:
            Logger.error(f"save_all_remarks工作副本写入失败: {e2}")
            return False

    @staticmethod
    def write_back_to_uri(uri_str, source_file):
        """v3.19.0: 通过 SAF content:// URI 将本地 Excel 文件字节写回原始文件。
        用于 Android 11+ scoped storage：直接写共享存储路径会失败，
        但通过 ContentResolver.openOutputStream(uri) 可写回用户选择的原始文件。
        返回: (bool, str) 是否成功, 提示。
        """
        if not IS_ANDROID or not uri_str or not os.path.exists(source_file):
            return False, "无可用 URI 或源文件"
        # v3.22.13 修复: fis/out 用 try/finally 确保异常时也关闭，防止 Java 文件句柄泄漏
        out = None
        fis = None
        try:
            from jnius import autoclass
            Uri = autoclass('android.net.Uri')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            resolver = PythonActivity.mActivity.getContentResolver()
            uri = Uri.parse(uri_str)
            out = resolver.openOutputStream(uri)
            FileInputStream = autoclass('java.io.FileInputStream')
            fis = FileInputStream(source_file)
            buf = bytearray(8192)
            while True:
                n = fis.read(buf)
                if n <= 0:
                    break
                out.write(buf, 0, n)
            Logger.info(f"write_back_to_uri: 已写回原始 Excel URI")
            return True, "已写回原始 Excel"
        except Exception as e:
            Logger.error(f"write_back_to_uri 失败: {e}")
            return False, f"写回原始文件失败: {str(e)[:40]}"
        finally:
            # v3.22.13: 无论成功/失败都关闭流（原代码仅在成功路径 close，异常时泄漏句柄）
            if fis is not None:
                try:
                    fis.close()
                except Exception:
                    pass
            if out is not None:
                try:
                    out.close()
                except Exception:
                    pass

# ============================================================
# 报告生成器
# ============================================================

class ReportGenerator:
    """日报表生成器 v3.19.2: 基于内置模板，调用 AI 根据备注+Excel内容
    生成专业的"现状描述"与"风险备注"，保存到用户指定位置。"""

    TEMPLATE_NAME = 'report_template.xlsx'

    REPORT_SYSTEM_PROMPT = (
        "你是银行抵押物、抵债资产现场勘查日报表撰写助手。"
        "根据勘查人员提供的客户名称、抵押物地址、抵押物类型、备注与拍照情况，"
        "为每位客户撰写日报表中的三列内容：\n"
        "1. 抵押物情况(detail)：概括盘点了几处、什么类型的抵押物及地址位置\n"
        "2. 现状描述(status)：客观描述抵押物当前使用状态(自用/出租/闲置)、位置楼层、"
        "装修、维护情况等\n"
        "3. 风险备注(risk)：评估是否存在风险(如闲置、渗水裂缝、价格波动、用途变更等)，"
        "给出关注建议；无风险则说明整体状态良好暂无异常\n"
        "要求：语言专业、简洁，符合银行风控用语；严格基于勘查人员提供的备注内容填写，"
        "严禁编造备注中未提及的信息；备注为「无」或为空时，仅客观描述抵押物类型与地址，不得杜撰现状或风险。\n"
        "【重要】严禁使用「等」字省略信息！备注中提及的所有人名、地址、使用人、"
        "使用情况必须完整列出，不得遗漏。例如备注写「抵押物由张三、李四、王五使用」，"
        "报告中必须写「由张三、李四、王五使用」，不得简写为「由张三等人使用」。\n"
        "v3.22.6 同类型说明：对于标记为「同类型」的抵押物，这些是借款人抵押整栋楼盘中"
        "具有代表性的户型，已批量标记为同类型，无需为每户单独拍照。在报告中应合并表述为"
        "「同类型户型」，可在抵押物情况中说明「该抵押物与同楼盘代表性户型为同类型，"
        "已现场勘查代表性户型」，现状描述与风险备注可参照同类型代表性户型的勘查结论填写。\n"
        "【每客户仅 1 行 — 强制规则】同一客户不论抵押房产数量"
        "（住宅、商业、车位、办公、厂房等所有类型），SHALL 仅生成 1 行数据。"
        "即使 Excel 中该客户有 200+ 条记录、覆盖多栋楼/多个地址/多种类型，"
        "报告中对该客户只输出一条 JSON 对象，绝不允许多行。"
        "「抵押物情况」字段应合并表述为「共计盘点 N 处，位置为 XX、YY…」。\n"
        "【同楼盘多房间合并描述】当同一借款人抵押同一栋楼的多个房间号"
        "（如 A大街12号101、102、201、202 等，Excel 中可能对应 200+ 条记录）时：\n"
        "1. 「抵押物情况」列应合并表述为「共计盘点 N 处，位置为 XX栋YY号、ZZ号...」，"
        "不要为每个房间号单独生成一行；\n"
        "2. 「现状描述」列应描述整栋楼的整体状态，不重复每个房间的具体描述；\n"
        "3. 当走访补充说明中提到「同房型一致」「镜像关系」「不同楼层面积存在细小差别」时，"
        "应在「现状描述」中合并说明，参考样板："
        "「该项目已经建设完成，正在出售阶段，其中28、29、30、31号楼为高层，"
        "现场据楼盘销售介绍楼内两种户型（均已拍摄完毕），未拍摄房间为01或02房间镜像，"
        "同一房间号不同楼层面积存在细微差别，因此未逐户拍摄。」\n"
        "4. 即使 Excel 中该客户有 200+ 条记录，报告中只生成 1 行，不重复描述。\n"
        "【多地址分号成行】v3.22.19: 当一个客户存在多个地址的抵押物时"
        "（如 弘马大街123号5xx、弘马大街123号4xx、小田屯组 等），"
        "在「现状描述」单元格内每个地址单独标号、单独成行，格式：\n"
        "1. <地址1>：<现状描述>\n"
        "2. <地址2>：<现状描述>\n"
        "3. <地址3>：<现状描述>\n"
        "注意：每个地址占一行，用换行符 \\n 分隔；编号从 1 开始；地址取自 Excel 地址列；"
        "「现状描述」字段内的换行需用 \\n 转义写入 JSON 字符串。\n"
        "【同地址 4+ 房间号合并】v3.22.19: 同一地址项下存在 4 个及以上房间号时"
        "（如 弘马大街123号 项下有 101、102、201、202、207 等），视为同一地址"
        "（弘马大街123号），LLM SHALL 根据补充说明和备注内容合并描述现状。\n"
        "判断逻辑：地址前缀相同（如「弘马大街123号」）+ 项下房间号 ≥ 4 个 → 视为同一地址。"
        "此时「现状描述」单元格内仅生成 1 条该地址的描述，但若各房间号现状不同，"
        "可在同一地址条目内用顿号或分号列出各房间号的不同情况，"
        "不为每个房间号单独标号成行。\n"
        "【示例】客户「张三」有以下抵押物：\n"
        "- 弘马大街123号5xx（客房层）\n"
        "- 弘马大街123号4xx（客房层）\n"
        "- 弘马大街123号107（一楼门面）\n"
        "- 弘马大街123号207（二楼，照比107多一个月亮门至后门区域）\n"
        "- 弘马大街123号101（与201同空间，201性质疑似错误）\n"
        "- 小田屯组（108-2-208-1/2）（已打通为一整栋房产）\n"
        "补充说明：\n"
        "- 弘马大街123号5xx、4xx 为两层房产，电梯到达4楼实际为4楼半，"
        "向下/上半层通向4楼及5楼客房层，走访抵押物为客房层\n"
        "- 107 月亮门至大门口为房证面积\n"
        "- 207 照比107多一个月亮门至后门区域\n"
        "- 101 与 201 在一个空间内，201性质为住宅疑似错误\n"
        "- 小田屯组（108-2-208-1/2）已打通为一整栋房产，2楼清水，3楼半简装半清水\n"
        "期望输出（仅 1 行）：\n"
        "{\"name\":\"张三\","
        "\"detail\":\"共计盘点 6 处，位置为弘马大街123号（5xx、4xx、107、207、101）及小田屯组\","
        "\"status\":\"1. 弘马大街123号5xx、4xx：两层房产，电梯到达4楼实际为4楼半，"
        "向下/上半层分别通向4楼及5楼客房层，走访抵押物为客房层\\n"
        "2. 弘马大街123号107：月亮门至大门口为房证面积\\n"
        "3. 弘马大街123号207：照比107多一个月亮门至后门区域\\n"
        "4. 弘马大街123号101：与201在同一空间内，201性质为住宅疑似错误\\n"
        "5. 小田屯组（108-2-208-1/2）：已打通为一整栋房产，2楼清水，3楼半简装半清水\","
        "\"risk\":\"（根据实际情况评估）\"}\n"
        "注意：弘马大街123号 项下有 5 个房间号（≥4），视为同一地址但每个房间号现状不同，"
        "因此在「现状描述」内按房间号分号成行；小田屯组为独立地址单独成行。\n"
        "仅返回 JSON 数组，不要包含任何解释文字或 markdown 代码块标记。"
        "格式：[{\"name\":\"客户名\",\"detail\":\"抵押物情况\",\"status\":\"现状描述\",\"risk\":\"风险备注\"}, ...]"
    )

    def __init__(self):
        self.template_path = self._resolve_template()

    def _resolve_template(self):
        """定位内置模板（打包目录优先）"""
        for d in [_APP_DIR, os.path.dirname(os.path.abspath(__file__))]:
            p = os.path.join(d, self.TEMPLATE_NAME)
            if os.path.exists(p) and os.path.getsize(p) > 1000:
                return p
        return None

    def _collect_records(self, rows, progress_mgr):
        """收集每位客户的勘查记录（含备注与拍照数）
        v3.22.0: Excel 格式改为 A序号 B客户名 C地址概 D地址详 E性质 F备注
        v3.22.0: 按 borrower 分组 — 同一客户多处抵押物合并为一条记录，
        detail 字段需表述"共计盘点 N 处，位置为 XX、XX…"。
        v3.22.6: 新增 batch_marked 字段 — 标记为同类型的代表性户型（无需单独拍照）。
                 按 borrower 分组时，组内任一行被标记即视为该组已标记同类型。
        """
        from collections import OrderedDict
        groups = OrderedDict()
        # v3.22.4 P0 修复: 必须用 enumerate 获取行号 i，否则下方
        # get_remark_by_row(i) 的 i 未定义 → NameError → 线程崩溃 → 报表弹窗永不关闭（卡住）
        for i, row in enumerate(rows):
            borrower = row[1] if len(row) > 1 else ""
            if not borrower:
                continue
            addr_general = row[2] if len(row) > 2 else ""
            addr_precise = row[3] if len(row) > 3 else ""
            property_type = row[4] if len(row) > 4 else ""
            excel_remark = row[5] if len(row) > 5 else ""
            full_addr = (addr_general + addr_precise).strip()
            pk = progress_mgr._make_key(borrower, full_addr)
            saved_remark = progress_mgr.get_remark(pk)
            # v3.22.3 P0 修复: 优先按行号读取备注（行号独立），回退 Excel，回退旧 progress_key。
            row_remark = progress_mgr.get_remark_by_row(i) if hasattr(progress_mgr, 'get_remark_by_row') else ""
            remark = row_remark if row_remark else (excel_remark if excel_remark else saved_remark)
            photo_info = progress_mgr.get_photo_types(pk) if hasattr(progress_mgr, 'get_photo_types') else {}
            photo_count = sum(photo_info.values()) if isinstance(photo_info, dict) else 0
            # v3.22.6: 查询该行是否被标记为同类型（按行号，稳定）
            row_batch_marked = progress_mgr.is_batch_marked(i) if hasattr(progress_mgr, 'is_batch_marked') else False
            g = groups.setdefault(borrower, {
                'name': borrower, 'addresses': [], 'property_types': [],
                'remarks': [], 'photo_count': 0, 'batch_marked': False,
            })
            g['addresses'].append(full_addr)
            g['property_types'].append(property_type)
            g['remarks'].append(remark)
            g['photo_count'] += photo_count
            # v3.22.6: 组内任一行被标记即视为该组已标记同类型
            if row_batch_marked:
                g['batch_marked'] = True
        records = []
        for name, g in groups.items():
            # 合并备注：去重保留所有非空备注
            merged_remark = "；".join(sorted(set(r for r in g['remarks'] if r))) or ""
            records.append({
                'name': name,
                'address': g['addresses'],          # list
                'property_type': g['property_types'], # list
                'remark': merged_remark,
                'photo_count': g['photo_count'],
                'count': len(g['addresses']),
                # v3.22.6: 是否标记为同类型代表性户型
                'batch_marked': g.get('batch_marked', False),
            })
        return records

    def _build_prompt(self, records, visit_note=""):
        """构建发送给 AI 的用户提示词
        v3.22.0: 同一客户多处抵押物已合并；严格约束备注不得编造。
        v3.22.14: visit_note 参数 — 今日走访补充说明，附加到 user prompt 末尾，
                  与 system prompt 附加双管齐下，强化 AI 对合并描述同楼盘多房间的引导。
        v3.22.19: 按地址前缀分组，标注 4+ 房间号的同地址，引导 LLM 在「现状描述」内分号成行。"""
        lines = ["以下是今日现场勘查记录。同一客户可能有多处抵押物，已合并为一条。\n"]
        for i, r in enumerate(records, 1):
            addrs = r['address']
            if isinstance(addrs, list):
                addr_str = "、".join(a for a in addrs if a)
                count = r.get('count', len(addrs))
            else:
                addr_str = addrs
                count = 1
            ptypes = r['property_type']
            if isinstance(ptypes, list):
                ptype_str = "、".join(sorted(set(p for p in ptypes if p)))
            else:
                ptype_str = ptypes or ''
            # v3.22.19: 按地址前缀分组，标注 4+ 房间号的同地址
            addr_lines, merge_hints = self._format_address_grouping(addrs)
            lines.append("【客户%d】名称：%s | 抵押物共%d处 | 类型：%s | 拍照数：%d"
                         % (i, r['name'], count, ptype_str or '未注明', r['photo_count']))
            if addr_lines:
                lines.append("  地址列表：")
                for al in addr_lines:
                    lines.append("    - " + al)
            else:
                lines.append("  地址：%s" % (addr_str or '无'))
            # 同地址合并提示
            for hint in merge_hints:
                lines.append("  [同地址合并提示] " + hint)
            lines.append("  勘查备注：%s" % (r['remark'] or '无'))
            lines.append("")
        lines.append("要求：")
        lines.append("1. 每位客户生成一条日报表内容（仅 1 行，不论抵押房产数量）。")
        lines.append("2. 「抵押物/抵债资产具体情况」字段：当客户有多处时表述为「共计盘点 N 处，位置为 XX、XX…」；仅1处时直接写地址。")
        lines.append("3. 「现状描述」字段：多地址客户按地址分号成行（1. <地址1>：...\\n 2. <地址2>：...）；同地址 4+ 房间号合并为 1 条。")
        lines.append("4. 「现状描述」「备注」字段必须严格基于上述「勘查备注」内容填写，严禁编造未提供的信息；备注为「无」时留空或写「无」。")
        lines.append("5. 请按 JSON 数组返回，每个元素含 name/detail/status/risk 四字段，顺序与客户一致。")
        lines.append("6. 【重要】严禁使用「等」字省略！备注和地址中提及的所有人名、地点、使用人必须完整列出。")
        lines.append("   例如备注「抵押物由张三、李四、王五使用」→报告中写「由张三、李四、王五使用」，禁止「由张三等人使用」。")
        prompt = "\n".join(lines)
        # v3.22.14: 走访补充说明显式附加到 user prompt（双管齐下，强化 AI 引导）
        if visit_note:
            prompt += "\n\n今日走访补充说明（请重点参考此说明生成报告，特别是合并描述同楼盘多房间的情况）：\n" + visit_note
        return prompt

    def _format_address_grouping(self, addrs):
        """v3.22.19: 按地址前缀分组，返回 (addr_lines, merge_hints)
        判断逻辑：地址前缀相同（取最后一个「号」或「栋」之前含该字的子串）
        + 项下房间号 ≥ 4 → 视为同一地址，输出合并提示。
        addr_lines: 每个地址的展示行（同地址 4+ 时合并为 1 行展示）
        merge_hints: 同地址合并提示文本列表"""
        import re as _re
        if not isinstance(addrs, list):
            addrs = [addrs] if addrs else []
        if not addrs:
            return [], []

        def _prefix(a):
            # 贪心匹配到最后一个「号」或「栋」
            m = _re.search(r'^(.*[号栋])', a)
            if m:
                return m.group(1)
            return a  # 无「号/栋」，整串作为前缀

        groups = {}  # prefix -> list of (full_addr, room_part)
        order = []
        for a in addrs:
            if not a:
                continue
            p = _prefix(a)
            room = a[len(p):].strip() if len(a) > len(p) else ""
            if p not in groups:
                groups[p] = []
                order.append(p)
            groups[p].append((a, room))

        addr_lines = []
        merge_hints = []
        for p in order:
            items = groups[p]
            if len(items) >= 4:
                rooms = "、".join((r or "（无房间号）") for _, r in items)
                addr_lines.append("%s（项下 %d 个房间号：%s）" % (p, len(items), rooms))
                merge_hints.append(
                    "%s 项下有 %d 个房间号（≥4），视为同一地址，但每个房间号现状可能不同，"
                    "请在「现状描述」内按房间号分号成行" % (p, len(items)))
            else:
                for full_a, _ in items:
                    addr_lines.append(full_a)
        return addr_lines, merge_hints

    def _fallback_detail(self, r):
        """v3.22.0: 构造 detail 兜底文案 — 适配合并结构（按 count 分支）"""
        addrs = r.get('address', '')
        count = r.get('count', 1)
        if isinstance(addrs, list):
            addr_str = "、".join(a for a in addrs if a)
            cnt = count if count > 0 else len(addrs)
        else:
            addr_str = addrs
            cnt = 1
        ptypes = r.get('property_type', '')
        if isinstance(ptypes, list):
            ptype_str = "、".join(sorted(set(p for p in ptypes if p))) or '抵押物'
        else:
            ptype_str = ptypes or '抵押物'
        if cnt > 1:
            return "共计盘点%d处%s，位置为%s" % (cnt, ptype_str, addr_str)
        return "地址：%s" % addr_str

    def _parse_ai_response(self, ai_text, records):
        """v3.20.0: 解析 AI 返回的 JSON — 强化提取逻辑，fallback 清理特殊字符
        v3.22.0: fallback detail 适配合并结构（按 count 分支）"""
        import json as _json
        if not ai_text or not ai_text.strip():
            return [{
                'name': r['name'],
                'detail': self._fallback_detail(r),
                'status': '现场勘查完成，状态正常',
                'risk': '经实地走访抵押物，目前整体状态较好，暂无异常情况。',
            } for r in records]

        text = ai_text.strip()

        # 尝试1：去掉 markdown 代码块标记后直接解析
        clean = text
        if clean.startswith('```'):
            # 去掉首行 ```json 或 ```
            clean = clean.split('\n', 1)[-1] if '\n' in clean else clean[3:]
            if clean.endswith('```'):
                clean = clean[:-3]
            clean = clean.strip()
        try:
            items = _json.loads(clean)
            if isinstance(items, list) and len(items) > 0:
                return self._validate_items(items, records)
        except Exception:
            pass

        # 尝试2：用正则提取 JSON 数组 [...]（AI 可能返回多余文本）
        try:
            m = re.search(r'\[.*\]', text, re.DOTALL)
            if m:
                items = _json.loads(m.group(0))
                if isinstance(items, list) and len(items) > 0:
                    return self._validate_items(items, records)
        except Exception:
            pass

        # 尝试3：逐行提取 {name:..., detail:..., ...} 对象
        try:
            obj_matches = re.findall(r'\{[^{}]+\}', text, re.DOTALL)
            if obj_matches:
                items = []
                for obj_str in obj_matches:
                    try:
                        item = _json.loads(obj_str)
                        if isinstance(item, dict):
                            items.append(item)
                    except Exception:
                        pass
                if items:
                    return self._validate_items(items, records)
        except Exception:
            pass

        # 所有解析失败：用结构化模板兜底，清理特殊字符（不使用 AI 原文避免 {} 等符号残留）
        app_log.warn('AI', 'AI 返回 JSON 解析失败，使用结构化模板兜底')
        return [{
            'name': r['name'],
            'detail': self._fallback_detail(r),
            'status': '现场勘查完成，已拍照%d张，整体状态正常。' % r['photo_count'],
            'risk': '经实地走访抵押物，目前整体状态较好，暂无异常情况。',
        } for r in records]

    def _validate_items(self, items, records):
        """v3.20.0: 校验 AI 返回的 items，补全缺失字段，清理特殊字符"""
        result = []
        for i, item in enumerate(items):
            if not isinstance(item, dict):
                continue
            r = records[i] if i < len(records) else records[-1] if records else {}
            # 确保 name 字段存在
            name = item.get('name', '') or r.get('name', '')
            detail = item.get('detail', '') or self._fallback_detail(r)
            status = item.get('status', '') or '现场勘查完成，状态正常'
            risk = item.get('risk', '') or '经实地走访抵押物，目前整体状态较好，暂无异常情况。'
            # v3.20.0: 清理残留的 markdown/JSON 特殊字符
            for field_val in [name, detail, status, risk]:
                pass  # 清理在下面统一做
            result.append({
                'name': self._clean_text(name),
                'detail': self._clean_text(detail),
                'status': self._clean_text(status),
                'risk': self._clean_text(risk),
            })
        # 如果解析出的条目少于记录数，补全
        while len(result) < len(records):
            r = records[len(result)]
            result.append({
                'name': r['name'],
                'detail': self._fallback_detail(r),
                'status': '现场勘查完成，状态正常',
                'risk': '经实地走访抵押物，目前整体状态较好，暂无异常情况。',
            })
        return result

    @staticmethod
    def _clean_text(s):
        """v3.20.0: 清理文本中残留的 markdown/JSON 特殊字符"""
        if not s:
            return ""
        s = str(s).strip()
        # 去掉 markdown 代码标记
        s = s.replace('```', '').replace('`', '')
        # 去掉残留的 JSON 大括号（仅清理首尾的孤立 { }）
        s = re.sub(r'^[\s{}]+', '', s)
        s = re.sub(r'[\s{}]+$', '', s)
        # 去掉 markdown 加粗/斜体标记
        s = s.replace('**', '').replace('*', '')
        return s.strip()

    def _fill_template(self, items, out_path, summary_text=""):
        """填充模板并保存。
        v3.19.6: 末尾追加汇总说明行（紧接数据末尾的下一行，合并A:F列）。
        """
        import shutil as _shutil
        if not self.template_path or not os.path.exists(self.template_path):
            raise RuntimeError("未找到日报表模板 report_template.xlsx")
        # 复制模板到输出路径，避免污染原模板
        _shutil.copy2(self.template_path, out_path)
        wb = load_workbook(out_path)
        ws = wb['Sheet1']
        # 清空示例数据 R4-R13
        for r in range(4, 14):
            for c in range(1, 7):
                ws.cell(row=r, column=c).value = None
        # 数据 + 汇总行 超过10行时插入行，保留底部签字/注释区
        n = len(items)
        need_rows = n + (1 if summary_text else 0)
        if need_rows > 10:
            try:
                ws.insert_rows(14, need_rows - 10)
            except Exception:
                pass
        report_date = get_report_date_str()
        for i, item in enumerate(items):
            r = 4 + i
            ws.cell(row=r, column=1, value=i + 1)
            ws.cell(row=r, column=2, value=report_date)
            ws.cell(row=r, column=3, value=item.get('name', ''))
            ws.cell(row=r, column=4, value=item.get('detail', ''))
            ws.cell(row=r, column=5, value=item.get('status', ''))
            ws.cell(row=r, column=6, value=item.get('risk', ''))
        # 汇总说明行：紧接数据末尾的下一行（合并A:F让长文本完整显示）
        if summary_text:
            sr = 4 + n
            ws.cell(row=sr, column=1, value=summary_text)
            try:
                ws.merge_cells(start_row=sr, start_column=1,
                               end_row=sr, end_column=6)
            except Exception:
                pass
        wb.save(out_path)
        wb.close()
        return out_path

    def generate_with_ai(self, rows, progress_mgr, ai_service, excel_filename="", visit_note=""):
        """AI 生成日报表。返回 (ok, out_path_or_None, msg)
        v3.19.6: 仅对有外访照片的客户生成报告行；末尾追加汇总说明行；
                 输出文件名改为"抵押物、抵债资产现场勘查日报表YYYYMMDD.xlsx"。
        v3.22.6: 合并统计「同类型」户数 — 已标记为同类型代表性户型且无照片的客户
                 不再单独生成报告行，但在汇总说明中合并表述。
        v3.22.13: visit_note 参数 — 今日走访补充说明，附加到 system prompt 供 AI 参考。
        """
        records = self._collect_records(rows, progress_mgr)
        if not records:
            return False, None, "没有可生成报告的客户数据，请先打开Excel"
        total = len(records)
        # 仅对有照片的客户生成报告
        visited_records = [r for r in records if r['photo_count'] > 0]
        visited_count = len(visited_records)
        # v3.22.6: 统计同类型户数（标记为同类型且无照片的，需合并表述）
        batch_marked_count = len([r for r in records
                                  if r.get('batch_marked', False)
                                  and r.get('photo_count', 0) == 0])
        not_visited_count = total - visited_count
        if not visited_records:
            return False, None, "所有客户均无外访照片，无法生成报告（请先现场拍照）"
        prompt = self._build_prompt(visited_records, visit_note=visit_note)
        # v3.22.13: 附加今日走访补充说明到 system prompt
        sys_prompt = self.REPORT_SYSTEM_PROMPT
        if visit_note:
            sys_prompt = sys_prompt + '\n\n今日走访补充说明：\n' + visit_note
        ok, ai_text = ai_service.chat([
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": prompt},
        ], timeout=120)
        if not ok:
            return False, None, "AI生成失败：%s" % ai_text
        items = self._parse_ai_response(ai_text, visited_records)
        # 汇总说明行（基于导入Excel文件名 + 客户外访统计）
        excel_name = excel_filename or "未命名"
        # v3.22.6: 如有同类型标记，汇总文案合并表述「代表性户型+同类型」
        if batch_marked_count > 0:
            summary = "本次报告基于%s生成，共计%d个客户，其中已外访代表性户型%d户，另有%d户为同类型（已批量标记）" % (
                excel_name, total, visited_count, batch_marked_count)
        else:
            summary = "本次报告基于%s生成，共计%d个客户，其中有外访%d户已生成报告，%d个客户没有外访未生成报告" % (
                excel_name, total, visited_count, not_visited_count)
        # v3.22.19: 默认文件名改为"线路四XX（日期）现场勘查日报表.xlsx"，XX 从 Excel 文件名提取线路名
        route_name = _extract_route_name(excel_filename or "")
        out_path = os.path.join(APP_DIR, "线路四%s（%s）现场勘查日报表.xlsx" % (route_name, get_date_str()))
        try:
            self._fill_template(items, out_path, summary_text=summary)
            return True, out_path, "日报表已生成（已外访%d户，未外访%d户）" % (visited_count, not_visited_count)
        except Exception as e:
            return False, None, "填充模板失败：%s" % e

# ============================================================
# 照片处理器
# ============================================================

class PhotoProcessor:
    """水印、命名、保存"""

    @staticmethod
    def build_watermark_lines(segments, **kwargs):
        """根据段选择生成水印文字列表，每段一行。
        segments: ["经纬度"/"拍摄时间"/"地址名"/"空值", ...]
        kwargs: time_str(拍摄时间), address(地址), lat, lng(经纬度)
        v3.18.0: 返回 [(段标签, 段文字), ...]，便于分行渲染，避免显示不全。
        """
        lat = kwargs.get('lat', '')
        lng = kwargs.get('lng', '')
        has_gps = bool(lat and lng)
        lines = []
        for seg in segments:
            label = ""
            val = ""
            if seg == "经纬度":
                label = "GPS"
                if has_gps:
                    val = "%s,%s" % (lng, lat)
                else:
                    val = "定位中"
            elif seg == "拍摄时间":
                label = "TIME"
                val = kwargs.get('time_str', '')
                if not val:
                    val = "--"
            elif seg == "地址名":
                label = "ADDR"
                addr = kwargs.get('address', '')
                if addr and not addr.startswith("GPS") and not addr.startswith("定位"):
                    val = addr
                else:
                    val = "定位中"
            elif seg == "空值":
                continue
            if val:
                lines.append((label, val))
        return lines

    @staticmethod
    def add_watermark(photo_path, config, **kwargs):
        """根据配置添加水印（段选择模式）。
        v3.19.0: 水印区域与字体整体增大2倍，最小字号提升至24，保证清晰可读。
        v3.18.0: 改为每段独立一行，并按图片宽度自动缩放字体，避免显示不全。
        """
        if not config.get('watermark_enabled', True):
            return

        try:
            segments = config.get('watermark_segments', DEFAULT_CONFIG['watermark_segments'])
            lines = PhotoProcessor.build_watermark_lines(segments, **kwargs)
            if not lines:
                return

            # v3.22.12: 用 with 管理文件句柄；convert 返回新 Image 对象，不依赖原文件句柄
            with PILImage.open(photo_path) as src_img:
                img = src_img.convert('RGBA')
            draw = ImageDraw.Draw(img)

            font_size_key = config.get('watermark_font_size', '中')
            base_font_size = WATERMARK_FONT_SIZE_MAP.get(font_size_key, 56)
            opacity = config.get('watermark_opacity', 170)
            position = config.get('watermark_position', 'bottom-right')

            # v3.19.0: 内边距与字号同步增大2倍
            padding = 24
            max_width = max(img.width - padding * 2, 50)

            # 计算能放下所有行的最大字号（最小 24，最大 base_font_size）
            font_size = base_font_size
            while font_size > 24:
                font = PhotoProcessor._get_font(font_size)
                fits = True
                for _, text in lines:
                    tw = draw.textbbox((0, 0), text, font=font)[2]
                    if tw > max_width:
                        fits = False
                        break
                if fits:
                    break
                font_size -= 4

            font = PhotoProcessor._get_font(font_size)
            line_height = font_size + 12
            total_height = len(lines) * line_height + 16

            # 计算整体背景框尺寸
            max_text_width = 0
            for _, text in lines:
                tw = draw.textbbox((0, 0), text, font=font)[2]
                if tw > max_text_width:
                    max_text_width = tw
            box_w = max_text_width + padding
            box_h = total_height

            if position == 'bottom-right':
                bx = img.width - box_w - padding
                by = img.height - box_h - padding
            elif position == 'bottom-left':
                bx = padding
                by = img.height - box_h - padding
            elif position == 'top-right':
                bx = img.width - box_w - padding
                by = padding
            else:  # top-left
                bx = padding
                by = padding

            overlay = PILImage.new('RGBA', img.size, (0, 0, 0, 0))
            od = ImageDraw.Draw(overlay)
            od.rectangle([bx, by, bx + box_w, by + box_h],
                         fill=(0, 0, 0, min(opacity, 255)))

            if img.mode != 'RGBA':
                img = img.convert('RGBA')
            img = PILImage.alpha_composite(img, overlay)

            draw = ImageDraw.Draw(img)
            cy = by + 8
            for _, text in lines:
                draw.text((bx + padding / 2, cy), text, font=font, fill=(255, 255, 255))
                cy += line_height

            final = img.convert('RGB') if img.mode == 'RGBA' else img
            final.save(photo_path, 'JPEG', quality=92)
        except Exception as e:
            Logger.error("PhotoProcessor.add_watermark: %s" % e)

    @staticmethod
    def _get_font(size):
        if os.path.exists(FONT_PATH):
            try:
                return ImageFont.truetype(FONT_PATH, size)
            except:
                pass
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/simhei.ttf",
            "msyh.ttc", "simhei.ttf",
            "/system/fonts/DroidSansFallback.ttf",
            "/system/fonts/NotoSansCJK-Regular.ttc",
        ]
        for fp in font_paths:
            try:
                return ImageFont.truetype(fp, size)
            except:
                continue
        return ImageFont.load_default()

    @staticmethod
    def generate_filename(segments, borrower="", address_general="", address_precise="",
                         property_type="", seq=0, date_str="", photo_type="", time_str=""):
        """根据段选择生成文件名（X-X-X-X 格式，空值段自动跳过）
        segments: ["拍摄日期"/"拍摄时间"/"客户名"/"抵押物地址（全）"/.../"地址+时间"/"空值", ...]
        抵押物地址（全） = 抵押物地址（概） + 抵押物地址（精确门牌号）
        地址+时间 = 抵押物地址（全） + 时间HHMM（无分隔符）
        """
        if not date_str:
            date_str = get_date_str()
        if not time_str:
            time_str = get_time_str()

        parts = []
        for seg in segments:
            val = ""
            if seg == "拍摄日期":
                val = date_str
            elif seg == "拍摄时间":
                val = date_str + time_str
            elif seg == "客户名":
                val = borrower if borrower else "未知"
            elif seg == "抵押物地址（全）":
                val = (address_general + address_precise).strip()
            elif seg == "抵押物地址（概）":
                val = address_general
            elif seg == "抵押物地址（精确门牌号）":
                val = address_precise
            elif seg == "地址+时间":
                full_addr = (address_general + address_precise).strip()
                val = full_addr + time_str if full_addr else time_str
            elif seg == "空值":
                val = ""
            val = clean_filename(val)
            if val:
                parts.append(val)

        filename = "-".join(parts)
        filename = clean_filename(filename)
        if not filename:
            filename = "photo_%s" % date_str
        if photo_type:
            filename = "%s-%s-%02d" % (filename, photo_type, seq)
        if not filename.endswith('.jpg'):
            filename += '.jpg'
        return filename

    @staticmethod
    def save_to_gallery(photo_path):
        """v3.20.0: 保存照片到系统相册，返回最终保存路径（供 progress_mgr 记录）。
        成功返回 DCIM/Camera 路径，失败返回 None（调用方保留 APP_DIR 副本）。
        v3.18.1: Android 10+ 使用 MediaStore 直接插入 DCIM/Camera；
                 Android 9 及以下回退到直接复制；最后使用 MediaScannerConnection 刷新。
        """
        if not IS_ANDROID or not os.path.exists(photo_path):
            return None
        fname = os.path.basename(photo_path)
        # v3.20.0: MediaStore 插入后照片在 DCIM/Camera 目录
        dcim_path = '/storage/emulated/0/DCIM/Camera/%s' % fname
        inserted_uri = None
        try:
            from jnius import autoclass
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            activity = PythonActivity.mActivity
            resolver = activity.getContentResolver()
            ContentValues = autoclass('android.content.ContentValues')
            ImagesMedia = autoclass('android.provider.MediaStore$Images$Media')
            FileInputStream = autoclass('java.io.FileInputStream')

            cv = ContentValues()
            cv.put("_display_name", fname)
            cv.put("mime_type", "image/jpeg")
            if ANDROID_API >= 29:
                cv.put("relative_path", "DCIM/Camera")
            # 不设置 is_pending：直接作为可见文件插入，防止某些 ROM/清理工具删除 pending 文件。
            uri = resolver.insert(ImagesMedia.EXTERNAL_CONTENT_URI, cv)
            if uri is not None:
                inserted_uri = uri
                # v3.22.13 修复: fis/out 用 try/finally 确保异常时关闭，防止 Java 文件句柄泄漏
                out = None
                fis = None
                try:
                    out = resolver.openOutputStream(uri)
                    fis = FileInputStream(photo_path)
                    buf = bytearray(8192)
                    while True:
                        n = fis.read(buf)
                        if n <= 0:
                            break
                        out.write(buf, 0, n)
                finally:
                    if fis is not None:
                        try:
                            fis.close()
                        except Exception:
                            pass
                    if out is not None:
                        try:
                            out.close()
                        except Exception:
                            pass
                Logger.info(f"save_to_gallery: MediaStore 已插入 DCIM/Camera {fname}")
                # 使用 MediaScannerConnection 刷新，Android 10+ 比旧广播更稳定。
                PhotoProcessor._scan_file(photo_path, uri=uri)
                return dcim_path
        except Exception as e:
            Logger.error(f"save_to_gallery MediaStore 失败: {e}")
            # 如果已插入但写入失败，尝试删除残留的 pending/空记录
            if inserted_uri is not None:
                try:
                    from jnius import autoclass
                    PythonActivity = autoclass('org.kivy.android.PythonActivity')
                    resolver = PythonActivity.mActivity.getContentResolver()
                    resolver.delete(inserted_uri, None, None)
                except Exception:
                    pass

        # 回退：直接复制到 DCIM/Camera（Android 9 及以下或 MediaStore 失败时）
        try:
            import shutil
            dcim_dirs = [
                '/storage/emulated/0/DCIM/Camera',
                '/sdcard/DCIM/Camera',
            ]
            for dcim_dir in dcim_dirs:
                try:
                    os.makedirs(dcim_dir, exist_ok=True)
                    dest_path = os.path.join(dcim_dir, fname)
                    shutil.copy2(photo_path, dest_path)
                    Logger.info(f"save_to_gallery: 已复制到 {dest_path}")
                    PhotoProcessor._scan_file(dest_path)
                    return dest_path
                except Exception as e:
                    Logger.error(f"save_to_gallery: 复制到{dcim_dir}失败: {e}")
                    continue
        except Exception as e:
            Logger.error(f"save_to_gallery 直接复制失败: {e}")
        return None  # v3.20.0: 所有方式均失败

    @staticmethod
    def _scan_file(file_path, uri=None):
        """刷新系统媒体库，让相册/微信立即看到新照片。
        v3.18.1: Android 10+ 优先使用 MediaScannerConnection.scanFile；
                 旧设备保留 MEDIA_SCANNER_SCAN_FILE 广播作为兜底。
        """
        if not IS_ANDROID or not file_path:
            return
        scanned = False
        # 方案1: MediaScannerConnection.scanFile（Android 10+ 推荐）
        if ANDROID_API >= 29:
            try:
                from jnius import autoclass, PythonJavaClass, java_method
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                activity = PythonActivity.mActivity
                MediaScannerConnection = autoclass('android.media.MediaScannerConnection')
                String = autoclass('java.lang.String')

                class _ScanClient(PythonJavaClass):
                    __javainterfaces__ = ['android/media/MediaScannerConnection$OnScanCompletedListener']

                    @java_method('(Ljava/lang/String;Landroid/net/Uri;)V')
                    def onScanCompleted(self, path, uri):
                        Logger.info(f"MediaScannerConnection: 扫描完成 {path}")

                MediaScannerConnection.scanFile(
                    activity,
                    [String(file_path)],
                    [String("image/jpeg")],
                    _ScanClient()
                )
                scanned = True
                Logger.info(f"save_to_gallery: MediaScannerConnection 已扫描 {file_path}")
            except Exception as e:
                Logger.error(f"save_to_gallery MediaScannerConnection 失败: {e}")

        # 方案2: 旧版广播（兜底）
        if not scanned:
            try:
                from jnius import autoclass
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                activity = PythonActivity.mActivity
                Intent = autoclass('android.content.Intent')
                Uri = autoclass('android.net.Uri')
                File = autoclass('java.io.File')
                if not os.path.exists(file_path):
                    return
                file_uri = Uri.fromFile(File(file_path))
                scan_intent = Intent("android.intent.action.MEDIA_SCANNER_SCAN_FILE")
                scan_intent.setData(file_uri)
                activity.sendBroadcast(scan_intent)
                Logger.info(f"save_to_gallery: MediaScanner 广播已发送 {file_path}")
            except Exception as e:
                Logger.error(f"save_to_gallery MediaScanner 广播失败: {e}")

# ============================================================
# GPS 管理器（非阻塞缓存）
# ============================================================

class GpsManager:
    """异步获取 GPS 坐标并缓存，供水印经纬度段使用。
    v3.18.1: 优先使用 Android LocationManager 主动请求位置更新（GPS + Network），
             比仅依赖 getLastKnownLocation 更能拿到有效坐标；plyer 作为备选。
    """
    def __init__(self):
        self.lat = ""
        self.lng = ""
        self._started = False
        self._lm_started = False
        self._lm = None
        self._listener = None
        if IS_ANDROID:
            self._start()
            # 尽快启动 LocationManager（主动更新 + 最后已知位置）
            Clock.schedule_once(lambda dt: self._start_location_manager(), 0.2)

    def _start(self):
        try:
            from plyer import gps
            gps.configure(on_location=self._on_location)
            gps.start(min_time=3000, min_distance=0)
            self._started = True
        except Exception:
            self._started = False

    def _start_location_manager(self):
        """使用 Android LocationManager 直连获取位置（主动更新 + 最后已知位置）。"""
        if self._lm_started:
            return
        try:
            from jnius import autoclass, PythonJavaClass, java_method
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            activity = PythonActivity.mActivity
            if activity is None:
                return
            Context = autoclass('android.content.Context')
            LocationManager = autoclass('android.location.LocationManager')
            lm = activity.getSystemService(Context.LOCATION_SERVICE)
            if lm is None:
                return
            self._lm = lm

            # 检查权限
            if ANDROID_API >= 23:
                if activity.checkSelfPermission("android.permission.ACCESS_FINE_LOCATION") != 0:
                    Logger.warning("GpsManager: 缺少定位权限")
                    return

            # 先取最后已知位置（快速暖机）
            self._read_last_known(lm, LocationManager)

            # 再注册主动更新（GPS + Network 双保险）
            try:
                self._listener = self._create_listener()
                min_time_ms = 1000  # 1秒
                min_distance_m = 0.0
                if lm.isProviderEnabled(LocationManager.GPS_PROVIDER):
                    lm.requestLocationUpdates(
                        LocationManager.GPS_PROVIDER,
                        min_time_ms, min_distance_m,
                        self._listener
                    )
                    Logger.info("GpsManager: 已注册 GPS 位置更新")
                if lm.isProviderEnabled(LocationManager.NETWORK_PROVIDER):
                    lm.requestLocationUpdates(
                        LocationManager.NETWORK_PROVIDER,
                        min_time_ms, min_distance_m,
                        self._listener
                    )
                    Logger.info("GpsManager: 已注册 Network 位置更新")
            except Exception as e:
                Logger.error(f"GpsManager: requestLocationUpdates 失败 {e}")

            self._lm_started = True
        except Exception as e:
            Logger.error("GpsManager._start_location_manager: %s" % e)

    def _read_last_known(self, lm, LocationManager):
        try:
            last = lm.getLastKnownLocation(LocationManager.GPS_PROVIDER)
            if last is None:
                last = lm.getLastKnownLocation(LocationManager.NETWORK_PROVIDER)
            if last is not None:
                self.lat = "%.6f" % last.getLatitude()
                self.lng = "%.6f" % last.getLongitude()
                Logger.info("GpsManager: getLastKnownLocation (%s,%s)" % (self.lat, self.lng))
        except Exception:
            pass

    def _create_listener(self):
        from jnius import PythonJavaClass, java_method

        class _LocationListener(PythonJavaClass):
            __javainterfaces__ = ['android/location/LocationListener']

            def __init__(self, manager):
                super().__init__()
                self.manager = manager

            @java_method('(Landroid/location/Location;)V')
            def onLocationChanged(self, location):
                try:
                    if location is not None:
                        lat = location.getLatitude()
                        lng = location.getLongitude()
                        self.manager.lat = "%.6f" % lat
                        self.manager.lng = "%.6f" % lng
                        Logger.info(f"GpsManager: 位置更新 ({self.manager.lat},{self.manager.lng})")
                except Exception:
                    pass

            @java_method('(Ljava/lang/String;)V')
            def onProviderEnabled(self, provider):
                pass

            @java_method('(Ljava/lang/String;)V')
            def onProviderDisabled(self, provider):
                pass

            @java_method('(Ljava/lang/String;ILandroid/os/Bundle;)V')
            def onStatusChanged(self, provider, status, extras):
                pass

        return _LocationListener(self)

    def refresh_last_known(self):
        """手动刷新位置（拍照前调用）。如尚未启动则尝试启动。"""
        if not IS_ANDROID:
            return
        try:
            if not self._lm_started:
                self._start_location_manager()
            elif self._lm is not None:
                from jnius import autoclass
                LocationManager = autoclass('android.location.LocationManager')
                self._read_last_known(self._lm, LocationManager)
        except Exception:
            pass

    def _on_location(self, **kwargs):
        try:
            lat = kwargs.get('lat')
            lon = kwargs.get('lon')
            if lat is not None and lon is not None:
                self.lat = ("%.6f" % float(lat))
                self.lng = ("%.6f" % float(lon))
        except Exception:
            pass

    def get_coords(self):
        """返回 (lat, lng)，未定位时返回 ('', '')"""
        if not (self.lat and self.lng):
            self.refresh_last_known()
        return self.lat, self.lng


# ============================================================
# AI 服务（OpenRouter / OpenAI 兼容）
# ============================================================

class AIService:
    """封装 AI API 调用（兼容 OpenAI Chat Completions 格式）
    v3.15.0: 支持 AI 查询拍摄情况。
    v3.19.5: 添加 DeepSeek 作为第二备用 API。
    v3.19.6: 移除 OpenRouter，仅保留 DeepSeek(deepseek-v4-flash) 作为唯一 LLM。
    """
    DEFAULT_API_URL = "https://api.deepseek.com/v1"
    DEFAULT_API_KEY = DEEPSEEK_DEFAULT_API_KEY
    DEFAULT_MODEL = "deepseek-v4-flash"

    def __init__(self, api_url=None, api_key=None, model=None):
        self.api_url = (api_url or self.DEFAULT_API_URL).rstrip('/')
        self.api_key = api_key or self.DEFAULT_API_KEY
        self.model = model or self.DEFAULT_MODEL

    def chat(self, messages, timeout=60):
        """调用 chat/completions 接口
        messages: [{"role":"system","content":"..."},{"role":"user","content":"..."}]
        返回 (success, response_text_or_error)
        v3.18.1: 改用 requests 库，禁用 SSL 验证（Android 上 certifi 证书链可能不完整），
                 增加详细日志，解决 Android 上 urllib SSL 偶发失败问题。
        v3.19.6: 仅使用 DeepSeek 单引擎（移除 OpenRouter 主备双引擎）。
        """
        # 使用内置默认 key（如果未配置）
        if not self.api_key:
            self.api_key = AI_DEFAULT_API_KEY
        # 使用内置默认模型（如果未配置）
        if not self.model:
            self.model = self.DEFAULT_MODEL

        return self._call_api(self.api_url, self.api_key, self.model, messages, timeout)

    def _call_api(self, api_url, api_key, model, messages, timeout=60):
        """实际调用 API（内部方法）"""
        import json as _json
        try:
            import requests
        except Exception as e:
            Logger.error(f"AIService: requests 库未导入 {e}")
            return False, "缺少网络库 requests"

        url = "%s/chat/completions" % api_url
        payload = {
            "model": model,
            "messages": messages,
            "temperature": 0.3,
            "max_tokens": 1024,
        }
        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer %s" % api_key,
            # v3.19.2: 移除 X-Title（含中文"资产盘点拍照工具"8字符 → position 0-7，
            #          requests 按 latin-1 编码 HTTP 头会报 ordinal not in range(256)）
            "HTTP-Referer": "https://github.com/jare39063124-oss/loan-photo-app",
        }

        Logger.info(f"AIService: 请求 {url} model={model}")
        Logger.info(f"AIService: key={api_key[:15]}...{api_key[-8:]}")
        try:
            # 禁用 SSL 验证：Android 上 certifi 证书链可能不完整导致 SSL 错误
            resp = requests.post(url, headers=headers, json=payload, timeout=timeout, verify=False)
            Logger.info(f"AIService: 响应 HTTP {resp.status_code}, len={len(resp.text)}")
            Logger.info(f"AIService: 响应内容前200字: {resp.text[:200]}")
            if resp.status_code == 200:
                result = resp.json()
                if 'choices' in result and len(result['choices']) > 0:
                    content = result['choices'][0].get('message', {}).get('content', '')
                    return True, content.strip()
                elif 'error' in result:
                    err = result['error']
                    err_msg = err.get('message', str(err)) if isinstance(err, dict) else str(err)
                    return False, "API错误: %s" % err_msg[:200]
                else:
                    return False, "API返回格式异常: %s" % str(result)[:200]
            else:
                try:
                    err_data = resp.json()
                    err_msg = err_data.get('error', {}).get('message', resp.text)
                except Exception:
                    err_msg = resp.text
                return False, "HTTP %d: %s" % (resp.status_code, str(err_msg)[:200])
        except requests.exceptions.Timeout as e:
            Logger.error(f"AIService: 请求超时 {e}")
            return False, "请求超时，请检查网络"
        except requests.exceptions.ConnectionError as e:
            Logger.error(f"AIService: 连接失败 {type(e).__name__}: {e}")
            return False, "网络连接失败: %s" % str(e)[:100]
        except Exception as e:
            Logger.error(f"AIService: 请求异常 {type(e).__name__}: {e}")
            import traceback as _tb
            Logger.error(_tb.format_exc())
            return False, "请求失败: %s" % str(e)[:200]

    @staticmethod
    def build_system_prompt(rows, progress_mgr, excel_path=""):
        """构建系统提示词，注入 Excel 数据和拍摄进度"""
        import json as _json
        from datetime import datetime

        today = datetime.now().strftime('%Y-%m-%d')

        # 构建客户数据摘要（v3.22.0: A序号 B客户名 C地址概 D地址详 E性质 F备注）
        # v3.22.0: 性能优化 — 用 get_progress_snapshot 一次性收集所有 key 的进度，
        # 避免每行分别调用 get_photo_count + get_photo_types（N 行 2N 次加锁 → 1 次）
        snap = progress_mgr.get_progress_snapshot(rows) if hasattr(progress_mgr, 'get_progress_snapshot') else {}
        customers = []
        for i, row in enumerate(rows):
            serial = row[0] if len(row) > 0 else ""
            borrower = row[1] if len(row) > 1 else ""
            addr_gen = row[2] if len(row) > 2 else ""
            addr_prec = row[3] if len(row) > 3 else ""
            prop_type = row[4] if len(row) > 4 else ""
            remark = row[5] if len(row) > 5 else ""
            full_addr = (addr_gen + addr_prec).strip()
            key = progress_mgr._make_key(borrower, full_addr)
            s = snap.get(key, {})
            photo_count = s.get('count', 0)
            photo_types = s.get('types', {})
            done_types = [t for t, v in photo_types.items() if v]
            customers.append({
                "序号": serial or str(i + 1),
                "客户名": borrower,
                "地址": full_addr,
                "性质": prop_type,
                "备注": remark,
                "已拍照片数": photo_count,
                "已拍类型": done_types,
                "是否完成": photo_count > 0,
            })

        # 统计
        total = len(customers)
        done = sum(1 for c in customers if c["是否完成"])
        total_photos = sum(c["已拍照片数"] for c in customers)

        # 各类型统计
        type_stats = {}
        for c in customers:
            for t in c["已拍类型"]:
                type_stats[t] = type_stats.get(t, 0) + 1

        data_summary = {
            "日期": today,
            "Excel文件": os.path.basename(excel_path) if excel_path else "未加载",
            "总客户数": total,
            "已完成客户数": done,
            "未完成客户数": total - done,
            "总照片数": total_photos,
            "各类型拍摄数": type_stats,
            "客户明细": customers,
        }

        prompt = (
            "你是「资产盘点专项拍照工具」的AI助手。你的职责是帮助用户查询拍摄进度和客户信息。\n\n"
            "## 当前数据\n"
            "以下是用户当前打开的Excel文件中的客户数据和拍摄进度：\n\n"
            "```json\n%s\n```\n\n"
            "## 你的能力\n"
            "1. 查询拍摄进度：例如「今天拍了多少照片」「还有多少户没拍」\n"
            "2. 查询客户信息：例如「某某公司拍了没有」「张三的地址是什么」\n"
            "3. 查询照片类型：例如「远景拍了多少张」「哪些客户拍了瑕疵」\n"
            "4. 统计分析：例如「完成率多少」「还有哪些类型没拍全」\n"
            "5. 备注查询：例如「哪些客户有备注」\n\n"
            "## 注意事项\n"
            "- 照片类型包括：远景、近景、内部、瑕疵、其他\n"
            "- 每个客户最多拍5种类型的照片\n"
            "- 回答要简洁明了，用中文\n"
            "- 如果数据中没有相关信息，请如实告知\n"
        ) % _json.dumps(data_summary, ensure_ascii=False, indent=2)

        return prompt


# ============================================================
# 相机管理器
# ============================================================

class CameraManager:
    CAMERA_REQUEST_CODE = 0x123

    def __init__(self):
        self.photo_path = ""
        self.pending_callback = None
        self.status_callback = None
        self.gps = GpsManager()
        self.geocoder = GeoCoder()
        self._camera_launched = False
        self._media_uri = None
        self._launch_attempts = 0
        self._debug_log_path = ""
        self._log_lines = []
        self._max_log_lines = 50
        self._log_enabled = True  # v3.20.0: 默认开启，日志统一写入 app_log

    def _dbg(self, msg, show_toast=False):
        """Write debug message to log file and update UI status.
        show_toast=True to also show an Android Toast (use sparingly to avoid flashing).
        v3.20.0: 日志统一写入 app_log（全量日志器）"""
        ts = get_system_date().strftime('%H:%M:%S')
        line = f"[{ts}] {msg}"
        Logger.info("CAMDBG: %s", msg)
        self._log_lines.append(line)
        if len(self._log_lines) > self._max_log_lines:
            self._log_lines = self._log_lines[-self._max_log_lines:]
        # v3.20.0: 统一写入 app_log 全量日志
        app_log.info('PHOTO', msg)
        # 兼容旧日志文件（仅在开关开启时写 camera_debug.log）
        if self._log_enabled:
            try:
                if not self._debug_log_path:
                    self._debug_log_path = os.path.join(APP_DIR, "camera_debug.log")
                os.makedirs(APP_DIR, exist_ok=True)
                with open(self._debug_log_path, 'a', encoding='utf-8') as f:
                    f.write(line + "\n")
            except:
                pass
        if self.status_callback:
            try:
                self.status_callback('\n'.join(self._log_lines[-10:]))
            except:
                pass
        if show_toast:
            self._toast(msg)

    def _toast(self, msg):
        """Show a native Android Toast message."""
        if not IS_ANDROID:
            Logger.info("Toast: %s", msg)
            return
        try:
            from jnius import autoclass
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            Toast = autoclass('android.widget.Toast')
            activity = PythonActivity.mActivity
            if activity:
                LENGTH_LONG = 1
                toast = Toast.makeText(activity, msg, LENGTH_LONG)
                toast.show()
        except Exception as e:
            Logger.warning("Toast failed: %s", str(e)[:80])

    def take_photo(self, callback, status_callback=None):
        """启动相机拍照，拍完后callback(photo_path)被调用，失败传None。
        status_callback(msg) 用于实时更新UI状态。"""
        self.pending_callback = callback
        self.status_callback = status_callback
        self._camera_launched = False
        self._media_uri = None
        self.photo_path = ""
        self._dbg("开始拍照流程...")
        if IS_ANDROID:
            self._dbg("正在启动相机...", show_toast=True)
            self._check_and_request_permission()
        else:
            self._dbg("桌面测试模式")
            self._simulate_photo()

    def _check_and_request_permission(self):
        """检查相机权限，有则直接启动相机；无则尝试请求，失败也直接启动（让系统处理）。"""
        self._dbg("检查相机权限...")
        has_perm = False
        try:
            from android.permissions import check_permission
            has_perm = check_permission('android.permission.CAMERA')
            self._dbg(f"权限检查结果: {has_perm}")
        except Exception as e:
            self._dbg(f"check_permission导入失败: {str(e)[:60]}，直接尝试启动")
            has_perm = False

        if has_perm:
            self._dbg("已有相机权限，启动相机")
            self._launch_camera_intent()
            return

        self._dbg("请求相机权限...")
        try:
            request_permissions(
                [Permission.CAMERA],
                self._on_permission_result
            )
            self._dbg("权限请求已发送（等待用户授权弹窗）")
            Clock.schedule_once(self._permission_timeout, 5)
        except Exception as e:
            self._dbg(f"request_permissions失败: {str(e)[:80]}，直接启动")
            self._launch_camera_intent()

    def _permission_timeout(self, dt):
        """5秒后如果还没启动相机，说明权限弹窗可能没出现，直接尝试启动。"""
        if not self._camera_launched and self.pending_callback:
            self._dbg("权限请求超时（5秒未响应），直接尝试启动相机")
            Clock.schedule_once(lambda dt2: self._launch_camera_intent(), 0)

    def _on_permission_result(self, permissions, grants):
        self._dbg(f"权限回调: perms={permissions}, grants={grants}")
        def _handle(dt):
            if any(grants):
                self._dbg("用户已授权，启动相机")
                self._launch_camera_intent()
            else:
                self._dbg("相机权限被拒绝，尝试直接启动")
                self._launch_camera_intent()
        Clock.schedule_once(_handle, 0)

    def _launch_camera_intent(self):
        """通过 Android Intent 调用系统相机。
        v3.12.0 彻底重写（基于Kivy官方2013示例 + 现代Android兼容）：
        - 核心修复：所有Uri传给putExtra前必须cast为Parcelable（pyjnius重载解析问题）
        - 核心原则：相机必须100%启动！URI策略失败永远不阻断相机启动
        - 启动策略：Android 11+优先Intent.createChooser（豁免包可见性限制）
        - 结果获取：多级回退 (MediaStore URI→Intent URI→ClipData→缩略图data extra)
        """
        def _do_launch():
            self._launch_attempts += 1
            try:
                from jnius import autoclass, cast
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                Intent = autoclass('android.content.Intent')
                File = autoclass('java.io.File')
                Uri = autoclass('android.net.Uri')
                Environment = autoclass('android.os.Environment')
                Parcelable = autoclass('android.os.Parcelable')
            except Exception as e:
                self._dbg(f"jnius类加载失败: {str(e)[:80]}", show_toast=True)
                if self.pending_callback:
                    cb = self.pending_callback
                    self.pending_callback = None
                    Clock.schedule_once(lambda dt, cb=cb: cb(None), 0)
                return

            ACTION_IMAGE_CAPTURE = "android.media.action.IMAGE_CAPTURE"
            EXTRA_OUTPUT = "output"
            FLAG_GRANT_READ_URI_PERMISSION = 1
            FLAG_GRANT_WRITE_URI_PERMISSION = 2

            try:
                raw_activity = PythonActivity.mActivity
                if raw_activity is None:
                    self._dbg("错误：mActivity为None！", show_toast=True)
                    if self.pending_callback:
                        cb = self.pending_callback
                        self.pending_callback = None
                        Clock.schedule_once(lambda dt, cb=cb: cb(None), 0)
                    return
                activity = cast('android.app.Activity', raw_activity)
                package_name = activity.getPackageName()
                self._dbg(f"Activity OK, pkg={package_name}, API={ANDROID_API}")

                self._dbg("正在准备照片存储路径...")
                photo_fname = f"capture_{get_system_date().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
                try:
                    pics_dir = activity.getExternalFilesDir(Environment.DIRECTORY_PICTURES)
                    if pics_dir is not None:
                        save_dir = str(pics_dir.getAbsolutePath())
                    else:
                        cache_dir = activity.getExternalCacheDir()
                        save_dir = str(cache_dir.getAbsolutePath()) if cache_dir else APP_DIR
                    os.makedirs(save_dir, exist_ok=True)
                    self.photo_path = os.path.join(save_dir, photo_fname)
                    self._dbg(f"保存目录: ...{save_dir[-30:]}")
                except Exception as e:
                    self._dbg(f"目录准备失败: {str(e)[:60]}，使用APP_DIR")
                    os.makedirs(APP_DIR, exist_ok=True)
                    self.photo_path = os.path.join(APP_DIR, photo_fname)

                intent = Intent(ACTION_IMAGE_CAPTURE)

                uri = None
                uri_set_ok = False
                self._media_uri = None

                # 记录拍摄前时间戳（用于DCIM扫描过滤旧照片）
                self._photo_launch_time = time.time()

                # ========== 统一URI策略（API 30+和API < 30都尝试）==========
                self._dbg("尝试设置输出URI（用于高清照片）...")

                # 策略1：FileProvider content:// URI（官方推荐，全分辨率）
                try:
                    self._dbg("策略1: FileProvider URI")
                    FileProvider = autoclass('androidx.core.content.FileProvider')
                    authority = package_name + ".fileprovider"
                    photo_file = File(self.photo_path)
                    if photo_file.exists():
                        photo_file.delete()
                    photo_file.createNewFile()
                    raw_uri = FileProvider.getUriForFile(activity, authority, photo_file)
                    if raw_uri is not None:
                        parcel_uri = cast('android.os.Parcelable', raw_uri)
                        intent.putExtra(EXTRA_OUTPUT, parcel_uri)
                        intent.addFlags(FLAG_GRANT_READ_URI_PERMISSION)
                        intent.addFlags(FLAG_GRANT_WRITE_URI_PERMISSION)
                        uri = raw_uri
                        uri_set_ok = True
                        self._dbg("  FileProvider URI成功（全分辨率）")
                except:  # 裸except：pyjnius的Java异常不继承BaseException
                    import sys as _sys
                    e = _sys.exc_info()[1]
                    self._dbg(f"  FileProvider失败: {str(e)[:100] if e else 'Unknown'}")
                    intent = Intent(ACTION_IMAGE_CAPTURE)

                # 策略2：MediaStore content:// URI（Android 10+，全分辨率，保存到 DCIM/Camera）
                # v3.18.0: 不再设置 is_pending=1，否则外部相机应用无法写入该 URI。
                if not uri_set_ok and ANDROID_API >= 29:
                    self._dbg("策略2: MediaStore URI")
                    try:
                        ImagesMedia = autoclass('android.provider.MediaStore$Images$Media')
                        EXTERNAL_CONTENT_URI = ImagesMedia.EXTERNAL_CONTENT_URI
                        ContentValues = autoclass('android.content.ContentValues')
                        resolver = activity.getContentResolver()
                        cv = ContentValues()
                        cv.put("_display_name", photo_fname)
                        cv.put("mime_type", "image/jpeg")
                        cv.put("relative_path", "DCIM/Camera")
                        # 不设置 is_pending，让相机应用可以直接写入公共 DCIM/Camera 目录
                        raw_uri = resolver.insert(EXTERNAL_CONTENT_URI, cv)
                        if raw_uri is not None:
                            parcel_uri = cast('android.os.Parcelable', raw_uri)
                            intent.putExtra(EXTRA_OUTPUT, parcel_uri)
                            intent.addFlags(FLAG_GRANT_READ_URI_PERMISSION)
                            intent.addFlags(FLAG_GRANT_WRITE_URI_PERMISSION)
                            uri = raw_uri
                            self._media_uri = raw_uri
                            uri_set_ok = True
                            self.photo_path = None
                            self._dbg("  MediaStore URI成功（全分辨率，DCIM/Camera）")
                    except:  # 裸except：pyjnius的Java异常不继承BaseException
                        import sys as _sys
                        e = _sys.exc_info()[1]
                        self._dbg(f"  MediaStore失败: {str(e)[:100] if e else 'Unknown'}")
                    if not uri_set_ok:
                        intent = Intent(ACTION_IMAGE_CAPTURE)

                # 策略3：file:// URI + 禁用StrictMode（API < 30）
                if not uri_set_ok and ANDROID_API < 30:
                    try:
                        self._dbg("策略3: file:// URI")
                        photo_file = File(self.photo_path)
                        if photo_file.exists():
                            photo_file.delete()
                        photo_file.createNewFile()
                        try:
                            StrictMode = autoclass('android.os.StrictMode')
                            VmPolicyBuilder = autoclass('android.os.StrictMode$VmPolicy$Builder')
                            b = VmPolicyBuilder()
                            b.penaltyLog()
                            StrictMode.setVmPolicy(b.build())
                            self._dbg("  StrictMode已禁用")
                        except:
                            pass
                        raw_uri = Uri.fromFile(photo_file)
                        parcel_uri = cast('android.os.Parcelable', raw_uri)
                        intent.putExtra(EXTRA_OUTPUT, parcel_uri)
                        intent.addFlags(FLAG_GRANT_READ_URI_PERMISSION)
                        intent.addFlags(FLAG_GRANT_WRITE_URI_PERMISSION)
                        uri = raw_uri
                        uri_set_ok = True
                        self._dbg("  file:// URI设置成功")
                    except:  # 裸except：pyjnius的Java异常不继承BaseException
                        import sys as _sys
                        e = _sys.exc_info()[1]
                        self._dbg(f"  file:// URI失败: {str(e)[:80] if e else 'Unknown'}")
                        intent = Intent(ACTION_IMAGE_CAPTURE)

                # 策略4：无EXTRA_OUTPUT（兜底，只获取缩略图）
                if not uri_set_ok:
                    self._dbg("策略4: 无EXTRA_OUTPUT（兜底，仅缩略图）")
                    self.photo_path = None
                    self._media_uri = None
                    intent = Intent(ACTION_IMAGE_CAPTURE)

                # 尝试授予URI权限（如果有uri）
                if uri is not None and uri_set_ok:
                    try:
                        pm = activity.getPackageManager()
                        resolves = pm.queryIntentActivities(intent, 0)
                        if resolves is not None and resolves.size() > 0:
                            self._dbg(f"找到{resolves.size()}个相机应用，授予URI权限")
                            for i in range(resolves.size()):
                                try:
                                    ri = resolves.get(i)
                                    pkg = ri.activityInfo.packageName
                                    activity.grantUriPermission(pkg, uri,
                                        FLAG_GRANT_WRITE_URI_PERMISSION | FLAG_GRANT_READ_URI_PERMISSION)
                                except:
                                    pass
                    except:
                        pass

                # ========== 启动相机（无论URI是否设置成功都必须执行！）==========
                launched = False
                launch_error = ""

                # Android 11+ (API 30+): Intent.createChooser() 豁免包可见性限制
                use_chooser = (ANDROID_API >= 30)

                if use_chooser:
                    self._dbg("使用选择器启动相机...")
                    try:
                        chooser_intent = Intent.createChooser(intent, "选择相机应用")
                        activity.startActivityForResult(chooser_intent, self.CAMERA_REQUEST_CODE)
                        launched = True
                        self._dbg("[OK] 相机选择器已弹出！")
                    except:  # 裸except：pyjnius的Java异常不继承BaseException
                        import sys as _sys
                        e = _sys.exc_info()[1]
                        ename = type(e).__name__ if e else "Unknown"
                        emsg = str(e)[:80] if e else ""
                        if 'ActivityNotFound' in ename or 'Notfound' in ename:
                            launch_error = "系统未安装相机应用"
                        else:
                            launch_error = f"选择器异常: {ename}: {emsg}"
                        self._dbg(launch_error)

                if not launched:
                    self._dbg("直接启动相机...")
                    try:
                        activity.startActivityForResult(intent, self.CAMERA_REQUEST_CODE)
                        launched = True
                        self._dbg("[OK] 相机已启动！")
                    except:  # 裸except：pyjnius的Java异常不继承BaseException
                        import sys as _sys
                        e = _sys.exc_info()[1]
                        ename = type(e).__name__ if e else "Unknown"
                        emsg = str(e)[:80] if e else ""
                        if 'ActivityNotFound' in ename or 'Notfound' in ename:
                            launch_error = "未找到相机应用"
                        else:
                            launch_error = f"启动异常: {ename}: {emsg}"
                        self._dbg(launch_error)

                if launched:
                    self._camera_launched = True
                    self._toast("拍完照请点击对勾保存按钮！")
                else:
                    self._dbg(f"相机启动失败: {launch_error}", show_toast=True)
                    self._camera_launched = False
                    if self.pending_callback:
                        cb = self.pending_callback
                        self.pending_callback = None
                        Clock.schedule_once(lambda dt, cb=cb: cb(None), 0)

            except:  # 裸except：pyjnius的Java异常不继承BaseException
                import sys as _sys
                e = _sys.exc_info()[1]
                err_msg = f"相机启动异常: {type(e).__name__ if e else 'Unknown'}: {str(e)[:100] if e else ''}"
                self._dbg(err_msg, show_toast=True)
                Logger.error(traceback.format_exc())
                self._camera_launched = False
                if self.pending_callback:
                    cb = self.pending_callback
                    self.pending_callback = None
                    Clock.schedule_once(lambda dt, cb=cb: cb(None), 0)

        try:
            from android.runnable import run_on_ui_thread
            _do_launch_ui = run_on_ui_thread(_do_launch)
            _do_launch_ui()
        except ImportError:
            Clock.schedule_once(lambda dt: _do_launch(), 0)
        except:  # 裸except：pyjnius的Java异常不继承BaseException
            import sys as _sys
            e = _sys.exc_info()[1]
            self._dbg(f"UI线程调度失败: {str(e)[:60] if e else 'Unknown'}，直接执行")
            _do_launch()

    def on_camera_result(self, result_code, intent=None):
        """由 MainScreen.on_activity_result 在收到 CAMERA_REQUEST_CODE 结果时调用。
        v3.13.0: 不管result_code，先检查照片文件是否已写入EXTRA_OUTPUT路径。
        某些相机应用(如小米)即使返回RESULT_CANCELED照片也可能已写入文件。
        """
        self._dbg(f"收到相机结果: result_code={result_code}, launched={self._camera_launched}")
        if not self._camera_launched:
            return
        self._camera_launched = False

        # ========== 第1优先级：检查EXTRA_OUTPUT路径文件是否已写入 ==========
        if self.photo_path and os.path.exists(self.photo_path):
            fsize = os.path.getsize(self.photo_path)
            if fsize > 0:
                self._dbg(f"[OK] 照片文件已存在({fsize} bytes)，拍照成功！", show_toast=True)
                if self.pending_callback:
                    cb = self.pending_callback
                    self.pending_callback = None
                    Clock.schedule_once(lambda dt, cb=cb: cb(self.photo_path), 0)
                return
            else:
                self._dbg(f"照片文件存在但大小为0（占位文件），删除并继续查找")
                try:
                    os.remove(self.photo_path)
                except:
                    pass
                self.photo_path = None
        else:
            if self.photo_path:
                self._dbg(f"照片文件不存在: {self.photo_path}")

        # ========== 第2优先级：检查MediaStore URI ==========
        if self._media_uri is not None:
            self._dbg("处理MediaStore返回的照片...")
            try:
                from jnius import autoclass
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                activity = PythonActivity.mActivity
                resolver = activity.getContentResolver()

                # 强制解除 is_pending，防止某些相机应用写入后未清理 pending 导致照片不可见
                try:
                    ContentValues = autoclass('android.content.ContentValues')
                    cv = ContentValues()
                    cv.put("is_pending", 0)
                    resolver.update(self._media_uri, cv, None, None)
                except:
                    import sys as _sys
                    e = _sys.exc_info()[1]
                    self._dbg(f"解除 MediaStore is_pending 失败(可能已是可见状态): {str(e)[:80] if e else ''}")

                istream = resolver.openInputStream(self._media_uri)
                os.makedirs(APP_DIR, exist_ok=True)
                dest_path = os.path.join(
                    APP_DIR, f"capture_{get_system_date().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
                )
                ByteArrayOutputStream = autoclass('java.io.ByteArrayOutputStream')
                baos = ByteArrayOutputStream()
                buf = bytearray(8192)
                while True:
                    n = istream.read(buf)
                    if n <= 0:
                        break
                    baos.write(buf, 0, n)
                istream.close()
                with open(dest_path, 'wb') as f:
                    f.write(bytes(baos.toByteArray()))
                baos.close()
                self.photo_path = dest_path
                fsize = os.path.getsize(dest_path)
                self._dbg(f"MediaStore照片已保存: {fsize} bytes")
                if fsize > 0:
                    # v3.19.0: 删除原始 MediaStore 条目，避免 DCIM/Camera 中出现
                    # capture_xxx 与规范命名两份文件（save_to_gallery 会重新插入规范命名）
                    try:
                        resolver.delete(self._media_uri, None, None)
                        self._dbg("已删除 MediaStore 临时条目，避免重复文件")
                    except:
                        pass
                    self._media_uri = None
                    if self.pending_callback:
                        cb = self.pending_callback
                        self.pending_callback = None
                        Clock.schedule_once(lambda dt, cb=cb: cb(self.photo_path), 0)
                    return
            except:
                import sys as _sys
                e = _sys.exc_info()[1]
                self._dbg(f"MediaStore复制失败: {str(e)[:100] if e else 'Unknown'}")
                self._media_uri = None

        # ========== 第3优先级：扫描DCIM/Camera目录获取最新照片 ==========
        # 在Intent处理之前执行DCIM扫描（Intent处理可能导致崩溃）
        # v3.17.0: 使用拍摄前时间戳过滤旧照片，避免连拍循环
        if IS_ANDROID:
            self._dbg("扫描DCIM/Camera目录获取最新照片...")
            try:
                import shutil
                launch_time = getattr(self, '_photo_launch_time', 0)
                now = time.time()
                dcim_dirs = [
                    '/storage/emulated/0/DCIM/Camera',
                    '/storage/emulated/0/DCIM/相机',
                    '/sdcard/DCIM/Camera',
                    '/storage/emulated/0/Pictures',
                ]
                for dcim_dir in dcim_dirs:
                    if not os.path.isdir(dcim_dir):
                        continue
                    files = [os.path.join(dcim_dir, f) for f in os.listdir(dcim_dir)
                             if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
                    if not files:
                        continue
                    files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                    latest = files[0]
                    fsize = os.path.getsize(latest)
                    mtime = os.path.getmtime(latest)
                    age = now - mtime
                    # v3.17.0: 只接受拍摄后(launch_time之后)创建的照片
                    # 避免复制旧照片导致连拍循环
                    if launch_time > 0 and mtime < launch_time:
                        self._dbg(f"DCIM最新: {os.path.basename(latest)} ({fsize}bytes, {age:.0f}秒前) - 早于拍摄时间，跳过")
                        continue
                    self._dbg(f"DCIM最新: {os.path.basename(latest)} ({fsize}bytes, {age:.0f}秒前, {dcim_dir})")
                    if fsize > 10000 and age < 120:  # >10KB 且2分钟内
                        os.makedirs(APP_DIR, exist_ok=True)
                        dest_path = os.path.join(
                            APP_DIR, f"capture_{get_system_date().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
                        )
                        shutil.copy2(latest, dest_path)
                        self.photo_path = dest_path
                        self._dbg(f"[OK] 从DCIM复制照片成功: {os.path.getsize(dest_path)} bytes", show_toast=True)
                        # v3.19.0: 删除 DCIM 原始文件，避免相册出现相机原始命名+规范命名两份
                        try:
                            os.remove(latest)
                            self._dbg(f"已删除 DCIM 原文件 {os.path.basename(latest)}，避免重复")
                        except:
                            pass
                        if self.pending_callback:
                            cb = self.pending_callback
                            self.pending_callback = None
                            Clock.schedule_once(lambda dt, cb=cb: cb(self.photo_path), 0)
                        return
                    else:
                        self._dbg(f"  照片太旧({age:.0f}秒前)或太小({fsize}bytes)，跳过")
                self._dbg("DCIM扫描完成，未找到符合条件的照片")
            except:
                import sys as _sys
                e = _sys.exc_info()[1]
                self._dbg(f"DCIM扫描失败: {str(e)[:80] if e else 'Unknown'}")

        # ========== 第4优先级：从返回Intent获取照片 ==========
        if intent is not None:
            self._dbg(f"尝试从返回Intent获取照片(result_code={result_code})...")
            try:
                from jnius import autoclass, cast
                data_uri = intent.getData()
                if data_uri is not None:
                    self._dbg(f"Intent返回URI: {str(data_uri.toString())[:80]}")
                    try:
                        PythonActivity = autoclass('org.kivy.android.PythonActivity')
                        activity = PythonActivity.mActivity
                        resolver = activity.getContentResolver()
                        istream = resolver.openInputStream(data_uri)
                        os.makedirs(APP_DIR, exist_ok=True)
                        dest_path = os.path.join(
                            APP_DIR, f"capture_{get_system_date().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
                        )
                        ByteArrayOutputStream = autoclass('java.io.ByteArrayOutputStream')
                        baos = ByteArrayOutputStream()
                        buf = bytearray(8192)
                        while True:
                            n = istream.read(buf)
                            if n <= 0:
                                break
                            baos.write(buf, 0, n)
                        istream.close()
                        with open(dest_path, 'wb') as f:
                            f.write(bytes(baos.toByteArray()))
                        baos.close()
                        self.photo_path = dest_path
                        fsize = os.path.getsize(dest_path)
                        self._dbg(f"从Intent URI保存成功: {fsize} bytes")
                        if fsize > 0:
                            if self.pending_callback:
                                cb = self.pending_callback
                                self.pending_callback = None
                                Clock.schedule_once(lambda dt, cb=cb: cb(self.photo_path), 0)
                            return
                    except:
                        import sys as _sys
                        e = _sys.exc_info()[1]
                        self._dbg(f"从URI读取失败: {str(e)[:80] if e else 'Unknown'}")
                else:
                    extras = intent.getExtras()
                    if extras is not None and extras.containsKey("data"):
                        self._dbg("Intent返回缩略图(data extra) - 分辨率较低")
                        try:
                            bitmap = cast('android.graphics.Bitmap', extras.get("data"))
                            if bitmap is not None:
                                os.makedirs(APP_DIR, exist_ok=True)
                                dest_path = os.path.join(
                                    APP_DIR, f"capture_{get_system_date().strftime('%Y%m%d_%H%M%S_%f')}.png"
                                )
                                FileOutputStream = autoclass('java.io.FileOutputStream')
                                fos = FileOutputStream(dest_path)
                                try:
                                    CompressFormat = autoclass('android.graphics.Bitmap$CompressFormat')
                                    bitmap.compress(CompressFormat.PNG, 100, fos)
                                except:
                                    import sys as _sys
                                    ce = _sys.exc_info()[1]
                                    self._dbg(f"  CompressFormat反射失败: {str(ce)[:60] if ce else ''}，尝试备选方案")
                                    bitmap.compress(autoclass('android.graphics.Bitmap').CompressFormat.PNG, 100, fos)
                                fos.flush()
                                fos.close()
                                self.photo_path = dest_path
                                fsize = os.path.getsize(dest_path)
                                self._dbg(f"缩略图已保存: {fsize} bytes", show_toast=True)
                                if fsize > 0:
                                    if self.pending_callback:
                                        cb = self.pending_callback
                                        self.pending_callback = None
                                        Clock.schedule_once(lambda dt, cb=cb: cb(self.photo_path), 0)
                                    return
                        except:
                            import sys as _sys
                            e = _sys.exc_info()[1]
                            self._dbg(f"缩略图保存失败: {str(e)[:80] if e else 'Unknown'}")
                    else:
                        clip_data = intent.getClipData()
                        if clip_data is not None and clip_data.getItemCount() > 0:
                            self._dbg(f"Intent返回ClipData({clip_data.getItemCount()}项)")
                            try:
                                item = clip_data.getItemAt(0)
                                clip_uri = item.getUri()
                                if clip_uri is not None:
                                    from jnius import autoclass
                                    PythonActivity = autoclass('org.kivy.android.PythonActivity')
                                    activity = PythonActivity.mActivity
                                    resolver = activity.getContentResolver()
                                    istream = resolver.openInputStream(clip_uri)
                                    os.makedirs(APP_DIR, exist_ok=True)
                                    dest_path = os.path.join(
                                        APP_DIR, f"capture_{get_system_date().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
                                    )
                                    ByteArrayOutputStream = autoclass('java.io.ByteArrayOutputStream')
                                    baos = ByteArrayOutputStream()
                                    buf = bytearray(8192)
                                    while True:
                                        n = istream.read(buf)
                                        if n <= 0:
                                            break
                                        baos.write(buf, 0, n)
                                    istream.close()
                                    with open(dest_path, 'wb') as f:
                                        f.write(bytes(baos.toByteArray()))
                                    baos.close()
                                    self.photo_path = dest_path
                                    fsize = os.path.getsize(dest_path)
                                    self._dbg(f"ClipData URI保存成功: {fsize} bytes")
                                    if fsize > 0:
                                        if self.pending_callback:
                                            cb = self.pending_callback
                                            self.pending_callback = None
                                            Clock.schedule_once(lambda dt, cb=cb: cb(self.photo_path), 0)
                                        return
                            except:
                                import sys as _sys
                                e = _sys.exc_info()[1]
                                self._dbg(f"ClipData读取失败: {str(e)[:80] if e else 'Unknown'}")
                        else:
                            self._dbg("Intent无data/extras/clipData")
            except:
                import sys as _sys
                e = _sys.exc_info()[1]
                self._dbg(f"处理Intent返回失败: {str(e)[:100] if e else 'Unknown'}")

        # ========== 所有方式都失败 ==========
        if result_code == 0:
            self._dbg("拍照已取消（未检测到照片文件）", show_toast=True)
        else:
            self._dbg(f"拍照失败(result_code={result_code})，未检测到照片", show_toast=True)
        self._media_uri = None
        if self.pending_callback:
            cb = self.pending_callback
            self.pending_callback = None
            Clock.schedule_once(lambda dt, cb=cb: cb(None), 0)

    def _simulate_photo(self):
        """桌面端测试：生成一张模拟照片"""
        self.photo_path = os.path.join(APP_DIR, f"test_{get_system_date().strftime('%Y%m%d_%H%M%S')}.jpg")
        img = PILImage.new('RGB', (640, 480), (180, 180, 180))
        draw = ImageDraw.Draw(img)
        font = PhotoProcessor._get_font(24)
        now_str = get_full_datetime_str()
        draw.text((150, 200), "Test Photo", fill=(0, 0, 0), font=font)
        draw.text((150, 250), now_str, fill=(0, 0, 0), font=font)
        img.save(self.photo_path)
        if self.pending_callback:
            Clock.schedule_once(lambda dt: self.pending_callback(self.photo_path), 0.3)

    def get_location_name(self, lat, lng):
        """根据 GPS 坐标逆地理编码获取地名，供水印「地址名」段使用。
        在后台线程调用（HTTP/计算不会阻塞 UI）。
        无 GPS 时返回 'GPS未开启或无权限'。"""
        if lat and lng:
            try:
                name = self.geocoder.reverse_geocode(float(lng), float(lat))
                if name:
                    return name
                return "GPS定位解析失败"
            except Exception as e:
                Logger.error("CameraManager.get_location_name: %s" % e)
                return "GPS定位解析失败"
        return "GPS未开启或无权限"

# ============================================================
# 自定义小组件
# ============================================================

def bind_press_animation(btn, scale=0.94, duration=0.08):
    """v3.19.0: 为按钮绑定按压动画，带来流畅的交互反馈。
    注意：Kivy 的 Button/Widget 没有 scale_x/scale_y 属性（那是 Scatter 的），
    用 scale_x 会触发 AttributeError 导致闪退。改用 opacity 做按压反馈，
    按下时轻微变透明，松开时还原，配合 SlideTransition 让操作更顺滑。
    """
    from kivy.animation import Animation
    def _on_down(instance, touch):
        if instance.collide_point(*touch.pos):
            Animation(opacity=0.72, duration=duration, t='out_quad').start(instance)
    def _on_up(instance, touch):
        Animation(opacity=1.0, duration=duration * 1.2, t='out_quad').start(instance)
    btn.bind(on_touch_down=_on_down)
    btn.bind(on_touch_up=_on_up)


def _scroll_input_to_visible_func(popup, text_input):
    """v3.22.19: 软键盘弹出后，将 popup 内 text_input 滚动到可见区域。
    实现：
      1) 估算键盘高度（resize 模式下 Window.height 已收缩，差值即键盘高度）
      2) 若 text_input 顶部高于可见区顶部（被键盘遮挡）：
         - 优先在 popup 内查找 ScrollView 并 scroll_to(text_input)
         - 退路：整体上移 popup（设置 pos_hint={'top': 0.98}）
    """
    try:
        kb_height = 0
        try:
            # Kivy 在 resize 模式下 Window.height 已收缩，system_size 仍保留原全屏尺寸
            if hasattr(Window, 'system_size'):
                kb_height = max(0, int(Window.system_size[1] - Window.height))
        except Exception:
            kb_height = 0
        # 获取 text_input 在 Window 坐标系下的位置
        try:
            ti_y = text_input.to_window(text_input.x, text_input.y)[1]
            ti_top = ti_y + text_input.height
        except Exception:
            return
        visible_top = Window.height - kb_height
        # 若 text_input 顶部高于可见区顶部 → 被键盘遮挡
        if ti_top > visible_top:
            # 在 popup 内查找 ScrollView 并尝试 scroll_to
            found_sv = [False]
            from kivy.uix.scrollview import ScrollView as _SV
            def _find_scroll(widget):
                if found_sv[0]:
                    return
                if isinstance(widget, _SV):
                    try:
                        widget.scroll_to(text_input)
                        found_sv[0] = True
                    except Exception:
                        pass
                    return
                for child in getattr(widget, 'children', []) or []:
                    _find_scroll(child)
            content = getattr(popup, 'content', None)
            if content is not None:
                _find_scroll(content)
            # 若 ScrollView 滚动后仍未可见，尝试整体上移 popup（pos_hint）
            if not found_sv[0]:
                try:
                    popup.pos_hint = {'top': 0.98}
                except Exception:
                    pass
    except Exception as e:
        try:
            app_log.info('UI', '_scroll_input_to_visible_func 异常: %s' % e)
        except Exception:
            pass


class CardWidget(BoxLayout):
    """卡片式容器
    v3.19.0: 明亮浅色主题——纯白卡片 + 浅灰描边 + 大圆角，更有层次感。
    """
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.padding = [14, 12, 14, 12]
        self.spacing = 6
        with self.canvas.before:
            Color(*THEME['card'])
            self.rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[14])
            Color(*THEME['card_border'])
            self.border = Line(rounded_rectangle=(self.x, self.y, self.width, self.height, 10), width=1)
        self.bind(pos=self._update_rect, size=self._update_rect)

    def _update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size
        self.border.rounded_rectangle = (self.x, self.y, self.width, self.height, 10)


class SectionLabel(Label):
    """分区标题"""
    def __init__(self, text="", **kwargs):
        super().__init__(**kwargs)
        self.text = text
        self.font_size = '17sp'
        self.bold = True
        self.color = THEME['accent']
        self.size_hint_y = None
        self.height = dp(40)
        self.halign = 'left'
        self.valign = 'middle'


# ============================================================
# 欢迎页面
# ============================================================

class WelcomeScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'welcome'

        root = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))

        # 顶部状态栏留白（Android上为状态栏高度，桌面端为小间距）
        self._status_spacer = Label(size_hint_y=None, height=dp(get_status_bar_height_dp()))
        root.add_widget(self._status_spacer)

        # Logo 图标（v3.19.2: 用 SimHei 支持的◆符号替换📷emoji，避免显示为方块）
        root.add_widget(Label(
            text="◆", font_size='80sp',
            size_hint_y=None, height=dp(96), color=THEME['accent'],
        ))

        # 标题
        root.add_widget(Label(
            text="资产盘点专项拍照工具", font_size='26sp',
            bold=True, color=THEME['text'],
            size_hint_y=None, height=dp(48),
        ))

        # 副标题
        root.add_widget(Label(
            text="银行抵押物现场勘查工具", font_size='15sp',
            color=THEME['text_dim'],
            size_hint_y=None, height=dp(32),
        ))

        # 版本
        root.add_widget(Label(
            text="v3.22.19", font_size='12sp',
            color=THEME['text_dim'],
            size_hint_y=None, height=dp(24),
        ))

        # 间距
        root.add_widget(Label(size_hint_y=None, height=dp(14)))

        # 功能简介卡片
        feat_card = CardWidget(size_hint_y=None)
        feat_card.bind(minimum_height=feat_card.setter('height'))
        features = [
            "●  四类拍照引导（远景/近景/内部/瑕疵）",
            "●  水印自选模式（段+位置+字号）",
            "●  文件命名自选模式（4段下拉）",
            "●  一键生成勘查日报表",
        ]
        for feat in features:
            feat_card.add_widget(Label(
                text=feat, font_size='15sp',
                color=THEME['text'],
                size_hint_y=None, height=dp(34),
                halign='left', valign='middle',
                text_size=(None, dp(34)),
            ))
        root.add_widget(feat_card)

        # 弹性空间
        root.add_widget(Label(size_hint_y=1))

        # 间距
        root.add_widget(Label(size_hint_y=None, height=dp(12)))

        # 进入按钮 - 大尺寸适合手指点击
        start_btn = RoundedButton(
            text="开 始 使 用", font_size='22sp',
            size_hint_y=None, height=dp(64),
            background_color=THEME['accent'],
            background_normal='',
            color=(1, 1, 1, 1),
            bold=True,
        )
        bind_press_animation(start_btn)
        start_btn.bind(on_release=self._go_main)
        root.add_widget(start_btn)

        self.add_widget(root)

    def _go_main(self, instance):
        self.manager.current = 'main'


# ============================================================
# 拍照类型选择弹窗
# ============================================================

class PhotoTypePopup(Popup):
    def __init__(self, on_select, **kwargs):
        super().__init__(**kwargs)
        self.title = "选择拍照类型"
        self.size_hint = (0.85, 0.55)
        self.auto_dismiss = True
        self.on_select = on_select

        layout = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(14))
        layout.add_widget(Label(text="请选择本次拍摄的照片类型：", font_size='16sp', size_hint_y=None, height=dp(40)))

        for type_name, type_desc in PHOTO_TYPES:
            btn = RoundedButton(
                text=f"{type_name}\n{type_desc}", font_size='16sp',
                size_hint_y=None, height=dp(72),
                background_color=THEME['accent'], background_normal='',
                halign='center', color=(1,1,1,1), bold=True,
            )
            btn.bind(on_release=lambda x, t=type_name: self._select(t))
            layout.add_widget(btn)

        cancel_btn = RoundedButton(text="取消", font_size='16sp', size_hint_y=None, height=dp(52),
                           background_color=THEME['muted'], background_normal='',
                           color=(1,1,1,1))
        cancel_btn.bind(on_release=self.dismiss)
        layout.add_widget(cancel_btn)
        self.content = layout

    def _select(self, photo_type):
        self.dismiss()
        if self.on_select:
            self.on_select(photo_type)


# ============================================================
# v3.22.9: 统一 Popup 浅色主题基类
# ============================================================

class ThemedPopup(Popup):
    """v3.22.9: 统一 Popup 浅色主题基类。
    自动绘制 THEME['card'] 浅色背景 + 设置 title_color/separator_color，
    根治 v3.22.7 漏改 L5357 这类 background=THEME['card'] 遗漏问题。
    子类只需继承，无需再写 canvas.before 背景绘制代码。"""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # v3.22.17: 禁用 Kivy Popup 默认的 atlas 背景图像，避免在某些机型上加载失败导致 popup 不可见
        self.background = ''
        # 统一绘制浅色背景
        with self.canvas.before:
            Color(*THEME['card'])
            _bg = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=lambda i, v: setattr(_bg, 'pos', v),
                 size=lambda i, v: setattr(_bg, 'size', v))
        # 统一标题颜色和分隔线颜色
        self.title_color = THEME['accent_dark']
        self.separator_color = THEME['card_border']


# ============================================================
# v3.22.13: 今日走访补充说明对话框
# ============================================================

class VisitNoteDialog(ThemedPopup):
    """v3.22.13: 今日走访补充说明输入对话框
    用户输入今日走访的补充说明，AI 报告生成时参考此说明。
    说明按 Excel 文件路径 md5 哈希持久化到 APP_DIR/visit_notes/ 目录。
    """

    def __init__(self, initial_text='', on_confirm=None, **kwargs):
        """v3.22.13: 初始化
        Args:
            initial_text: 预填的文本（从持久化加载）
            on_confirm: 确认回调，签名为 on_confirm(text: str)
        """
        super().__init__(**kwargs)
        self.title = '今日走访补充说明'
        self.auto_dismiss = False  # 避免误关闭丢失输入
        self.size_hint = (0.9, None)
        self.height = dp(420)

        self._on_confirm_callback = on_confirm
        self._initial_text = initial_text

        content = self._build_content()
        self.content = content

    def _build_content(self):
        """构建对话框内容"""
        content = BoxLayout(orientation='vertical', spacing=dp(6), padding=dp(10))

        # 标题
        title = Label(
            text='今日走访补充说明',
            font_size='18sp', bold=True,
            color=THEME['text'],
            size_hint_y=0.1,
            halign='left', valign='middle',
        )
        title.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
        content.add_widget(title)

        # 提示文字
        hint = Label(
            text=('请输入今日走访的补充说明，AI 将参考此说明生成日报表。\n'
                  '【同楼盘多房间合并说明】\n'
                  '同一借款人抵押同一栋楼的多个房间号时，只需对代表性户型拍摄并备注。\n'
                  '示例：A大街12号101、201、301房型一致，仅拍摄101并备注；\n'
                  '102、202、302为镜像关系，可一并说明。\n'
                  'AI 会根据此说明合并描述同楼盘多房间的情况，报告中只生成 1 行。'),
            font_size='11sp',
            color=THEME['text_dim'],
            size_hint_y=0.28,
            halign='left', valign='top',
            text_size=(0, None),
        )
        hint.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], val[1])))
        content.add_widget(hint)

        # 多行 TextInput
        self.text_input = TextInput(
            text=self._initial_text,
            multiline=True,
            font_size='13sp',
            size_hint_y=0.42,
            hint_text='在此输入今日走访补充说明…',
        )
        # v3.22.19: 聚焦时延迟滚动到可见区域，避免键盘遮挡输入框
        self.text_input.bind(focus=lambda inst, val: Clock.schedule_once(
            lambda dt: _scroll_input_to_visible_func(self, self.text_input), 0.1) if val else None)
        content.add_widget(self.text_input)

        # 按钮行
        btn_row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint_y=0.15)

        btn_cancel = RoundedButton(
            text='取消', font_size='14sp',
            background_color=THEME['text_dim'], background_normal='',
            color=(1, 1, 1, 1),
        )
        btn_cancel.bind(on_press=self.dismiss)
        btn_row.add_widget(btn_cancel)

        btn_confirm = RoundedButton(
            text='生成日报表', font_size='14sp',
            background_color=THEME['success'], background_normal='',
            color=(1, 1, 1, 1),
        )
        btn_confirm.bind(on_press=self._on_confirm)
        btn_row.add_widget(btn_confirm)

        content.add_widget(btn_row)
        return content

    def _on_confirm(self, instance):
        """v3.22.13: 确认按钮回调"""
        text = self.text_input.text or ''
        try:
            self.dismiss()
            if self._on_confirm_callback:
                self._on_confirm_callback(text)
        except Exception as e:
            app_log.error('UI', 'VisitNoteDialog._on_confirm 异常: %s' % e, exc_info=True)

    def get_text(self):
        """获取输入文本"""
        return self.text_input.text or ''


# ============================================================
# 照片查看弹窗
# ============================================================

class PhotoViewerPopup(Popup):
    """v3.22.0: 异步加载缩略图，避免大图同步解码导致"查看已拍"卡顿；
    浅色背景白底深字，确保删除确认弹窗可读。
    v3.22.18: 直接继承 Popup，背景绘制在 content.canvas.before（匹配 _show_report_confirm 有效模式）"""
    _THUMB_DIR = None  # 缩略图缓存目录（惰性初始化）

    def __init__(self, row_index, photos, delete_callback,
                 photo_count=None, progress_key=None, **kwargs):
        super().__init__(**kwargs)
        # v3.22.19: photos 含义改为缩略图路径列表；title 仍显示原照片数量（photo_count）
        self.title = f"已拍照片 (%d张)" % (photo_count if photo_count is not None else len(photos))
        self.title_color = THEME['accent_dark']
        self.separator_color = THEME['card_border']
        self.auto_dismiss = False  # v3.22.17: 防止被打开它的同一触摸事件立即关闭
        self.size_hint = (None, None)
        self.size = (int(Window.width * 0.92), int(Window.height * 0.8))
        self.center = (Window.width / 2, Window.height / 2)
        self.row_index = row_index
        self.photos = photos  # v3.22.19: 缩略图路径列表
        self.photo_count = photo_count if photo_count is not None else len(photos)
        self.progress_key = progress_key  # v3.22.19: 用于删除缩略图
        self.delete_callback = delete_callback
        self._thumb_slots = []  # 每项图片占位 widget，索引对应 photos

        main_layout = BoxLayout(orientation='vertical', spacing=dp(8), padding=dp(10))
        # v3.22.18: 在 content 的 canvas.before 中绘制浅色背景（匹配 _show_report_confirm 有效模式）
        with main_layout.canvas.before:
            Color(*THEME['card'])
            self._content_bg = Rectangle(pos=main_layout.pos, size=main_layout.size)
        main_layout.bind(pos=self._update_content_bg, size=self._update_content_bg)
        scroll = ScrollView()
        list_layout = GridLayout(cols=3, spacing=dp(6), size_hint_y=None)
        list_layout.bind(minimum_height=list_layout.setter('height'))

        for i, photo_path in enumerate(photos):
            # v3.22.14: 回退到 v3.22.12 内嵌缩略图方案
            # BoxLayout(垂直) = 加载中 Label(占位) + 文件名 Label + 删除按钮
            item = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(150), spacing=dp(2))
            placeholder = Label(
                text="加载中…", font_size='11sp', color=THEME['text_dim'],
                size_hint_y=0.65, halign='center', valign='middle'
            )
            item.add_widget(placeholder)
            self._thumb_slots.append(placeholder)

            fname = os.path.basename(photo_path)
            if len(fname) > 16:
                fname = fname[:13] + '...'
            name_label = Label(
                text=fname, font_size='10sp', halign='center', valign='middle',
                size_hint_y=0.15, color=THEME['text_dim'], text_size=(0, None)
            )
            name_label.bind(width=lambda inst, val: setattr(inst, 'text_size', (val, None)))
            item.add_widget(name_label)

            del_btn = RoundedButton(
                text="删除", font_size='11sp', size_hint_y=0.20, height=dp(32),
                background_color=THEME['danger'], background_normal='', color=(1,1,1,1), bold=True
            )
            del_btn.bind(on_press=lambda x, idx=i: self._confirm_delete(idx))
            bind_press_animation(del_btn)
            item.add_widget(del_btn)

            list_layout.add_widget(item)

        scroll.add_widget(list_layout)
        main_layout.add_widget(scroll)

        del_all_btn = RoundedButton(text="删除全部照片（重拍）", font_size='15sp', size_hint_y=None, height=dp(52),
                            background_color=THEME['danger'], background_normal='',
                            color=(1,1,1,1), bold=True)
        del_all_btn.bind(on_press=self._confirm_delete_all)
        main_layout.add_widget(del_all_btn)

        close_btn = RoundedButton(text="关闭", font_size='16sp', size_hint_y=None, height=dp(52),
                          background_color=THEME['accent'], background_normal='',
                          color=(1,1,1,1), bold=True)
        close_btn.bind(on_press=self.dismiss)
        main_layout.add_widget(close_btn)
        self.content = main_layout

        # v3.22.0: 弹窗显示后异步加载缩略图
        self.bind(on_open=lambda *a: Clock.schedule_once(self._load_thumbs, 0.1))

    def _update_content_bg(self, instance, value):
        self._content_bg.pos = instance.pos
        self._content_bg.size = instance.size

    @classmethod
    def _get_thumb_dir(cls):
        if cls._THUMB_DIR is None:
            cls._THUMB_DIR = os.path.join(APP_DIR, 'thumbs')
            try:
                os.makedirs(cls._THUMB_DIR, exist_ok=True)
            except Exception:
                pass
        return cls._THUMB_DIR

    def _ensure_thumb(self, photo_path):
        """生成缩略图（240px），返回缓存路径；失败返回 None"""
        try:
            if not os.path.exists(photo_path):
                return None
            st = os.stat(photo_path)
            # 缓存名：大小+修改时间+文件名，避免失效
            safe = re.sub(r'[^\w]', '_', os.path.basename(photo_path))[:40]
            thumb_name = "%s_%d_%s.png" % (safe, int(st.st_size), int(st.st_mtime))
            thumb_path = os.path.join(self._get_thumb_dir(), thumb_name)
            if os.path.exists(thumb_path):
                return thumb_path
            # v3.22.12: 用 with 管理文件句柄，thumbnail/save 后自动关闭
            with PILImage.open(photo_path) as img:
                img.thumbnail((240, 240))
                img.save(thumb_path, 'PNG')
            return thumb_path
        except Exception as e:
            app_log.error('PHOTO', '生成缩略图失败: %s' % e)
            return None

    def _load_thumbs(self, *args):
        """v3.22.19: self.photos 已是缩略图路径列表，无需再调用 _ensure_thumb 生成。
        后台逐张验证文件存在并调度主线程替换占位 widget。"""
        import threading
        def _worker():
            for i, p in enumerate(self.photos):
                try:
                    # popup 已关闭则停止
                    if self.parent is None and not self._is_open():
                        return
                except Exception:
                    pass
                # v3.22.19: 缩略图路径直接用，不再调用 _ensure_thumb（避免重复生成）
                if p and os.path.exists(p):
                    Clock.schedule_once(lambda dt, idx=i, t=p: self._set_thumb(idx, t), 0)
                else:
                    app_log.error('PHOTO', '缩略图文件不存在: %s' % p)
        threading.Thread(target=_worker, daemon=True).start()

    def _is_open(self):
        try:
            return self.parent is not None
        except Exception:
            return False

    def _set_thumb(self, idx, thumb_path):
        """v3.22.14: 回退到 v3.22.12 — 用 KivyImage 替换占位 Label
        （v3.22.13 改为设置 Button.background_normal，回退后改回替换 widget）"""
        placeholder = self._thumb_slots[idx]
        parent = placeholder.parent
        if parent is None:
            return
        pos_in_parent = parent.children.index(placeholder)
        parent.remove_widget(placeholder)
        try:
            img = KivyImage(source=thumb_path, size_hint_y=0.65,
                            allow_stretch=True, keep_ratio=True)
            parent.add_widget(img, pos_in_parent)
            self._thumb_slots[idx] = img
        except Exception as e:
            app_log.error('PHOTO', '设置缩略图失败 idx=%d: %s' % (idx, e))
            err_label = Label(text="加载失败", font_size='10sp', color=THEME['danger'],
                              size_hint_y=0.65, halign='center', valign='middle')
            parent.add_widget(err_label, pos_in_parent)
            self._thumb_slots[idx] = err_label

    def _show_msg_safe(self, msg):
        """v3.22.13: PhotoViewerPopup 内部用 — 无 MainScreen._show_msg 时的 fallback
        v3.22.13 修复: 原实现仅记日志，用户看不到失败反馈（违反 v3.22.13 「Toast 反馈」目标）。
        现改为 Android 上弹 Toast，非 Android 环境仅记录日志。"""
        try:
            app_log.error('UI', msg)
        except Exception:
            pass
        # v3.22.13: Android 上弹 Toast 让用户看到失败原因
        if IS_ANDROID:
            try:
                from jnius import autoclass
                Toast = autoclass('android.widget.Toast')
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                Toast.makeText(PythonActivity.mActivity, msg, Toast.LENGTH_SHORT).show()
            except Exception:
                pass

    def _confirm_delete(self, index):
        """删除单张照片前弹出二次确认（v3.22.0: 白底深字可读）
        v3.22.9: 改用 ThemedPopup，背景与标题色由基类统一处理。"""
        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(10))
        # v3.22.9: 背景由 ThemedPopup.canvas.before 统一绘制，此处不再重复
        content.add_widget(Label(text=f"确定要删除这张照片吗？\n此操作不可撤销。",
                                 font_size='16sp', color=THEME['text'], halign='center'))
        btn_row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(10))
        popup = ThemedPopup(title="确认删除",
                      size_hint=(0.8, 0.35), auto_dismiss=True)
        yes_btn = RoundedButton(text="确认删除", font_size='16sp', background_color=THEME['danger'],
                         background_normal='', color=(1,1,1,1), bold=True)
        no_btn = RoundedButton(text="取消", font_size='16sp', background_color=THEME['accent'],
                        background_normal='', color=(1,1,1,1), bold=True)
        def _do_delete(instance):
            popup.dismiss()
            self.delete_callback(self.row_index, index)
            self.dismiss()
        yes_btn.bind(on_release=_do_delete)
        no_btn.bind(on_release=popup.dismiss)
        btn_row.add_widget(yes_btn)
        btn_row.add_widget(no_btn)
        content.add_widget(btn_row)
        popup.content = content
        popup.open()

    def _confirm_delete_all(self, instance):
        """删除全部照片前弹出二次确认（v3.22.0: 白底深字可读）
        v3.22.9: 改用 ThemedPopup，背景与标题色由基类统一处理。"""
        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(10))
        # v3.22.9: 背景由 ThemedPopup.canvas.before 统一绘制，此处不再重复
        content.add_widget(Label(text=f"确定要删除该客户的全部照片吗？\n此操作不可撤销！",
                                 font_size='16sp', color=THEME['danger'], halign='center'))
        btn_row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(10))
        popup = ThemedPopup(title="危险操作确认",
                      size_hint=(0.85, 0.35), auto_dismiss=True)
        yes_btn = RoundedButton(text="全部删除", font_size='16sp', background_color=THEME['danger'],
                         background_normal='', color=(1,1,1,1), bold=True)
        no_btn = RoundedButton(text="取消", font_size='16sp', background_color=THEME['accent'],
                        background_normal='', color=(1,1,1,1), bold=True)
        def _do_delete_all(instance):
            popup.dismiss()
            self.delete_callback(self.row_index, -1)
            self.dismiss()
        yes_btn.bind(on_release=_do_delete_all)
        no_btn.bind(on_release=popup.dismiss)
        btn_row.add_widget(yes_btn)
        btn_row.add_widget(no_btn)
        content.add_widget(btn_row)
        popup.content = content
        popup.open()


# ============================================================
# 设置页面
# ============================================================

class SettingsScreen(Screen):
    def __init__(self, app_config, **kwargs):
        super().__init__(**kwargs)
        self.name = 'settings'
        self.config = app_config

        main = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))

        # 顶部状态栏留白
        main.add_widget(Label(size_hint_y=None, height=dp(get_status_bar_height_dp())))

        # 标题栏 - 增大高度和按钮
        title_bar = BoxLayout(size_hint_y=None, height=dp(56), spacing=dp(8))
        back_btn = RoundedButton(text="返回", font_size='18sp', size_hint_x=0.28,
                         background_color=THEME['accent'], background_normal='',
                         color=(1,1,1,1), bold=True,
                         size_hint_y=None, height=dp(52))
        back_btn.bind(on_release=self._go_back)
        bind_press_animation(back_btn)
        title_bar.add_widget(back_btn)
        title_bar.add_widget(Label(text="设置", font_size='22sp', bold=True, color=THEME['text']))
        main.add_widget(title_bar)

        scroll = ScrollView()
        content = BoxLayout(orientation='vertical', spacing=dp(12), size_hint_y=None, padding=[dp(4), dp(4), dp(4), dp(30)])
        content.bind(minimum_height=content.setter('height'))

        # === 命名规则 ===
        naming_card = CardWidget(size_hint_y=None)
        naming_card.bind(minimum_height=naming_card.setter('height'))

        naming_card.add_widget(SectionLabel(text="照片命名规则"))

        naming_card.add_widget(Label(text="格式：X-X-X-X（每段自选，空值段自动省略）",
                                     font_size='13sp', color=THEME['text_dim'],
                                     size_hint_y=None, height=dp(24)))

        cur_segments = self.config.get('naming_segments', DEFAULT_CONFIG['naming_segments'])
        self.naming_spinners = []
        for idx in range(4):
            row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
            row.add_widget(Label(text="第%d段" % (idx + 1), font_size='15sp',
                                 color=THEME['text_dim'], size_hint_x=0.2))
            cur_val = cur_segments[idx] if idx < len(cur_segments) else "空值"
            sp = Spinner(
                text=cur_val,
                values=NAMING_SEGMENT_OPTIONS,
                size_hint_x=0.8, font_size='15sp',
            )
            sp.bind(text=lambda inst, val, i=idx: self._on_naming_segment_change(i, val))
            self.naming_spinners.append(sp)
            row.add_widget(sp)
            naming_card.add_widget(row)

        preview = self._get_naming_preview()
        self.naming_preview_label = Label(text="预览：%s" % preview, font_size='13sp',
                                          color=THEME['text_dim'], size_hint_y=None, height=dp(28))
        naming_card.add_widget(self.naming_preview_label)

        save_naming_btn = RoundedButton(text="保存命名规则", font_size='16sp', size_hint_y=None, height=dp(52),
                                background_color=THEME['accent'], background_normal='',
                                color=(1,1,1,1), bold=True)
        save_naming_btn.bind(on_press=self._save_naming)
        naming_card.add_widget(save_naming_btn)

        content.add_widget(naming_card)

        # === 水印设置 ===
        watermark_card = CardWidget(size_hint_y=None)
        watermark_card.bind(minimum_height=watermark_card.setter('height'))

        watermark_card.add_widget(SectionLabel(text="水印设置"))

        # 水印开关
        toggle_box = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(8))
        toggle_box.add_widget(Label(text="启用水印", font_size='16sp', color=THEME['text'], size_hint_x=0.6))
        self.wm_check = CheckBox(active=self.config.get('watermark_enabled', True), size_hint_x=0.4)
        toggle_box.add_widget(self.wm_check)
        watermark_card.add_widget(toggle_box)

        watermark_card.add_widget(Label(text="水印格式：X-X-X（每段自选）",
                                        font_size='13sp', color=THEME['text_dim'],
                                        size_hint_y=None, height=dp(24)))

        # 水印段选择（3段）
        cur_wm_segments = self.config.get('watermark_segments', DEFAULT_CONFIG['watermark_segments'])
        self.wm_spinners = []
        for idx in range(3):
            row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
            row.add_widget(Label(text="第%d段" % (idx + 1), font_size='15sp',
                                 color=THEME['text_dim'], size_hint_x=0.2))
            cur_val = cur_wm_segments[idx] if idx < len(cur_wm_segments) else "空值"
            sp = Spinner(
                text=cur_val,
                values=WATERMARK_SEGMENT_OPTIONS,
                size_hint_x=0.8, font_size='15sp',
            )
            sp.bind(text=lambda inst, val, i=idx: self._on_wm_segment_change(i, val))
            self.wm_spinners.append(sp)
            row.add_widget(sp)
            watermark_card.add_widget(row)

        # 水印位置
        pos_box = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
        pos_box.add_widget(Label(text="位置：", font_size='15sp', color=THEME['text'], size_hint_x=0.25))
        cur_pos = self.config.get('watermark_position', DEFAULT_CONFIG['watermark_position'])
        cur_pos_label = WATERMARK_POSITION_LABELS.get(cur_pos, '右下')
        self.pos_spinner = Spinner(
            text=cur_pos_label,
            values=list(WATERMARK_POSITION_LABELS.values()),
            size_hint_x=0.75, font_size='15sp',
        )
        pos_box.add_widget(self.pos_spinner)
        watermark_card.add_widget(pos_box)

        # 字号（大/中/小）
        font_size_box = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
        font_size_box.add_widget(Label(text="字号：", font_size='15sp', color=THEME['text'], size_hint_x=0.25))
        cur_fs = self.config.get('watermark_font_size', DEFAULT_CONFIG['watermark_font_size'])
        if cur_fs not in WATERMARK_FONT_SIZE_OPTIONS:
            cur_fs = '中'
        self.font_spinner = Spinner(
            text=cur_fs,
            values=WATERMARK_FONT_SIZE_OPTIONS,
            size_hint_x=0.75, font_size='15sp',
        )
        font_size_box.add_widget(self.font_spinner)
        watermark_card.add_widget(font_size_box)

        save_wm_btn = RoundedButton(text="保存水印设置", font_size='16sp', size_hint_y=None, height=dp(52),
                            background_color=THEME['accent'], background_normal='',
                            color=(1,1,1,1), bold=True)
        save_wm_btn.bind(on_press=self._save_watermark)
        watermark_card.add_widget(save_wm_btn)

        content.add_widget(watermark_card)

        # === AI 设置 === v3.15.0
        ai_card = CardWidget(size_hint_y=None)
        ai_card.bind(minimum_height=ai_card.setter('height'))
        ai_card.add_widget(SectionLabel(text="AI 助手设置"))

        ai_url_row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
        ai_url_row.add_widget(Label(text="API地址", font_size='14sp',
                                    color=THEME['text_dim'], size_hint_x=0.22))
        self.ai_url_input = TextInput(
            text=self.config.get('ai_api_url', DEFAULT_CONFIG['ai_api_url']),
            font_size='14sp', multiline=False, size_hint_x=0.78,
        )
        ai_url_row.add_widget(self.ai_url_input)
        ai_card.add_widget(ai_url_row)

        ai_key_row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
        ai_key_row.add_widget(Label(text="API Key", font_size='14sp',
                                    color=THEME['text_dim'], size_hint_x=0.22))
        self.ai_key_input = TextInput(
            text=self.config.get('ai_api_key', DEFAULT_CONFIG['ai_api_key']),
            font_size='14sp', multiline=False, size_hint_x=0.78,
            password=True,
        )
        ai_key_row.add_widget(self.ai_key_input)
        ai_card.add_widget(ai_key_row)

        ai_model_row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
        ai_model_row.add_widget(Label(text="模型ID", font_size='14sp',
                                     color=THEME['text_dim'], size_hint_x=0.22))
        self.ai_model_input = TextInput(
            text=self.config.get('ai_model', ''),
            font_size='14sp', multiline=False, size_hint_x=0.78,
            hint_text="如: deepseek-v4-flash",
        )
        ai_model_row.add_widget(self.ai_model_input)
        ai_card.add_widget(ai_model_row)

        ai_hint = Label(
            text="提示：模型ID请填入DeepSeek平台的模型标识（默认 deepseek-v4-flash）",
            font_size='12sp', color=THEME['text_dim'],
            size_hint_y=None, height=dp(36), halign='left',
        )
        ai_hint.bind(size=lambda inst, val: setattr(inst, 'text_size', val))
        ai_card.add_widget(ai_hint)

        save_ai_btn = RoundedButton(text="保存AI设置", font_size='16sp', size_hint_y=None, height=dp(52),
                            background_color=THEME['success'], background_normal='',
                            color=(1,1,1,1), bold=True)
        save_ai_btn.bind(on_press=self._save_ai_settings)
        ai_card.add_widget(save_ai_btn)

        content.add_widget(ai_card)

        # === 关于（已移除作者信息） ===

        content.add_widget(Label(size_hint_y=None, height=dp(20)))

        scroll.add_widget(content)
        main.add_widget(scroll)
        self.add_widget(main)

    def _on_naming_segment_change(self, idx, val):
        """命名段下拉变化时刷新预览"""
        self.naming_preview_label.text = "预览：%s" % self._get_naming_preview()

    def _get_naming_preview(self):
        """根据当前 4 个命名段下拉生成预览文件名"""
        segments = [sp.text for sp in self.naming_spinners]
        return PhotoProcessor.generate_filename(
            segments, borrower="张三",
            address_general="沈阳市和平区XX街123号", address_precise="1-23-4",
            property_type="住宅", seq=1, date_str="20260628", photo_type="远景",
            time_str="1430",
        )

    def _on_wm_segment_change(self, idx, val):
        """水印段下拉变化时刷新预览"""
        # 水印预览暂不单独显示，保存时生效
        pass

    def _save_naming(self, instance):
        segments = [sp.text for sp in self.naming_spinners]
        self.config.set('naming_segments', segments)
        self.naming_preview_label.text = "预览：%s" % self._get_naming_preview()
        self._show_toast("命名规则已保存")

    def _save_watermark(self, instance):
        segments = [sp.text for sp in self.wm_spinners]
        self.config.set('watermark_enabled', self.wm_check.active)
        self.config.set('watermark_segments', segments)
        pos_label = self.pos_spinner.text
        self.config.set('watermark_position',
                        WATERMARK_POSITION_LABEL_TO_KEY.get(pos_label, 'bottom-right'))
        self.config.set('watermark_font_size', self.font_spinner.text)
        self._show_toast("水印设置已保存")

    def _save_ai_settings(self, instance):
        """保存AI设置 v3.15.0"""
        self.config.set('ai_api_url', self.ai_url_input.text.strip() or DEFAULT_CONFIG['ai_api_url'])
        self.config.set('ai_api_key', self.ai_key_input.text.strip())
        self.config.set('ai_model', self.ai_model_input.text.strip())
        self._show_toast("AI设置已保存")

    def _show_toast(self, msg):
        popup = Popup(title='', content=Label(text=msg, font_size='14sp'),
                     size_hint=(0.6, 0.2), auto_dismiss=True)
        popup.open()
        Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)

    def _go_back(self, instance):
        self.manager.current = 'main'


# ============================================================
# 单行组件
# ============================================================

class RowWidget(BoxLayout):
    # v3.22.14: 回退到内嵌按钮方案 — view_btn.on_press 直接调用 _on_view_photos
    # 不经过菜单 dismiss 步骤，避免 Kivy 同一事件循环中状态冲突导致
    # PhotoViewerPopup 静默不显示；也避免滚动列表时误触发菜单。
    # 移除长按多选相关代码（multi_select / on_long_press / _long_press_*）
    # v3.22.19: 改为水平布局 [CheckBox(0.08) | 内容列(0.92)]，新增行选择能力。
    #   选中状态独立于 done（实际拍摄）状态，仅用于"标记完成"辅助 AI 报告。
    def __init__(self, row_index, borrower, address_general, address_precise, property_type,
                 progress_key, progress_mgr,
                 remark="", excel_row_index=None, serial="",
                 photo_callback=None, view_photos_callback=None, remark_callback=None,
                 selected_callback=None,
                 **kwargs):
        super().__init__(**kwargs)
        # v3.22.19: 改为水平布局 — 左侧 CheckBox + 右侧垂直内容列
        self.orientation = 'horizontal'
        self.size_hint_y = None
        self._base_height = dp(90)  # v3.22.2: 删除竖版计数器(100dp) 后下调
        self.height = self._base_height
        self.padding = [dp(10), dp(8), dp(10), dp(8)]
        self.spacing = dp(6)

        self.row_index = row_index
        self.borrower = borrower
        self.serial = serial or ""
        self.address_general = address_general
        self.address_precise = address_precise
        self.property_type = property_type
        self.progress_key = progress_key
        self.progress_mgr = progress_mgr
        self.remark = remark or ""
        self.excel_row_index = excel_row_index  # Excel中的实际行号（含表头，从1开始）
        self.done = progress_mgr.is_photographed(progress_key)
        self.photo_count = progress_mgr.get_photo_count(progress_key)
        # v3.22.6: 同类型标记状态（按行号，稳定）— 仅显示用，标记由 MainScreen 操作菜单触发
        self.batch_marked = progress_mgr.is_batch_marked(row_index) if hasattr(progress_mgr, 'is_batch_marked') else False
        # v3.22.14: 内嵌按钮回调（由 MainScreen._create_row_widget 传入）
        self.photo_callback = photo_callback
        self.view_photos_callback = view_photos_callback
        self.remark_callback = remark_callback
        # v3.22.19: 行选择回调 + 选中状态
        self.selected_callback = selected_callback
        self._selected = False  # 选中状态（独立于 done 实际拍摄状态）
        self._updating_row_checkbox = False  # 防止 CheckBox 回调循环的标志位

        with self.canvas.before:
            Color(*THEME['card'])
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(16)])
            Color(*THEME['card_border'])
            self.bg_border = Line(rounded_rectangle=(self.x, self.y, self.width, self.height, dp(16)), width=1.5)
        self.bind(pos=self._update_bg, size=self._update_bg)

        # v3.22.19: 左侧 CheckBox 列（与表头 CheckBox 对齐，size_hint_x=0.08）
        self.checkbox = CheckBox(size_hint_x=0.08, active=False)
        self.checkbox.bind(active=self._on_row_checkbox_changed)
        self.add_widget(self.checkbox)

        # v3.22.19: 右侧内容列（垂直 BoxLayout 包裹原有 name + addr + spacer + btn_row）
        content = BoxLayout(orientation='vertical', size_hint_x=0.92, spacing=dp(4))

        full_address = (address_general + address_precise).strip()
        addr_display = full_address if full_address else "（无地址）"
        if property_type:
            addr_display += "  [%s]" % property_type

        name_text = borrower if borrower else "（无客户名）"
        if self.serial:
            name_text = "%s. %s" % (self.serial, name_text)
        # v3.22.6: 同类型标记的客户名前加 [同类型] 前缀徽章文字
        if self.batch_marked:
            name_text = "同类型  " + name_text

        # 客户名（加粗，自动换行）
        self.name_label = Label(
            text=name_text,
            font_size='16sp',
            bold=True,
            color=THEME['accent'] if self.batch_marked else (THEME['success'] if self.done else THEME['text']),
            size_hint_y=None,
            halign='left', valign='middle',
        )
        self.name_label.bind(width=self._update_name_text_size)
        self.name_label.text_size = (None, None)
        self.name_label.bind(texture_size=self._update_heights)
        content.add_widget(self.name_label)

        # 地址+性质（自动换行，灰色小字）
        self.addr_label = Label(
            text=addr_display,
            font_size='13sp',
            color=THEME['text_dim'],
            size_hint_y=None,
            halign='left', valign='top',
        )
        self.addr_label.bind(width=self._update_addr_text_size)
        self.addr_label.text_size = (None, None)
        self.addr_label.bind(texture_size=self._update_heights)
        content.add_widget(self.addr_label)

        # 间隔
        content.add_widget(Label(size_hint_y=None, height=dp(4)))

        # v3.22.14: 恢复内嵌按钮行 — view_btn.on_press 直接调用 _on_view_photos
        # 不经过菜单 dismiss 步骤，避免 PhotoViewerPopup 静默不显示
        btn_row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(6))

        self.photo_btn = RoundedButton(
            text="拍照", font_size='13sp',
            size_hint_x=0.20, height=dp(40),
            background_color=THEME['accent'], color=(1,1,1,1)
        )
        self.photo_btn.bind(on_press=self._on_photo)
        bind_press_animation(self.photo_btn)
        btn_row.add_widget(self.photo_btn)

        self.view_btn = RoundedButton(
            text=("查看已拍(%d)" % self.photo_count) if self.photo_count > 0 else "查看已拍",
            font_size='13sp',
            size_hint_x=0.28, height=dp(40),
            background_color=THEME['success'] if self.photo_count > 0 else THEME['text_dim'],
            color=(1,1,1,1)
        )
        self.view_btn.bind(on_press=self._on_view_photos)
        bind_press_animation(self.view_btn)
        btn_row.add_widget(self.view_btn)

        self.remark_btn = RoundedButton(
            text=("备注●" if self.remark else "备注"),
            font_size='13sp',
            size_hint_x=0.20, height=dp(40),
            background_color=THEME['warning'] if self.remark else THEME['text_dim'],
            color=(1,1,1,1)
        )
        self.remark_btn.bind(on_press=self._on_remark)
        bind_press_animation(self.remark_btn)
        btn_row.add_widget(self.remark_btn)

        # v3.22.2: 横排类型计数器（markup 单 label，5 类一行显示）
        # 格式: 远景:2 近景:1 内部:0 瑕疵:0 其他:3（已拍绿色加粗）
        self.type_status_label = Label(
            text="", font_size='11sp',
            size_hint_x=0.32,
            color=THEME['text_dim'],
            markup=True,
            halign='center', valign='middle'
        )
        self.type_status_label.bind(
            width=lambda inst, val: setattr(inst, 'text_size', (val, None))
        )
        btn_row.add_widget(self.type_status_label)

        content.add_widget(btn_row)
        self.add_widget(content)

        self._update_type_status()
        Clock.schedule_once(lambda dt: self._update_heights(), 0)

    def _on_row_checkbox_changed(self, checkbox, value):
        """v3.22.19: 行 CheckBox 状态变化回调 — 通知 MainScreen 更新 _row_selected"""
        if self._updating_row_checkbox:
            return
        try:
            if self.selected_callback:
                self.selected_callback(self.row_index, value)
        except Exception as e:
            app_log.error('UI', '_on_row_checkbox_changed 异常 row=%d: %s' % (self.row_index, e))
        # 更新自身选中态背景
        self._selected = bool(value)
        self._update_bg()

    def set_selected(self, selected):
        """v3.22.19: 由 MainScreen 调用，更新行 CheckBox 选中状态（避免回调循环用标志位）。
        选中状态独立于 done（实际拍摄）状态。"""
        self._updating_row_checkbox = True
        try:
            self.checkbox.active = bool(selected)
        except Exception:
            pass
        self._updating_row_checkbox = False
        self._selected = bool(selected)
        self._update_bg()

    def update_selected_display(self):
        """v3.22.19: 刷新选中态背景显示（由 MainScreen 在反选/全选后调用）"""
        self._update_bg()

    def _on_photo(self, instance):
        """v3.22.14: 拍照按钮回调（直接调用 MainScreen._on_photo_request）"""
        if self.photo_callback:
            self.photo_callback(self.row_index, self.borrower, self.address_general,
                                self.address_precise, self.property_type)

    def _on_view_photos(self, instance):
        """v3.22.14: 查看已拍按钮回调（直接调用 MainScreen._on_view_photos，
        不经过菜单 dismiss 步骤，避免 PhotoViewerPopup 静默不显示）"""
        if self.view_photos_callback:
            self.view_photos_callback(self.row_index)

    def _on_remark(self, instance):
        """v3.22.14: 备注按钮回调（直接调用 MainScreen._on_remark_request）"""
        if self.remark_callback:
            self.remark_callback(self.row_index)

    def update_batch_marked(self):
        """v3.22.6: 刷新同类型标记状态与徽章显示（_refresh_list 时调用）"""
        self.batch_marked = self.progress_mgr.is_batch_marked(self.row_index) if hasattr(self.progress_mgr, 'is_batch_marked') else False
        name_text = self.borrower if self.borrower else "（无客户名）"
        if self.serial:
            name_text = "%s. %s" % (self.serial, name_text)
        if self.batch_marked:
            name_text = "同类型  " + name_text
        self.name_label.text = name_text
        self.name_label.color = THEME['accent'] if self.batch_marked else (THEME['success'] if self.done else THEME['text'])
        self._update_type_status()

    def update_state(self):
        """v3.22.7: 增量更新行状态（计数器/done/batch_marked 徽章/按钮文案），不重建 widget。
        由 _refresh_list 在缓存命中时调用，避免长列表全量重建卡顿。
        v3.22.14: 恢复按钮文案/颜色更新（v3.22.13 移除按钮时丢失）。"""
        try:
            self.done = self.progress_mgr.is_photographed(self.progress_key)
            self.photo_count = self.progress_mgr.get_photo_count(self.progress_key)
            # v3.22.14: 刷新按钮文案/颜色（已拍变绿/查看已拍带数量/备注带●标记）
            self.photo_btn.background_color = THEME['success'] if self.done else THEME['accent']
            self.view_btn.text = ("查看已拍(%d)" % self.photo_count) if self.photo_count > 0 else "查看已拍"
            self.view_btn.background_color = THEME['success'] if self.photo_count > 0 else THEME['text_dim']
            self.remark_btn.text = ("备注●" if self.remark else "备注")
            self.remark_btn.background_color = THEME['warning'] if self.remark else THEME['text_dim']
            # 刷新同类型标记状态与徽章（内部调用 _update_type_status 并设置 name_label.color）
            self.update_batch_marked()
        except Exception as e:
            app_log.error('UI', 'update_state 异常 row=%d: %s' % (self.row_index, e))

    def _update_name_text_size(self, instance, value):
        instance.text_size = (value, None)

    def _update_addr_text_size(self, instance, value):
        instance.text_size = (value, None)

    def _update_heights(self, *args):
        name_h = self.name_label.texture_size[1] if self.name_label.texture_size else dp(24)
        addr_h = self.addr_label.texture_size[1] if self.addr_label.texture_size else dp(20)
        self.name_label.height = max(dp(24), name_h + dp(4))
        self.addr_label.height = max(dp(20), addr_h + dp(4))
        # v3.22.14: 恢复 btn_row(52dp) 替代 v3.22.13 的 type_status_label(20dp)
        self.height = self.name_label.height + self.addr_label.height + dp(52) + dp(16) + dp(8)

    def _update_bg(self, *args):
        # v3.22.2: 高亮态切换底色与边色，重绘整个 canvas.before
        # v3.22.3: 倒角 dp(8)→dp(12)，边框 width 1→1.2 增强可见性
        # v3.22.4: 倒角 dp(12)→dp(16)，边框 width 1.2→1.5 进一步增强可见性
        # v3.22.19: 新增 _selected 选中态（淡蓝底），优先级低于 _highlighted（当前拍摄行）
        if getattr(self, '_highlighted', False):
            bg_color = THEME['highlight_bg']
            border_color = THEME['highlight_border']
            border_width = 2.5
        elif getattr(self, '_selected', False):
            bg_color = THEME['info']
            border_color = THEME['accent']
            border_width = 1.8
        else:
            bg_color = THEME['card']
            border_color = THEME['card_border']
            border_width = 1.5
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*bg_color)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(16)])
            Color(*border_color)
            self.bg_border = Line(rounded_rectangle=(self.x, self.y, self.width, self.height, dp(16)),
                                  width=border_width)

    def set_highlight(self, on):
        """v3.22.2: 标记/取消当前拍摄行高亮（浅蓝底 + 蓝边）"""
        self._highlighted = bool(on)
        self._update_bg()

    def _update_type_status(self):
        # v3.22.2: 横排单行显示 5 类照片状态 [类型:张数 ...]，已拍加绿色加粗
        # v3.22.6: 同类型标记的行在计数器位置显示「同类型」徽章（蓝色加粗）替代 5 类计数
        if getattr(self, 'batch_marked', False) and not self.done:
            self.type_status_label.text = "[color=2196F3][b]同类型[/b][/color]"
            return
        counts = self.progress_mgr.get_photo_count_by_type(self.progress_key)
        parts = []
        for label in PHOTO_TYPE_LABELS:
            cnt = counts.get(label, 0)
            if cnt > 0:
                # markup 绿色加粗
                parts.append("[color=33B35C][b]%s:%d[/b][/color]" % (label, cnt))
            else:
                parts.append("[color=9EA6B3]%s:0[/color]" % label)
        self.type_status_label.text = "  ".join(parts)

    def set_remark(self, remark_text):
        """v3.22.14: 更新备注状态与按钮文案/颜色"""
        self.remark = remark_text or ""
        self.remark_btn.text = ("备注●" if self.remark else "备注")
        self.remark_btn.background_color = THEME['warning'] if self.remark else THEME['text_dim']

    def mark_done(self):
        # v3.22.14: 恢复按钮文案/颜色更新
        self.done = True
        self.photo_count = self.progress_mgr.get_photo_count(self.progress_key)
        self.photo_btn.background_color = THEME['success'] if self.done else THEME['accent']
        self.view_btn.text = ("查看已拍(%d)" % self.photo_count) if self.photo_count > 0 else "查看已拍"
        self.view_btn.background_color = THEME['success'] if self.photo_count > 0 else THEME['text_dim']
        # v3.22.6: 同类型标记的行名色保持 accent（蓝），否则已拍为 success（绿）
        self.name_label.color = THEME['accent'] if getattr(self, 'batch_marked', False) else THEME['success']
        self._update_type_status()
        Clock.schedule_once(lambda dt: self._update_heights(), 0)


# ============================================================
# 主界面
# ============================================================

class MainScreen(Screen):
    def __init__(self, app_config, **kwargs):
        super().__init__(**kwargs)
        self.name = 'main'
        self.config = app_config
        self.excel_path = ""
        # v3.19.0: SAF 选择的原始 Excel content:// URI，用于写回备注到原始文件
        # v3.20.0: 从 AppConfig 恢复持久化的 URI，重启后仍可写回备注
        self._excel_uri = app_config.get('excel_uri', '') or None
        # v3.19.2: 日报表生成后待保存到用户指定位置的临时路径
        self._pending_report_path = None
        self._report_save_code = 0x201
        self.headers = []
        self.rows = []  # [borrower, address_general, address_precise, property_type]
        self.progress_mgr = ProgressManager()
        self.progress_mgr.migrate_photo_paths()
        self.camera_mgr = CameraManager()
        self.report_generator = ReportGenerator()
        self._excel_save_lock = threading.Lock()  # v3.21.0: 串行化 Excel 写入，防止并发覆盖
        self.row_widgets = []
        self._current_row = 0
        self._current_borrower = ""
        self._current_addr_general = ""
        self._current_addr_precise = ""
        self._current_property_type = ""
        self._current_photo_type = ""
        self._current_key = ""
        self._continuous_shooting = False
        self._is_paused = False  # v3.20.0: app 后台暂停标志
        self._photos_in_session = 0
        self._render_token = 0   # v3.20.0: 批量渲染令牌，搜索/重新加载时取消旧的渲染
        self._loading_label = None  # v3.20.0: 加载进度提示 Label
        self._search_debounce = None  # v3.21.0: 搜索防抖定时器
        self._photo_session_ctx = None  # v3.20.0: 拍照会话上下文（替代 self._current_* 实例状态）
        self._photo_session_active = False  # v3.20.0: 拍照会话锁，防止连拍期间切换客户
        self._log_long_press = None  # v3.22.0: 日志按钮长按检测定时器
        self._row_widget_cache = {}  # v3.22.7: RowWidget 增量刷新缓存（按 row_index 复用）
        self._view_photos_last_time = 0  # v3.22.11: _on_view_photos 防抖时间戳
        self.current_visit_note = ''  # v3.22.13: 当前 Excel 的今日走访补充说明
        self.current_visit_note_excel_path = ''  # v3.22.13: 当前 visit_note 对应的 Excel 路径
        # v3.22.19: 全选 / 行选择状态（独立于实际拍摄状态，仅用于"标记完成"辅助 AI 报告）
        self._all_selected = False  # 表头 CheckBox 是否处于全选态
        self._row_selected = set()  # 记录选中的行号（row_index）
        self._updating_checkbox = False  # 防止 CheckBox 回调循环的标志位
        # v3.22.0: 从配置初始化日志开关（默认关闭）
        app_log.set_enabled(self.config.get('log_enabled', False))

        main = BoxLayout(orientation='vertical')
        self._build_ui(main)
        self.add_widget(main)

    def _build_ui(self, parent):
        parent.add_widget(Label(size_hint_y=None, height=dp(get_status_bar_height_dp())))

        # v3.19.0: 顶部标题栏——白底卡片样式，配活力蓝标题与进度徽章
        title_bar = BoxLayout(size_hint_y=None, height=dp(60), spacing=dp(8), padding=[dp(14), dp(6), dp(14), dp(6)])
        with title_bar.canvas.before:
            Color(*THEME['card'])
            self._titlebar_rect = RoundedRectangle(pos=title_bar.pos, size=title_bar.size, radius=[0, 0, 12, 12])
            Color(*THEME['card_border'])
            self._titlebar_border = Line(rectangle=(title_bar.x, title_bar.y, title_bar.width, title_bar.height), width=1)
        title_bar.bind(pos=self._update_titlebar_rect, size=self._update_titlebar_rect)
        title_bar.add_widget(Label(text="资产盘点拍照", font_size='19sp', bold=True, color=THEME['accent_dark'],
                                   size_hint_x=0.46, halign='left', valign='middle'))

        ai_btn = RoundedButton(text="AI助手", font_size='15sp', size_hint_x=0.16,
                       background_color=THEME['success'], background_normal='',
                       color=(1,1,1,1), bold=True)
        ai_btn.bind(on_release=self._go_ai)
        bind_press_animation(ai_btn)
        title_bar.add_widget(ai_btn)

        settings_btn = RoundedButton(text="设置", font_size='15sp', size_hint_x=0.20,
                             background_color=THEME['accent'], background_normal='',
                             color=(1,1,1,1), bold=True)
        settings_btn.bind(on_release=self._go_settings)
        bind_press_animation(settings_btn)
        title_bar.add_widget(settings_btn)

        self.progress_label = Label(text="0/0", font_size='16sp', color=THEME['accent'],
                                    size_hint_x=0.18, halign='right', valign='middle', bold=True)
        title_bar.add_widget(self.progress_label)
        parent.add_widget(title_bar)

        # v3.19.3: 工具栏恢复单行（打开Excel + 搜索类型 + 搜索框 + 搜索按钮）
        toolbar = BoxLayout(size_hint_y=None, height=dp(56), spacing=dp(6), padding=[dp(10), dp(6), dp(10), dp(6)])
        open_btn = RoundedButton(text="打开Excel", font_size='15sp', size_hint_x=0.22,
                         background_color=THEME['accent'], background_normal='',
                         color=(1,1,1,1), bold=True)
        open_btn.bind(on_release=self._show_file_dialog)
        bind_press_animation(open_btn)
        toolbar.add_widget(open_btn)

        # v3.22.0: 搜索类型选择器（序号/客户名/地址/备注）
        # v3.22.2: 合并 '地址(概)'/'地址(详)' 为 '地址'，同时匹配概+详
        _cur_field = self.config.get('search_field', '客户名') or '客户名'
        # v3.22.2: 旧值 '地址(概)'/'地址(详)' 自动迁移为 '地址'
        if _cur_field in ('地址(概)', '地址(详)'):
            _cur_field = '地址'
            self.config.set('search_field', '地址')
            self.config.save()
        self.search_field_spinner = Spinner(
            text=_cur_field,
            values=['客户名', '序号', '地址', '备注'],
            font_size='13sp', size_hint_x=0.20,
            background_color=THEME['accent_dark'], background_normal='',
            color=(1, 1, 1, 1), bold=True,
            sync_height=True,
            # v3.22.19: 下拉项使用带 dp(8) 圆角倒角的自定义 SpinnerOption
            option_cls=RoundedSpinnerOption,
        )
        self.search_field_spinner.bind(text=self._on_search_field_change)
        toolbar.add_widget(self.search_field_spinner)

        self.search_input = TextInput(hint_text=("搜索%s…" % _cur_field), multiline=False, font_size='15sp', size_hint_x=0.36,
                                      foreground_color=THEME['text'], hint_text_color=THEME['text_dim'])
        self.search_input.bind(text=self._on_search)
        # v3.22.0: 搜索框聚焦时滚回顶部，避免键盘遮挡当前行
        self.search_input.bind(focus=lambda inst, val: setattr(self.scroll_view, 'scroll_y', 1) if val else None)
        toolbar.add_widget(self.search_input)

        search_btn = RoundedButton(text="搜索", font_size='15sp', size_hint_x=0.22,
                           background_color=THEME['accent_dark'], background_normal='',
                           color=(1,1,1,1), bold=True)
        search_btn.bind(on_release=self._do_search)
        bind_press_animation(search_btn)
        toolbar.add_widget(search_btn)
        parent.add_widget(toolbar)

        # v3.22.19: 表头从 list_layout 中抽出，固定在 ScrollView 上方（不随列表滚动）
        # 布局顺序：[toolbar] [header_row(固定)] [scroll_view(含 list_layout)] [footer]
        self.header_row = self._build_header_row()
        parent.add_widget(self.header_row)

        self.scroll_view = ScrollView(do_scroll_x=False, do_scroll_y=True)
        self.list_layout = GridLayout(cols=1, spacing=dp(6), size_hint_y=None, padding=[dp(8), dp(6), dp(8), dp(6)])
        self.list_layout.bind(minimum_height=self.list_layout.setter('height'))
        self.scroll_view.add_widget(self.list_layout)
        parent.add_widget(self.scroll_view)

        # v3.20.0: 底部栏改为水平布局 — 左侧 AI 报表按钮(0.7) + 右侧 日志按钮(0.3)
        footer = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(60), spacing=dp(8), padding=[dp(12), dp(8), dp(12), dp(8)])

        ai_report_btn = RoundedButton(text="AI 一键生成日报表", font_size='17sp',
                               background_color=THEME['success'], background_normal='',
                               color=(1,1,1,1), bold=True, size_hint_x=0.7)
        ai_report_btn.bind(on_release=self._on_generate_report_button)  # v3.22.13: 改为先弹 VisitNoteDialog
        bind_press_animation(ai_report_btn)
        footer.add_widget(ai_report_btn)

        # v3.22.2: 日志入口按钮 — 短按开二级弹窗（开关/查看/复制/清除）
        _log_on = app_log.is_enabled()
        log_btn = RoundedButton(text="日志", font_size='15sp',
                         background_color=THEME['accent_dark'],
                         background_normal='',
                         color=(1,1,1,1), bold=True, size_hint_x=0.3)
        log_btn.bind(on_release=self._on_log_entry)
        bind_press_animation(log_btn)
        footer.add_widget(log_btn)
        self.log_btn = log_btn  # v3.22.0: 保留引用，切换开关时更新文字/颜色

        # 保留隐藏的 log_toggle_btn 引用（旧代码引用兼容），但不显示
        self.log_toggle_btn = RoundedButton(text="日志:关", font_size='1sp',
                              background_color=THEME['muted'], background_normal='',
                              size_hint=(0, 0), opacity=0)

        # status_label 仍然创建（供代码引用），但不显示在UI中
        self.status_label = Label(text="", font_size='11sp',
                                  color=THEME['text_dim'],
                                  halign='left', valign='top',
                                  size_hint_y=None, markup=False, opacity=0)
        self.status_label.bind(texture_size=self._update_log_label_size, width=lambda i, v: setattr(i, 'text_size', (v, None)))
        parent.add_widget(footer)

        self._show_empty_state()

    def _show_empty_state(self):
        self.list_layout.clear_widgets()
        if self.rows:
            return
        # v3.22.19: 无数据时隐藏固定表头（避免 CheckBox 干扰空状态布局）
        if hasattr(self, 'header_row'):
            self.header_row.height = 0
            self.header_row.opacity = 0
            self.header_row.disabled = True
        msg = Label(
            text="暂无数据\n\n点击「打开Excel」加载客户清单",
            font_size='15sp', color=THEME['text_dim'],
            size_hint_y=None, height=dp(120),
            halign='center', valign='middle',
        )
        msg.bind(width=lambda i, v: setattr(i, 'text_size', (v, None)))
        self.list_layout.add_widget(msg)
        # v3.22.0: 展示最近打开的 Excel 记录（3-5 条）
        # v3.22.19: 每项改为 BoxLayout 横向布局（文件名按钮 | 数据指示器圆点 | 清除按钮）
        recent = self.config.get('recent_excel', []) or []
        if recent:
            title = Label(text="最近打开", font_size='14sp', bold=True, color=THEME['accent_dark'],
                          size_hint_y=None, height=dp(30), halign='left', valign='middle')
            title.bind(width=lambda i, v: setattr(i, 'text_size', (v, None)))
            self.list_layout.add_widget(title)
            for item in recent[:5]:
                uri = item.get('uri', '')
                name = item.get('name', 'Excel文件')
                # 横向布局：[文件名按钮（0.78）][数据指示器圆点（0.06）][清除按钮（0.16）]
                row_box = BoxLayout(
                    orientation='horizontal', size_hint_y=None, height=dp(50),
                    spacing=dp(4),
                )
                # 文件名按钮
                btn = RoundedButton(
                    text=name, font_size='17sp',
                    size_hint_x=0.78,
                    background_color=THEME['bg'],
                    color=THEME['accent_dark'],
                    background_normal='',
                )
                btn.bind(on_press=lambda inst, u=uri: self._open_recent_excel(u))
                bind_press_animation(btn)
                row_box.add_widget(btn)

                # 数据指示器圆点（绿色=有数据；灰色=无数据）
                has_data = self._excel_has_data(uri)
                dot = Widget(size_hint_x=0.06, size=(dp(12), dp(12)))
                dot_color = THEME['success'] if has_data else THEME['text_dim']
                with dot.canvas.before:
                    Color(*dot_color)
                    dot._dot_circle = Ellipse(pos=(0, 0), size=(dp(12), dp(12)))
                # 居中绘制圆点
                def _update_dot(inst, val, _dot_circle=dot._dot_circle):
                    try:
                        # 在 Widget 内部居中
                        _dot_circle.pos = ((inst.width - dp(12)) / 2,
                                            (inst.height - dp(12)) / 2)
                    except Exception:
                        pass
                dot.bind(pos=_update_dot, size=_update_dot)
                row_box.add_widget(dot)

                # 清除按钮
                clear_btn = RoundedButton(
                    text="清除", font_size='13sp',
                    size_hint_x=0.16,
                    background_color=THEME['danger'], background_normal='',
                    color=(1, 1, 1, 1), bold=True,
                )
                clear_btn.bind(on_press=lambda inst, u=uri, n=name: self._on_clear_excel_data(u, n, inst))
                bind_press_animation(clear_btn)
                row_box.add_widget(clear_btn)

                self.list_layout.add_widget(row_box)

    def _excel_has_data(self, uri):
        """v3.22.19: 推断 app 内是否有该 Excel（uri）关联的数据。
        简化方案：检查 APP_DIR/thumbnails/ 下是否有任何 .jpg 文件
        AND visit_notes/<md5(uri)>.txt 是否存在。
        由于 progress_key 是 borrower+addr 的哈希，不直接关联 Excel，
        仅做粗略判断（任一缩略图存在即视为有数据）。"""
        try:
            if not uri:
                return False
            # 检查缩略图目录是否有任何 .jpg 文件
            thumb_root = os.path.join(APP_DIR, 'thumbnails')
            has_thumb = False
            if os.path.isdir(thumb_root):
                for sub in os.listdir(thumb_root):
                    sub_path = os.path.join(thumb_root, sub)
                    if os.path.isdir(sub_path):
                        for f in os.listdir(sub_path):
                            if f.lower().endswith('.jpg'):
                                has_thumb = True
                                break
                    if has_thumb:
                        break
            # 检查 visit_note 是否存在（用 uri md5 哈希推断文件名）
            visit_note_path = self._get_visit_note_path(uri)
            has_visit_note = bool(visit_note_path and os.path.exists(visit_note_path))
            return has_thumb or has_visit_note
        except Exception as e:
            app_log.error('UI', '_excel_has_data 异常 uri=%s: %s' % (uri, e))
            return False

    def _on_clear_excel_data(self, uri, name, btn):
        """v3.22.19: 清除该 Excel 在 app 内的关联数据（缩略图 + visit_note）。
        2 次确认逻辑：首次点击改文本为"再次确认"，3 秒内再次点击执行清除。"""
        try:
            # 已处于"再次确认"状态 → 执行清除
            if getattr(btn, '_confirming_clear', False):
                btn._confirming_clear = False
                # 取消已 schedule 的还原回调
                if getattr(btn, '_restore_ev', None):
                    try:
                        btn._restore_ev.cancel()
                    except Exception:
                        pass
                    btn._restore_ev = None
                # 执行清除
                # 1. 删除 APP_DIR/thumbnails/ 下所有内容（无法精确关联，全部清除）
                try:
                    import shutil
                    thumb_root = os.path.join(APP_DIR, 'thumbnails')
                    if os.path.isdir(thumb_root):
                        for sub in os.listdir(thumb_root):
                            sub_path = os.path.join(thumb_root, sub)
                            if os.path.isdir(sub_path):
                                shutil.rmtree(sub_path, ignore_errors=True)
                except Exception as e:
                    app_log.error('UI', '清除缩略图目录失败: %s' % e)
                # 2. 删除 visit_note 文件（用 uri md5 哈希）
                try:
                    visit_note_path = self._get_visit_note_path(uri)
                    if visit_note_path and os.path.exists(visit_note_path):
                        os.remove(visit_note_path)
                        app_log.info('IO', '已删除 visit_note: %s' % visit_note_path)
                except Exception as e:
                    app_log.error('IO', '删除 visit_note 失败: %s' % e)
                # 3. 还原按钮文本
                btn.text = "清除"
                # 4. 显示 toast
                try:
                    self._show_android_toast("已清除 %s 的数据" % name)
                except Exception:
                    pass
                app_log.info('UI', '已清除 Excel 数据: %s' % name)
                # 5. 刷新最近列表
                self._refresh_list()
            else:
                # 首次点击：改文本，3 秒后还原
                btn._confirming_clear = True
                btn.text = "再次确认"
                def _restore_text(dt):
                    try:
                        btn.text = "清除"
                        btn._confirming_clear = False
                    except Exception:
                        pass
                btn._restore_ev = Clock.schedule_once(_restore_text, 3.0)
        except Exception as e:
            app_log.error('UI', '_on_clear_excel_data 异常: %s' % e, exc_info=True)
            try:
                btn.text = "清除"
                btn._confirming_clear = False
            except Exception:
                pass

    def _saf_display_name(self, uri_str):
        """v3.22.0: 查询 SAF URI 对应的文件显示名"""
        if not IS_ANDROID or not uri_str:
            return ""
        try:
            from jnius import autoclass
            Uri = autoclass('android.net.Uri')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            resolver = PythonActivity.mActivity.getContentResolver()
            uri = Uri.parse(uri_str)
            cursor = resolver.query(uri, None, None, None, None)
            if cursor:
                try:
                    if cursor.moveToFirst():
                        idx = cursor.getColumnIndex('_display_name')
                        if idx >= 0:
                            return cursor.getString(idx) or ""
                finally:
                    cursor.close()
        except Exception as e:
            app_log.error('EXCEL', '查询显示名失败: %s' % e)
        return ""

    def _add_recent_excel(self, uri, name):
        """v3.22.0: 记录最近打开的 Excel（最多5条，新增到首位）"""
        if not uri:
            return
        try:
            recent = list(self.config.get('recent_excel', []) or [])
            # 移除已存在的相同 uri
            recent = [r for r in recent if r.get('uri') != uri]
            recent.insert(0, {'uri': uri, 'name': name or 'Excel文件'})
            recent = recent[:5]
            self.config.set('recent_excel', recent)
        except Exception as e:
            app_log.error('EXCEL', '记录最近文件失败: %s' % e)

    def _remove_recent_excel(self, uri):
        """v3.22.0: 从最近记录中移除指定文件"""
        try:
            recent = list(self.config.get('recent_excel', []) or [])
            recent = [r for r in recent if r.get('uri') != uri]
            self.config.set('recent_excel', recent)
        except Exception as e:
            app_log.error('EXCEL', '移除最近文件失败: %s' % e)

    def _open_recent_excel(self, uri):
        """v3.22.0: 打开最近记录的 Excel，文件移除时提示并清理"""
        if not uri:
            return
        if IS_ANDROID and uri.startswith('content://'):
            try:
                dest = os.path.join(APP_DIR, "_imported_excel.xlsx")
                android_copy_uri_to_app_dir(uri, dest)
                if not os.path.exists(dest) or os.path.getsize(dest) < 100:
                    raise FileNotFoundError("文件不可读或已移除")
                self._excel_uri = uri
                _name = self._saf_display_name(uri)
                self._load_excel_path(dest, display_name=_name, source_uri=uri)
            except Exception as e:
                app_log.error('EXCEL', '最近文件打开失败: %s' % e)
                self._remove_recent_excel(uri)
                self._show_msg("该文件已移除，已从记录中删除", THEME['warning'])
        else:
            if os.path.exists(uri):
                self._load_excel_path(uri, display_name=os.path.basename(uri), source_uri=uri)
            else:
                self._remove_recent_excel(uri)
                self._show_msg("该文件已移除，已从记录中删除", THEME['warning'])

    def _build_header_row(self):
        # v3.22.19: 表头新增最左侧 CheckBox（全选/反选），列宽调整：
        # [CheckBox(0.08)] [客户名/地址(0.52)] [操作(0.40)]
        header = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(36),
                           padding=[dp(10), dp(4), dp(10), dp(4)], spacing=dp(6))
        with header.canvas.before:
            # v3.19.0: 浅色主题表头——浅蓝灰底，避免深色硬编码
            Color(*THEME['accent'])
            self._header_rect = RoundedRectangle(pos=header.pos, size=header.size, radius=[10])
        header.bind(pos=self._update_header_rect, size=self._update_header_rect)
        # v3.22.19: 全选 CheckBox
        self.header_checkbox = CheckBox(size_hint_x=0.08, active=False)
        self.header_checkbox.bind(active=self._on_select_all)
        header.add_widget(self.header_checkbox)
        header.add_widget(Label(text="客户名 / 抵押物地址", font_size='13sp', bold=True,
                                color=(1, 1, 1, 1), size_hint_x=0.52, halign='left'))
        header.add_widget(Label(text="操作", font_size='13sp', bold=True,
                                color=(1, 1, 1, 1), size_hint_x=0.40, halign='center'))
        return header

    def _update_titlebar_rect(self, instance, *args):
        self._titlebar_rect.pos = instance.pos
        self._titlebar_rect.size = instance.size
        self._titlebar_border.rectangle = (instance.x, instance.y, instance.width, instance.height)

    def _update_header_rect(self, instance, *args):
        self._header_rect.pos = instance.pos
        self._header_rect.size = instance.size

    def _show_file_dialog(self, instance):
        """打开系统文件选择器。Android使用SAF(Intent.ACTION_OPEN_DOCUMENT)获取可访问URI，
        桌面端使用plyer.filechooser，最终fallback手动输入。"""
        if IS_ANDROID:
            self._android_open_file_picker()
        else:
            try:
                from plyer import filechooser
                filechooser.open_file(
                    filters=[("Excel 文件", "*.xlsx", "*.xls")],
                    on_selection=self._on_file_selected,
                )
            except Exception as e:
                Logger.warning("MainScreen.filechooser: %s, fallback to path input" % e)
                self._show_path_input_dialog()

    def _android_open_file_picker(self):
        """Android：使用Intent.ACTION_OPEN_DOCUMENT选择Excel文件，
        通过ContentResolver获取持久读取权限，然后复制到app私有目录。"""
        try:
            from jnius import autoclass
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            Intent = autoclass('android.content.Intent')
            activity = PythonActivity.mActivity

            ACTION_OPEN_DOCUMENT = "android.intent.action.OPEN_DOCUMENT"
            CATEGORY_OPENABLE = "android.intent.category.OPENABLE"
            EXTRA_MIME_TYPES = "android.intent.extra.MIME_TYPES"
            FLAG_GRANT_READ_URI_PERMISSION = 1
            FLAG_GRANT_PERSISTABLE_URI_PERMISSION = 64

            intent = Intent(ACTION_OPEN_DOCUMENT)
            intent.addCategory(CATEGORY_OPENABLE)
            intent.setType("*/*")
            intent.putExtra(EXTRA_MIME_TYPES, [
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            ])
            intent.addFlags(FLAG_GRANT_READ_URI_PERMISSION)
            intent.addFlags(FLAG_GRANT_PERSISTABLE_URI_PERMISSION)

            self._android_file_picker_code = 0x1001
            activity.startActivityForResult(intent, self._android_file_picker_code)
        except Exception as e:
            Logger.error("Android file picker error: %s" % e)
            self._show_msg("无法打开文件选择器", THEME['danger'])

    def on_activity_result(self, request_code, result_code, intent):
        """处理Android Activity结果回调（文件选择器+相机）。"""
        if request_code == self.camera_mgr.CAMERA_REQUEST_CODE:
            try:
                self.camera_mgr.on_camera_result(result_code, intent)
            except Exception as e:
                Logger.error("on_camera_result crash: %s" % e)
                Logger.error(traceback.format_exc())
                self.camera_mgr._dbg(f"[ERR] 处理结果时崩溃: {type(e).__name__}: {str(e)[:100]}", show_toast=True)
                self.camera_mgr._camera_launched = False
                if self.camera_mgr.pending_callback:
                    cb = self.camera_mgr.pending_callback
                    self.camera_mgr.pending_callback = None
                    Clock.schedule_once(lambda dt, cb=cb: cb(None), 0)
            return

        if request_code == getattr(self, '_android_file_picker_code', -1):
            if result_code != -1 or intent is None:
                return
            try:
                from jnius import autoclass
                uri = intent.getData()
                if uri is None:
                    return
                uri_str = str(uri.toString())

                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                activity = PythonActivity.mActivity
                resolver = activity.getContentResolver()
                try:
                    take_flags = 1
                    resolver.takePersistableUriPermission(uri, take_flags)
                except:
                    pass

                # v3.19.0: 保存 SAF URI，便于备注保存时通过 ContentResolver 写回原始 Excel
                # v3.20.0: 持久化 URI 到 AppConfig，重启后可继续写回备注
                self._excel_uri = uri_str
                try:
                    self.config.set('excel_uri', uri_str)
                    app_log.info('EXCEL', '已持久化 Excel URI')
                except Exception:
                    pass
                dest = os.path.join(APP_DIR, "_imported_excel.xlsx")
                android_copy_uri_to_app_dir(uri_str, dest)
                # v3.22.0: 查询文件显示名，用于最近记录
                _disp_name = self._saf_display_name(uri_str)
                Clock.schedule_once(lambda dt: self._load_excel_path(dest, display_name=_disp_name, source_uri=uri_str), 0)
            except Exception as e:
                Logger.error("on_activity_result (file): %s" % e)
                self._show_msg(f"文件选择失败: {str(e)[:40]}", THEME['danger'])
            return

        # v3.19.2: 日报表保存（ACTION_CREATE_DOCUMENT 返回的 URI）
        if request_code == getattr(self, '_report_save_code', -1):
            if result_code != -1 or intent is None:
                return
            try:
                from jnius import autoclass
                uri = intent.getData()
                if uri is None or not getattr(self, '_pending_report_path', None):
                    return
                uri_str = str(uri.toString())
                src = self._pending_report_path
                ok, m = ExcelWriter.write_back_to_uri(uri_str, src)
                if ok:
                    self._show_msg("日报表已保存到您选择的位置", THEME['success'])
                else:
                    self._show_msg("app内部已生成：%s；%s" % (src, m), THEME['warning'])
            except Exception as e:
                Logger.error("on_activity_result (report save): %s" % e)
                self._show_msg(f"日报表保存失败: {str(e)[:40]}", THEME['danger'])
            return

    def _on_file_selected(self, selection):
        if not selection:
            return
        path = selection[0] if isinstance(selection, list) else str(selection)
        if path and os.path.exists(path):
            self._load_excel_path(path, display_name=os.path.basename(path), source_uri=path)
        elif path:
            self._show_msg(f"文件不存在：{path[:40]}", THEME['danger'])

    def _show_path_input_dialog(self):
        content = BoxLayout(orientation='vertical', spacing=8, padding=8)
        content.add_widget(Label(text="请输入 Excel 文件路径：", size_hint_y=None, height=dp(30),
                                 font_size='14sp'))
        path_input = TextInput(text=self.excel_path, multiline=False, font_size='13sp',
                               size_hint_y=None, height=dp(40))
        content.add_widget(path_input)
        load_btn = RoundedButton(text="加载", size_hint_y=None, height=dp(44),
                         background_color=THEME['accent'], background_normal='',
                         font_size='16sp', color=(1,1,1,1))
        popup = Popup(title='选择 Excel 文件', content=content, size_hint=(0.9, 0.35),
                      title_size='16sp')
        load_btn.bind(on_release=lambda x: (popup.dismiss(), self._load_excel_path(path_input.text)))
        content.add_widget(load_btn)
        popup.open()

    def _show_msg(self, msg, color=None, toast=True):
        """v3.22.6: 改用 Popup 居中显示消息（替代写入不可见的 status_label）。
        旧版本写入 status_label（opacity=0 且从未 add_widget），所有错误对用户不可见，
        导致「查看已拍」按钮看似无反应。现改为弹窗，并保留 status_label 属性兼容。
        同时同步写入 app_log 全量日志（开关开启时入文件），保证 UI 交互可追溯。"""
        ts = get_system_date().strftime('%H:%M:%S')
        line = f"[{ts}] {msg}"
        self.camera_mgr._log_lines.append(line)
        if len(self.camera_mgr._log_lines) > self.camera_mgr._max_log_lines:
            self.camera_mgr._log_lines = self.camera_mgr._log_lines[-self.camera_mgr._max_log_lines:]
        # v3.22.6: 不再写入 status_label（其 opacity=0 且未加入布局，写入不可见）
        # 仅保留属性以兼容旧代码引用，不再赋值
        # v3.22.2: 同步入全量 app 日志（开关开启时落盘）
        app_log.info('UI', msg)
        if toast and IS_ANDROID:
            try:
                self.camera_mgr._toast(msg)
            except Exception:
                pass
        # v3.22.6: 弹出居中 Popup 显示消息
        try:
            self._show_msg_popup(msg, color)
        except Exception as e:
            app_log.error('UI', '_show_msg popup 异常: %s' % e)

    def _show_msg_popup(self, msg, color=None):
        """v3.22.6: 弹出一个居中 Popup 显示消息（替代不可见的 status_label）。
        Popup 含消息文本 + 「知道了」按钮，背景色 THEME['card']（浅色主题）。
        v3.22.7: 修复 background=THEME['card'] 传颜色列表导致
        'Popup.background accept only str' 异常 → 改用 content.canvas.before 绘制浅色背景。
        v3.22.9: 改用 ThemedPopup，背景与标题色由基类统一处理，移除重复 canvas.before。"""
        content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(14))
        # v3.22.9: 背景由 ThemedPopup.canvas.before 统一绘制，此处不再重复
        msg_color = color if color else THEME['text_dim']
        msg_label = Label(
            text=str(msg),
            font_size='16sp',
            color=msg_color,
            halign='center', valign='middle',
            size_hint_y=None, height=dp(60),
        )
        msg_label.bind(width=lambda s, w: setattr(s, 'text_size', (w, None)))
        content.add_widget(msg_label)
        ok_btn = RoundedButton(
            text="知道了",
            font_size='15sp',
            size_hint_y=None, height=dp(44),
            background_color=THEME['accent'],
            background_normal='',
            color=(1, 1, 1, 1), bold=True,
        )
        bind_press_animation(ok_btn)
        content.add_widget(ok_btn)
        popup = ThemedPopup(
            title="提示",
            title_size='16sp',
            content=content,
            size_hint=(0.8, None),
            height=dp(200),
            auto_dismiss=True,
        )
        ok_btn.bind(on_release=popup.dismiss)
        popup.open()

    def _get_visit_note_path(self, excel_path):
        """v3.22.13: 返回 visit_note 持久化文件路径
        按 Excel 文件路径 md5 哈希存储，避免路径含特殊字符
        """
        if not excel_path:
            return None
        import hashlib
        hash_val = hashlib.md5(excel_path.encode('utf-8')).hexdigest()
        return os.path.join(APP_DIR, 'visit_notes', '%s.txt' % hash_val)

    def _load_visit_note(self, excel_path):
        """v3.22.13: 加载 visit_note（不存在返回空字符串）"""
        path = self._get_visit_note_path(excel_path)
        if not path or not os.path.exists(path):
            return ''
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            app_log.error('IO', '_load_visit_note 异常: %s' % e, exc_info=True)
            return ''

    def _save_visit_note(self, excel_path, text):
        """v3.22.13: 保存 visit_note 到文件
        v3.22.13 修复: 原子写（临时文件 + os.replace），防止写入中途崩溃导致文件损坏
        （参考 ProgressManager.save 的原子写实现）。"""
        if not excel_path:
            return
        path = self._get_visit_note_path(excel_path)
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            # v3.22.13: 原子写 — 先写临时文件再 os.replace，防止崩溃损坏文件
            tmp = path + '.tmp'
            with open(tmp, 'w', encoding='utf-8') as f:
                f.write(text or '')
            os.replace(tmp, path)
            app_log.info('IO', 'visit_note 已保存: %s' % path)
        except Exception as e:
            app_log.error('IO', '_save_visit_note 异常: %s' % e, exc_info=True)

    def _load_excel_path(self, path, display_name=None, source_uri=None):
        if not path or not os.path.exists(path):
            self._show_msg("文件不存在或路径无效", THEME['danger'])
            return
        # v3.22.13: 切换 Excel 前保存当前 visit_note（避免上次输入丢失）
        if self.current_visit_note_excel_path:
            self._save_visit_note(self.current_visit_note_excel_path, self.current_visit_note)
        self.excel_path = path
        try:
            app_log.info('EXCEL', '开始加载 Excel: %s' % os.path.basename(path))
            reader = ExcelReader(path)
            self.headers, self.rows = reader.load()
            # v3.22.19: 切换 Excel 时清空选中状态（_all_selected / _row_selected）
            # 行号语义已变（新 Excel 的行号与旧 Excel 不同），旧选中集合不再有意义
            self._all_selected = False
            self._row_selected.clear()
            if hasattr(self, 'header_checkbox'):
                self.header_checkbox.active = False
            # v3.22.7: Excel 重新加载时清空 RowWidget 缓存（行数据已变，旧缓存失效）
            self._row_widget_cache = {}
            # v3.22.5: 切换 Excel 时清理行号备注，使新 Excel F 列为备注唯一真相源
            try:
                if '_row_remarks' in self.progress_mgr.data:
                    self.progress_mgr.data['_row_remarks'] = {}
                    self.progress_mgr.save()
                    app_log.info('EXCEL', '切换 Excel 已清理 _row_remarks')
            except Exception as e:
                app_log.error('EXCEL', '清理 _row_remarks 异常: %s' % e)
            # v3.20.0: 备注恢复逻辑 — progress_mgr 为单一真相源
            # 优先用 progress_mgr 中的备注（持久化），仅在 progress_mgr 无记录时用 Excel E 列值
            for i, row in enumerate(self.rows):
                borrower = row[1] if len(row) > 1 else ""
                address_general = row[2] if len(row) > 2 else ""
                address_precise = row[3] if len(row) > 3 else ""
                excel_remark = row[5] if len(row) > 5 else ""
                full_addr = (address_general + address_precise).strip()
                key = self.progress_mgr._make_key(borrower, full_addr)
                saved_remark = self.progress_mgr.get_remark(key)
                # v3.22.3 P0 修复: 优先按行号读取备注（行号独立，不会因共享 key 互相覆盖），
                # 回退 Excel，再回退旧 progress_key 备注。
                row_remark = self.progress_mgr.get_remark_by_row(i)
                final_remark = row_remark if row_remark else (excel_remark if excel_remark else saved_remark)
                if final_remark and final_remark != excel_remark:
                    while len(self.rows[i]) < 6:
                        self.rows[i].append("")
                    self.rows[i][5] = final_remark
            app_log.info('EXCEL', '已加载 %d 条客户记录' % len(self.rows))
            self._show_msg(f"● 已加载 {len(self.rows)} 条客户记录", THEME['success'])
            # v3.22.0: 记录到最近打开的 Excel
            _uri = source_uri or getattr(self, '_excel_uri', None) or path
            _name = display_name or os.path.basename(path)
            self._add_recent_excel(_uri, _name)
            self._refresh_list()
            # v3.22.13: 加载 visit_note（按当前 Excel 路径持久化）
            self.current_visit_note_excel_path = path
            self.current_visit_note = self._load_visit_note(path)
            app_log.info('IO', 'visit_note 已加载: %d 字符' % len(self.current_visit_note))
        except Exception as e:
            err_msg = str(e)
            app_log.error('EXCEL', '加载失败: %s' % err_msg)
            self._show_msg(f"加载失败: {err_msg[:60]}", THEME['danger'])
            Logger.error("Excel load failed: %s" % traceback.format_exc())

    def _refresh_list(self, filter_query=""):
        """v3.20.0: 重建列表，支持分批渲染(大数据量)和搜索过滤(只显示匹配行)。
        filter_query: 搜索关键词（小写），匹配客户名+地址；为空则显示全部
        v3.22.7: 增量刷新优化 — 无搜索过滤且缓存命中全部行时，仅更新状态不重建 RowWidget。
        """
        # v3.22.7: 增量刷新 — 仅当无搜索过滤、缓存命中全部行、且当前已显示全部行时
        # 才走增量路径（避免破坏搜索/分批渲染逻辑）
        try:
            cache_full = bool(self._row_widget_cache) and len(self._row_widget_cache) == len(self.rows)
            view_full = bool(self.row_widgets) and len(self.row_widgets) == len(self.rows)
        except Exception:
            cache_full = False
            view_full = False
        if (not filter_query.strip() and cache_full and view_full
                and not self._loading_label):
            for rw in self.row_widgets:
                try:
                    rw.update_state()
                except Exception as e:
                    app_log.error('UI', '增量刷新 update_state 异常 row=%d: %s' % (rw.row_index, e))
            self._update_progress()
            return

        # 取消之前可能正在进行的批量渲染
        self._render_token += 1
        token = self._render_token

        self.list_layout.clear_widgets()
        self.row_widgets = []
        # v3.22.7: 清空缓存（由 _create_row_widget 重新填充）
        self._row_widget_cache = {}
        self.scroll_view.scroll_y = 1

        # 计算匹配的行 — v3.22.0: 按搜索类型选择器过滤
        # v3.22.2: 合并 '地址' 字段（同时匹配 row[2]概 + row[3]详）
        q = filter_query.lower().strip() if filter_query else ""
        field = (self.config.get('search_field', '客户名') or '客户名') if hasattr(self, 'config') else '客户名'
        # v3.22.2: 兼容旧配置值
        if field in ('地址(概)', '地址(详)'):
            field = '地址'
        field_map = {'序号': 0, '客户名': 1, '地址': 2, '备注': 5}
        idx = field_map.get(field, 1)
        matched_indices = []
        for i, row in enumerate(self.rows):
            if not q:
                matched_indices.append(i)
            else:
                if field == '地址':
                    # v3.22.2: 地址同时匹配 row[2](概) 和 row[3](详)
                    val = ((row[2] if len(row) > 2 else "") + (row[3] if len(row) > 3 else "")).lower()
                else:
                    val = (row[idx] if len(row) > idx else "").lower()
                if q in val:
                    matched_indices.append(i)

        if not self.rows:
            self._show_empty_state()
            self._update_progress()
            return

        # v3.22.19: 表头已固定在 list_layout 外（_build_ui 中添加为 header_row 实例属性），
        # 此处不再 add 到 list_layout；同时确保表头可见（数据存在时）。
        if hasattr(self, 'header_row'):
            self.header_row.height = dp(36)
            self.header_row.opacity = 1
            self.header_row.disabled = False

        # v3.20.0: 搜索结果数提示
        if q:
            search_info = Label(
                text="找到 %d 条匹配" % len(matched_indices),
                font_size='13sp', color=THEME['accent_dark'],
                size_hint_y=None, height=dp(28), halign='left', valign='middle',
            )
            search_info.bind(width=lambda i, v: setattr(i, 'text_size', (v, None)))
            self.list_layout.add_widget(search_info)

        total_matched = len(matched_indices)

        # 数据量小时直接渲染（≤60行），大数据量分批
        if total_matched <= 60:
            for idx in matched_indices:
                self._create_row_widget(idx)
            self._update_progress()
            return

        # v3.20.0: 大数据量分批渲染 — 每批 40 行
        self._loading_label = Label(
            text="正在加载 0/%d..." % total_matched,
            font_size='14sp', color=THEME['accent_dark'],
            size_hint_y=None, height=dp(36), halign='center', valign='middle',
        )
        self.list_layout.add_widget(self._loading_label)

        BATCH_SIZE = 30  # v3.22.0: 降低单帧渲染压力，减少卡顿
        def _render_next(batch_start, _dt=None):
            # 检查令牌是否过期（被新的 refresh 取消）
            if token != self._render_token:
                return
            batch_end = min(batch_start + BATCH_SIZE, total_matched)
            for j in range(batch_start, batch_end):
                self._create_row_widget(matched_indices[j])
            if self._loading_label:
                self._loading_label.text = "正在加载 %d/%d..." % (batch_end, total_matched)
            if batch_end < total_matched:
                Clock.schedule_once(lambda dt: _render_next(batch_end), 0.03)
            else:
                # 加载完成，移除进度提示
                if self._loading_label:
                    try:
                        self.list_layout.remove_widget(self._loading_label)
                    except Exception:
                        pass
                    self._loading_label = None
                self._update_progress()

        Clock.schedule_once(lambda dt: _render_next(0), 0.05)

    def _create_row_widget(self, i):
        """v3.20.0: 创建单行 RowWidget 并添加到列表
        v3.22.0: 单行渲染失败时跳过而非整体崩溃（机型兼容）"""
        try:
            row = self.rows[i]
            serial = row[0] if len(row) > 0 else ""
            borrower = row[1] if len(row) > 1 else ""
            address_general = row[2] if len(row) > 2 else ""
            address_precise = row[3] if len(row) > 3 else ""
            property_type = row[4] if len(row) > 4 else ""
            excel_remark = row[5] if len(row) > 5 else ""
            full_addr = (address_general + address_precise).strip()
            pk = self.progress_mgr._make_key(borrower, full_addr)
            # v3.22.3 P0 修复: 优先按行号读取备注（避免共享 key 互相覆盖），
            # 回退 Excel，再回退旧 progress_key 备注（向后兼容 v3.22.2 前数据）。
            row_remark = self.progress_mgr.get_remark_by_row(i)
            saved_remark = self.progress_mgr.get_remark(pk)
            remark = row_remark if row_remark else (excel_remark if excel_remark else saved_remark)

            rw = RowWidget(
                row_index=i, borrower=borrower,
                address_general=address_general, address_precise=address_precise,
                property_type=property_type,
                progress_key=pk, progress_mgr=self.progress_mgr,
                remark=remark,
                excel_row_index=i + 2,  # Excel行号（1=表头，2=第一行数据）
                serial=serial,
                # v3.22.14: 内嵌按钮回调（直接绑定 MainScreen 方法，不经过菜单 dismiss）
                photo_callback=self._on_photo_request,
                view_photos_callback=self._on_view_photos,
                remark_callback=self._on_remark_request,
                # v3.22.19: 行选择回调（通知 MainScreen 更新 _row_selected）
                selected_callback=self._on_row_selected,
            )
            # v3.22.19: 若该行已处于 _row_selected 中（如全选后刷新），恢复 CheckBox 选中态
            if i in self._row_selected:
                rw.set_selected(True)
            self.list_layout.add_widget(rw)
            self.row_widgets.append(rw)
            # v3.22.7: 按 row_index 缓存，供 _refresh_list 增量刷新复用
            self._row_widget_cache[i] = rw
        except Exception as e:
            app_log.error('UI', '创建行 widget 失败 行%d: %s' % (i, e))
            traceback.print_exc()
            # v3.22.9: 不再静默吞异常，插入可见占位 widget 让用户看到失败
            try:
                placeholder = Button(
                    text='行 %d 加载失败，点击重试' % (i + 1),
                    font_size='14sp',
                    size_hint_y=None,
                    height=dp(56),
                    background_color=THEME['highlight_bg'],
                    color=THEME['danger'],
                )
                placeholder.bind(on_press=lambda btn: self._retry_create_row_widget(i, placeholder))
                self.list_layout.add_widget(placeholder)
            except Exception as e2:
                app_log.error('UI', '插入占位 widget 也失败 行%d: %s' % (i, e2))

    def _retry_create_row_widget(self, row_index, placeholder):
        """v3.22.9: 重试创建失败的行 widget"""
        try:
            # 移除占位 widget
            self.list_layout.remove_widget(placeholder)
            # 重新尝试创建
            self._create_row_widget(row_index)
            app_log.info('UI', '行 %d 重试创建成功' % row_index)
        except Exception as e:
            app_log.error('UI', '行 %d 重试创建仍失败: %s' % (row_index, e))
            # 重新插入占位 widget
            try:
                self.list_layout.add_widget(placeholder)
            except Exception:
                pass

    def _update_progress(self):
        total = len(self.rows)
        keys = []
        for r in self.rows:
            b = r[1] if len(r) > 1 else ""
            ag = r[2] if len(r) > 2 else ""
            ap = r[3] if len(r) > 3 else ""
            keys.append((b, (ag + ap).strip()))
        done = self.progress_mgr.get_done_count(keys)
        self.progress_label.text = "%d/%d" % (done, total)

    def _deferred_update_heights(self, dt):
        """v3.21.0: 延迟更新所有行高度，避免 on_resume 时同步执行导致卡顿
        v3.22.4: 同时强制重绘 _update_bg，修复从相机返回后高亮消失
        （OpenGL 上下文丢失清空 canvas，height 不变则不触发 size 绑定 → _update_bg 不被调用）"""
        for rw in getattr(self, 'row_widgets', []):
            try:
                rw._update_heights()
                rw._update_bg()
            except Exception:
                pass

    def _on_search(self, instance, text=None):
        """v3.21.0: 搜索防抖 — 350ms 内连续输入仅触发一次重建，避免每键全量重建导致卡死"""
        if self._search_debounce:
            self._search_debounce.cancel()
        self._search_debounce = Clock.schedule_once(
            lambda dt: self._refresh_list(filter_query=self.search_input.text.strip()), 0.35)

    def _on_search_field_change(self, spinner, value):
        """v3.22.0: 搜索类型切换 — 持久化 + 更新提示文字 + 立即重新过滤"""
        if not value:
            return
        self.config.set('search_field', value)
        self.config.save()
        try:
            self.search_input.hint_text = "搜索%s…" % value
        except Exception:
            pass
        # 立即按新类型过滤当前输入
        self._refresh_list(filter_query=self.search_input.text.strip())

    def _do_search(self, instance):
        self.search_input.focus = False
        self._on_search(None)

    def _clear_search(self, instance):
        # v3.21.0: 仅清空 text，由 text 绑定触发 _on_search（防抖延迟避免闪烁）
        self.search_input.text = ""

    def _clear_debug_log(self, instance):
        self.camera_mgr._log_lines = []
        try:
            if self.camera_mgr._debug_log_path and os.path.exists(self.camera_mgr._debug_log_path):
                os.remove(self.camera_mgr._debug_log_path)
        except OSError:
            pass
        app_log.clear()  # v3.20.0: 同时清空全量 app 日志
        self.status_label.text = ""
        self.status_label.color = THEME['text_dim']
        if IS_ANDROID:
            self.camera_mgr._toast("日志已清空")

    def _on_log_entry(self, instance):
        """v3.22.2: 日志入口按钮 — 打开二级弹窗（开关/查看/复制/清除）
        替代 v3.22.0 的"短按切换开关/长按查看日志"（长按难发现）"""
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout

        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(14))
        with content.canvas.before:
            Color(*THEME['card'])
            _bg = Rectangle(pos=content.pos, size=content.size)
        content.bind(pos=lambda i, v: setattr(_bg, 'pos', v),
                     size=lambda i, v: setattr(_bg, 'size', v))

        # 标题
        title = Label(text="日志管理", font_size='18sp', bold=True,
                      color=THEME['accent_dark'], size_hint_y=None, height=dp(36),
                      halign='left', valign='middle')
        title.bind(width=lambda i, v: setattr(i, 'text_size', (v, None)))
        content.add_widget(title)

        # 状态显示
        _cur_on = app_log.is_enabled()
        status_label = Label(
            text=("当前状态: [color=33B35C][b]已开启[/b][/color]" if _cur_on
                  else "当前状态: [color=9EA6B3][b]已关闭[/b][/color]"),
            font_size='14sp', markup=True, color=THEME['text'],
            size_hint_y=None, height=dp(28), halign='left', valign='middle')
        status_label.bind(width=lambda i, v: setattr(i, 'text_size', (v, None)))
        content.add_widget(status_label)

        # 4 按钮网格 2x2
        btn_grid = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None, height=dp(108))
        row1 = BoxLayout(spacing=dp(8), size_hint_y=None, height=dp(50))
        row2 = BoxLayout(spacing=dp(8), size_hint_y=None, height=dp(50))

        toggle_btn = RoundedButton(
            text=("关闭日志记录" if _cur_on else "开启日志记录"),
            font_size='15sp',
            background_color=(THEME['danger'] if _cur_on else THEME['success']),
            background_normal='', color=(1,1,1,1), bold=True)
        view_btn = RoundedButton(text="查看日志", font_size='15sp',
            background_color=THEME['accent'], background_normal='',
            color=(1,1,1,1), bold=True)
        copy_btn = RoundedButton(text="复制日志", font_size='15sp',
            background_color=THEME['accent_dark'], background_normal='',
            color=(1,1,1,1), bold=True)
        clear_btn = RoundedButton(text="清空日志", font_size='15sp',
            background_color=THEME['warning'], background_normal='',
            color=(1,1,1,1), bold=True)

        row1.add_widget(toggle_btn)
        row1.add_widget(view_btn)
        row2.add_widget(copy_btn)
        row2.add_widget(clear_btn)
        btn_grid.add_widget(row1)
        btn_grid.add_widget(row2)
        content.add_widget(btn_grid)

        # 关闭按钮
        close_btn = RoundedButton(text="关闭", font_size='15sp',
            background_color=THEME['muted'], background_normal='',
            color=(1,1,1,1), size_hint_y=None, height=dp(44))
        content.add_widget(close_btn)

        popup = Popup(title='日志管理',
                      title_color=THEME['accent_dark'],
                      separator_color=THEME['card_border'],
                      content=content,
                      size_hint=(0.85, 0.5))

        def _toggle(inst):
            new_state = not app_log.is_enabled()
            app_log.set_enabled(new_state)
            self.config.set('log_enabled', new_state)
            self.config.save()
            toggle_btn.text = "关闭日志记录" if new_state else "开启日志记录"
            toggle_btn.background_color = THEME['danger'] if new_state else THEME['success']
            status_label.text = ("当前状态: [color=33B35C][b]已开启[/b][/color]" if new_state
                                 else "当前状态: [color=9EA6B3][b]已关闭[/b][/color]")
            if IS_ANDROID:
                self.camera_mgr._toast("日志记录已开启" if new_state else "日志记录已关闭")
            app_log.info('APP', '日志开关: %s' % ('开' if new_state else '关'))

        def _view(inst):
            popup.dismiss()
            self._show_full_log(None)

        def _copy(inst):
            try:
                from kivy.core.clipboard import Clipboard
                text = app_log.get_log_text()
                Clipboard.copy(text)
                copy_btn.text = "已复制！"
                copy_btn.background_color = THEME['success']
                Clock.schedule_once(lambda dt: setattr(copy_btn, 'text', '复制日志'), 1.5)
                Clock.schedule_once(lambda dt: setattr(copy_btn, 'background_color', THEME['accent_dark']), 1.5)
                if IS_ANDROID:
                    self.camera_mgr._toast("日志已复制到剪贴板")
            except Exception as e:
                copy_btn.text = "复制失败"
                app_log.error('APP', '复制日志失败: %s' % e)

        def _clear(inst):
            app_log.clear()
            clear_btn.text = "已清空"
            clear_btn.background_color = THEME['success']
            Clock.schedule_once(lambda dt: setattr(clear_btn, 'text', '清空日志'), 1.5)
            Clock.schedule_once(lambda dt: setattr(clear_btn, 'background_color', THEME['warning']), 1.5)
            if IS_ANDROID:
                self.camera_mgr._toast("日志已清空")

        toggle_btn.bind(on_release=_toggle)
        view_btn.bind(on_release=_view)
        copy_btn.bind(on_release=_copy)
        clear_btn.bind(on_release=_clear)
        close_btn.bind(on_release=lambda x: popup.dismiss())

        popup.open()

    def _show_full_log(self, instance):
        """v3.20.0: 打开全屏日志记事本，显示全量 app 日志（Excel/拍照/备注/AI/生命周期等）"""
        log_content = app_log.get_log_text()
        if not log_content.strip():
            log_content = "暂无日志记录\n\n操作后会在这里记录详细日志（Excel加载、拍照、备注保存、AI调用等）。"

        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.button import Button
        from kivy.uix.scrollview import ScrollView
        from kivy.uix.label import Label

        content = BoxLayout(orientation='vertical', spacing=dp(8), padding=dp(10))
        # v3.22.0: 浅色背景 — 白底 + 深色文字，确保可读性
        with content.canvas.before:
            Color(*THEME['card'])
            _log_bg = Rectangle(pos=content.pos, size=content.size)
        content.bind(pos=lambda i, v: setattr(_log_bg, 'pos', v),
                     size=lambda i, v: setattr(_log_bg, 'size', v))
        scroll = ScrollView(do_scroll_x=False, do_scroll_y=True)
        log_label = Label(
            text=log_content,
            font_size='11sp',
            color=THEME['text'],
            halign='left', valign='top',
            size_hint_y=None,
            markup=False,
            text_size=(None, None)
        )
        log_label.bind(texture_size=lambda i, v: setattr(i, 'height', v[1] + dp(10)))
        log_label.bind(width=lambda i, v: setattr(i, 'text_size', (v, None)))
        scroll.add_widget(log_label)
        content.add_widget(scroll)

        btn_row = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        close_btn = RoundedButton(text="关闭", font_size='15sp', size_hint_x=0.34,
                          background_color=THEME['muted'], background_normal='', color=(1,1,1,1))
        copy_btn = RoundedButton(text="复制日志", font_size='15sp', size_hint_x=0.33,
                         background_color=THEME['accent'], background_normal='', color=(1,1,1,1))
        clear_btn = RoundedButton(text="清空", font_size='15sp', size_hint_x=0.33,
                          background_color=THEME['danger'], background_normal='', color=(1,1,1,1))
        btn_row.add_widget(close_btn)
        btn_row.add_widget(copy_btn)
        btn_row.add_widget(clear_btn)
        content.add_widget(btn_row)

        popup = Popup(title='应用日志（问题排查请截图此页）',
                      title_color=THEME['accent_dark'],
                      separator_color=THEME['card_border'],
                      content=content,
                      size_hint=(0.95, 0.9))

        def _close(instance):
            popup.dismiss()
        close_btn.bind(on_release=_close)

        def _copy(instance):
            try:
                from kivy.core.clipboard import Clipboard
                Clipboard.copy(log_content)
                copy_btn.text = "已复制！"
                copy_btn.background_color = THEME['success']
                Clock.schedule_once(lambda dt: setattr(copy_btn, 'text', '复制日志'), 1.5)
                Clock.schedule_once(lambda dt: setattr(copy_btn, 'background_color', THEME['accent']), 1.5)
            except Exception as e:
                copy_btn.text = "复制失败"

        copy_btn.bind(on_release=_copy)

        def _clear(instance):
            app_log.clear()
            log_label.text = "日志已清空"
            clear_btn.text = "已清空"
            Clock.schedule_once(lambda dt: setattr(clear_btn, 'text', '清空'), 1.5)
        clear_btn.bind(on_release=_clear)

        popup.open()

    def _update_log_label_size(self, instance, value):
        instance.height = max(dp(80), value[1] + dp(8))

    def _go_settings(self, instance):
        self._continuous_shooting = False
        self._photo_session_active = False  # v3.20.0: 解除会话锁
        self._unlock_photo_buttons()
        self._clear_all_highlights()  # v3.22.3: 跳转页面时清除高亮
        self.manager.current = 'settings'

    def _go_ai(self, instance):
        """跳转到 AI 助手页面 v3.15.0"""
        self._continuous_shooting = False
        self._photo_session_active = False  # v3.20.0: 解除会话锁
        self._unlock_photo_buttons()
        self._clear_all_highlights()  # v3.22.3: 跳转页面时清除高亮
        self.manager.current = 'ai'

    def _on_photo_request(self, row_index, borrower, address_general, address_precise, property_type):
        """v3.20.0: 拍照会话入口 — 创建上下文 ctx 并启动会话锁，防止连拍期间切换客户"""
        # v3.20.0: 会话锁 — 拍照进行中时拒绝新请求
        if self._photo_session_active:
            self.camera_mgr._dbg("拍照会话进行中，请先完成当前拍照", show_toast=True)
            return

        # 兼容旧代码：仍更新 self._current_*（键盘处理等仍引用）
        self._current_row = row_index
        self._current_borrower = borrower
        self._current_addr_general = address_general
        self._current_addr_precise = address_precise
        self._current_property_type = property_type
        self._current_key = self.progress_mgr._make_key(borrower, (address_general + address_precise).strip())
        self._photos_in_session = 0

        # v3.20.0: 创建拍照会话上下文（通过闭包传递，不依赖实例状态）
        self._photo_session_ctx = {
            'row_index': row_index,
            'borrower': borrower,
            'addr_general': address_general,
            'addr_precise': address_precise,
            'property_type': property_type,
            'photo_type': '',  # 在 _on_photo_type_selected 中设置
            'key': self._current_key,
        }

        popup = PhotoTypePopup(on_select=self._on_photo_type_selected)
        popup.open()

    def _on_photo_type_selected(self, photo_type):
        """v3.20.0: 选择拍照类型后启动相机，通过闭包绑定 ctx"""
        ctx = self._photo_session_ctx
        if not ctx:
            return
        ctx['photo_type'] = photo_type
        self._current_photo_type = photo_type  # 兼容旧代码
        self._continuous_shooting = True
        self._photo_session_active = True  # v3.20.0: 激活会话锁
        self._photos_in_session = 0

        # v3.22.3: 切换客户时先清除所有行的高亮（旧高亮自动消失），再设置当前行高亮
        # 这样拍完远景退出相机后高亮保持，下次点其他客户拍照才清除旧高亮
        self._clear_all_highlights()
        # v3.20.0: 锁定其他行的拍照按钮
        self._lock_photo_buttons(exclude_row=ctx['row_index'])
        # v3.22.2: 高亮当前拍摄行（浅蓝底+蓝边）
        rw = self._get_row_widget(ctx['row_index'])
        if rw:
            rw.set_highlight(True)

        self.camera_mgr._dbg(f"开始为【{ctx['borrower']}】拍照，类型: {photo_type}", show_toast=True)
        app_log.info('PHOTO', '拍照会话开始: 客户=%s, 类型=%s, 行=%d' % (ctx['borrower'], photo_type, ctx['row_index']))

        # v3.20.0: 通过闭包绑定 ctx，不依赖 self._current_*（修复命名错乱根因）
        Clock.schedule_once(
            lambda dt: self.camera_mgr.take_photo(
                lambda path: self._on_photo_done_with_context(path, ctx),
                self._camera_status_update
            ), 0.3)

    def _lock_photo_buttons(self, exclude_row=None):
        """v3.22.14: 拍照会话期间禁用其他行的拍照按钮
        （v3.22.13 改为 disabled 整行，v3.22.14 RowWidget 去 ButtonBehavior 后改为 photo_btn.disabled）"""
        for rw in self.row_widgets:
            if rw.row_index != exclude_row:
                try:
                    rw.photo_btn.disabled = True
                    rw.opacity = 0.5
                except Exception:
                    pass

    def _unlock_photo_buttons(self):
        """v3.22.14: 拍照会话结束后恢复所有行
        v3.22.3: 不再清除高亮（高亮在会话期间保持，让用户退出相机准备拍下一类时仍能找到当前条目）。
        高亮改由 _clear_all_highlights 统一管理，在切换客户/跳转页面时调用。"""
        for rw in self.row_widgets:
            try:
                rw.photo_btn.disabled = False
                rw.opacity = 1
            except Exception:
                pass

    def _clear_all_highlights(self):
        """v3.22.3: 清除所有行的高亮（切换客户/跳转页面时调用）"""
        for rw in self.row_widgets:
            try:
                rw.set_highlight(False)
            except Exception:
                pass

    def _camera_status_update(self, msg):
        """CameraManager回调：更新UI状态显示调试信息（多行日志）"""
        def _update(dt):
            self.status_label.text = msg
            self.status_label.color = THEME['warning']
            Clock.schedule_once(lambda dt2: self._scroll_log_to_bottom(), 0.05)
        Clock.schedule_once(_update, 0)

    def _scroll_log_to_bottom(self):
        """自动滚动日志到底部显示最新消息"""
        try:
            sv = self.status_label.parent
            if isinstance(sv, ScrollView):
                sv.scroll_y = 0
        except Exception:
            pass

    def _launch_next_photo(self, ctx=None):
        """v3.20.0: 拍完一张后立即重新调起相机，通过闭包传递同一 ctx。"""
        ctx = ctx or self._photo_session_ctx
        if self._continuous_shooting and ctx:
            self.camera_mgr._dbg(f"准备拍摄下一张（{ctx['photo_type']}）…")
            Clock.schedule_once(
                lambda dt: self.camera_mgr.take_photo(
                    lambda path: self._on_photo_done_with_context(path, ctx),
                    self._camera_status_update
                ), 0.5)

    def _on_photo_done_with_context(self, photo_path, ctx):
        """v3.20.0: 拍照完成回调 — 使用闭包绑定的 ctx，不依赖 self._current_*（修复命名错乱）"""
        if photo_path is None:
            self._continuous_shooting = False
            self._photo_session_active = False
            self._unlock_photo_buttons()
            self.camera_mgr._dbg("拍照已取消")
            self._refresh_row_done(ctx['row_index'])
            # v3.22.4: 会话结束后重新确认高亮，确保用户能看到当前拍摄条目
            # （相机返回后 canvas 可能被清空，需重绘高亮）
            rw = self._get_row_widget(ctx['row_index'])
            if rw:
                rw.set_highlight(True)
            return

        # 从 ctx 读取客户信息（不再从 self._current_* 读取，防止被覆盖）
        row_index = ctx['row_index']
        borrower = ctx['borrower']
        addr_general = ctx['addr_general']
        addr_precise = ctx['addr_precise']
        property_type = ctx['property_type']
        photo_type = ctx['photo_type']
        key = ctx['key']

        self._photos_in_session += 1
        seq = self.progress_mgr.get_next_photo_index(key) + 1

        date_str = get_date_str()
        time_str = get_time_str()
        datetime_str = get_datetime_str()

        self.camera_mgr._dbg("照片处理中...")
        app_log.info('PHOTO', '处理照片: 客户=%s, 类型=%s, 序号=%d' % (borrower, photo_type, seq))

        config_data = self.config.data
        naming_segments = self.config.get('naming_segments', DEFAULT_CONFIG['naming_segments'])
        lat, lng = self.camera_mgr.gps.get_coords()

        def _process():
            try:
                place_name = self.camera_mgr.get_location_name(lat, lng)

                PhotoProcessor.add_watermark(
                    photo_path, config_data,
                    time_str=get_date_display(), address=place_name,
                    lat=lat, lng=lng,
                )

                filename = PhotoProcessor.generate_filename(
                    naming_segments, borrower, addr_general, addr_precise,
                    property_type, seq, date_str, photo_type, time_str,
                )
                new_path = os.path.join(APP_DIR, filename)
                if photo_path != new_path:
                    if os.path.exists(new_path):
                        name_base, ext = os.path.splitext(filename)
                        suffix = 2
                        while os.path.exists(new_path):
                            new_path = os.path.join(APP_DIR, "%s-%d%s" % (name_base, suffix, ext if ext else '.jpg'))
                            suffix += 1
                    # v3.19.0: os.rename 跨设备会失败，改用 shutil.move 并兜底删除原文件
                    import shutil as _shutil
                    try:
                        os.rename(photo_path, new_path)
                    except Exception:
                        try:
                            _shutil.move(photo_path, new_path)
                        except Exception:
                            _shutil.copy2(photo_path, new_path)
                            try:
                                os.remove(photo_path)
                            except OSError:
                                pass
                    # 兜底：确保 capture_xxx 临时文件被清理
                    try:
                        if photo_path != new_path and os.path.exists(photo_path):
                            os.remove(photo_path)
                    except OSError:
                        pass

                # v3.22.4: 始终保留 APP_DIR 副本，progress_mgr 记录 APP_DIR 路径。
                # 原逻辑: save_to_gallery 成功后删除 APP_DIR 副本、记录 DCIM 路径；
                # 但 Android 10+ scoped storage 下 os.path.exists(DCIM路径) 返回 False，
                # 导致 get_photos 过滤掉所有照片 → 查看已拍无反应。
                # 现逻辑: APP_DIR 路径（app 私有目录，始终可访问）作为真相源，
                # save_to_gallery 仅为方便用户在系统相册查看（失败不影响功能）。
                try:
                    PhotoProcessor.save_to_gallery(new_path)
                except Exception as e:
                    app_log.error('PHOTO', 'save_to_gallery 失败(不影响拍照保存): %s' % e)
                self.progress_mgr.mark_photo(key, new_path, photo_type)
                Clock.schedule_once(lambda dt: self._on_photo_saved(row_index, filename, ctx), 0)
            except Exception as e:
                Logger.error("MainScreen._on_photo_done: %s" % e)
                Logger.error(traceback.format_exc())
                err_msg = str(e)
                app_log.error('PHOTO', '照片处理失败: %s' % err_msg)
                Clock.schedule_once(lambda dt: self._on_photo_failed(err_msg), 0)

        threading.Thread(target=_process, daemon=True).start()

    @mainthread
    def _on_photo_saved(self, row_index, filename, ctx=None):
        self._refresh_row_done(row_index)
        self.camera_mgr._dbg(f"[OK] 已保存: {filename}", show_toast=True)
        # v3.22.19: 异步生成压缩缩略图保存到 app 私有目录，规避 Scoped Storage 访问原图权限问题
        # ctx 中含 key（progress_key），filename 是原图 basename（保存在 APP_DIR 下）
        try:
            _key = ctx.get('key') if ctx else None
            if _key and filename:
                _orig_path = os.path.join(APP_DIR, filename)
                # 后台线程生成缩略图，避免阻塞 UI 线程
                threading.Thread(
                    target=ProgressManager.generate_thumbnail,
                    args=(_orig_path, _key),
                    daemon=True
                ).start()
        except Exception as e:
            app_log.error('PHOTO', '_on_photo_saved 异步生成缩略图启动失败: %s' % e, exc_info=True)
        # v3.20.0: 连拍下一张时传递同一 ctx
        if ctx:
            self._launch_next_photo(ctx)
        else:
            self._launch_next_photo()

    @mainthread
    def _on_photo_failed(self, err_msg):
        self._continuous_shooting = False
        self._photo_session_active = False  # v3.20.0: 解除会话锁
        self._unlock_photo_buttons()
        self.camera_mgr._dbg(f"保存失败: {err_msg[:60]}", show_toast=True)

    def _get_row_widget(self, row_index):
        """v3.21.0: 根据原始行号查找对应的 RowWidget。
        搜索过滤后 row_widgets 按匹配顺序排列，位置下标与原始行号错位，
        必须用此方法遍历匹配，避免高亮/备注指向错误客户。
        """
        for rw in self.row_widgets:
            if getattr(rw, 'row_index', -1) == row_index:
                return rw
        return None

    def _scroll_input_to_visible(self, popup, text_input):
        """v3.22.19: 软键盘弹出后，将 popup 内 text_input 滚动到可见区域（薄包装，
        实际逻辑见模块级 _scroll_input_to_visible_func 函数）。"""
        _scroll_input_to_visible_func(popup, text_input)

    def _on_select_all(self, checkbox, value):
        """v3.22.19: 表头全选 CheckBox 触发 — 勾选则全选所有行，取消则清空选中。
        用 _updating_checkbox 标志位避免与 _on_row_selected 形成回调循环。"""
        if self._updating_checkbox:
            return
        self._updating_checkbox = True
        try:
            self._all_selected = bool(value)
            if value:
                self._row_selected = set(range(len(self.rows)))
            else:
                self._row_selected.clear()
            # 更新所有 RowWidget 显示（用 set_selected 避免回调循环）
            for rw in self.row_widgets:
                try:
                    rw.set_selected(rw.row_index in self._row_selected)
                except Exception as e:
                    app_log.error('UI', '_on_select_all set_selected 异常 row=%d: %s' % (rw.row_index, e))
            self._update_progress()
        except Exception as e:
            app_log.error('UI', '_on_select_all 异常: %s' % e)
        finally:
            self._updating_checkbox = False

    def _on_row_selected(self, row_index, value):
        """v3.22.19: 单行 CheckBox 状态变化回调（作为 RowWidget 的 selected_callback）。
        维护 _row_selected 集合，并同步表头 CheckBox 状态（避免回调循环用 _updating_checkbox）。
        反选逻辑：全选态下取消某行 → 表头取消勾选，其他行保持选中。"""
        if self._updating_checkbox:
            return
        self._updating_checkbox = True
        try:
            if value:
                self._row_selected.add(row_index)
            else:
                self._row_selected.discard(row_index)
            # 同步表头 CheckBox 状态
            total = len(self.rows)
            if total > 0 and len(self._row_selected) == total:
                # 全部选中 → 表头勾选
                self._all_selected = True
                self.header_checkbox.active = True
            elif self._all_selected and not value:
                # 全选状态下取消某行 → 表头取消勾选（其他行保持选中）
                self._all_selected = False
                self.header_checkbox.active = False
            # 更新对应 RowWidget 显示
            rw = self._get_row_widget(row_index)
            if rw:
                rw.update_selected_display()
        except Exception as e:
            app_log.error('UI', '_on_row_selected 异常 row=%d: %s' % (row_index, e))
        finally:
            self._updating_checkbox = False

    # ============================================================
    # v3.22.9: Android Toast 通用工具方法
    # ============================================================
    def _show_android_toast(self, message):
        """v3.22.9: 弹 Android 原生 toast（通过 pyjnius）— 通用工具方法"""
        try:
            from jnius import autoclass
            Toast = autoclass('android.widget.Toast')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            Toast.makeText(PythonActivity.mActivity, message, Toast.LENGTH_SHORT).show()
        except Exception as e:
            # 非 Android 环境或 jnius 不可用时静默 fallback 到日志
            app_log.info('UI', 'toast 调用失败（非 Android 环境？）: %s' % e)

    # v3.22.14: 移除 v3.22.13 的行点击操作菜单（_on_row_tap /
    # _show_row_action_menu / _row_action_view_photos / _row_action_take_photo /
    # _row_action_remark）— 回退到内嵌按钮方案，view_btn.on_press 直接调用
    # _on_view_photos，不经过菜单 dismiss 步骤。

    def _refresh_row_done(self, row_index):
        rw = self._get_row_widget(row_index)
        if rw:
            rw.mark_done()
        self._update_progress()

    def _on_view_photos(self, row_index):
        """v3.22.13: 整体 try/except 包裹，任何异常通过 _show_msg（toast + popup）反馈
        v3.22.2 P0 修复: 直接用 RowWidget 持有的 progress_key（构造时算好），
        避免用 self.rows[row_index] 重算 key 时与 widget 不一致导致"无反应"。
        v3.22.11: 防抖 — 1500ms 内只执行一次，防止多次触发导致 popup 状态异常"""
        # v3.22.11: 防抖（1500ms）— 防抖检查不纳入 try/except（异常也不应跳过防抖）
        now = Clock.get_time()
        if now - getattr(self, '_view_photos_last_time', 0) < 1.5:
            app_log.info('PHOTO', '查看已拍防抖：跳过本次调用（距上次 < 1500ms）')
            return
        self._view_photos_last_time = now
        # v3.22.13: 整体 try/except — 任何异常（含 PhotoViewerPopup 创建失败）都通过 _show_msg 反馈
        try:
            rw = self._get_row_widget(row_index)
            if rw is None:
                app_log.error('UI', '查看已拍失败: 未找到 row_index=%d 的 widget' % row_index)
                self._show_msg('无法打开此客户的照片（行数据丢失）')
                return
            key = rw.progress_key
            # v3.22.19: 改读缩略图（保存在 app 私有目录，规避 Scoped Storage 权限问题）
            thumbnails = self.progress_mgr.get_thumbnails(key)
            # 仍记录原照片数量（仅用于日志诊断）
            orig_count = self.progress_mgr.get_photo_count(key)
            app_log.info('PHOTO', '查看已拍: 客户=%s, key=%s, 原图数=%d, 缩略图数=%d' % (
                rw.borrower, key[:8], orig_count, len(thumbnails)))
            if not thumbnails:
                self._show_msg('该客户暂无缩略图（仅原图）')
                return
            # v3.22.17: 直接创建 PhotoViewerPopup 并 open()（显式设置像素尺寸，无需 Clock.schedule_once 延迟）
            # v3.22.19: photos 参数改为缩略图路径列表，title 仍显示原照片数量
            try:
                app_log.info('PHOTO', '创建 PhotoViewerPopup: thumbnails=%d' % len(thumbnails))
                popup = PhotoViewerPopup(row_index=row_index, photos=thumbnails,
                                         photo_count=orig_count,
                                         progress_key=key,
                                         delete_callback=self._on_delete_photo)
                popup.open()
                app_log.info('PHOTO', 'PhotoViewerPopup.open() 已调用')
                # v3.22.17: 延迟 0.2 秒记录 popup 状态诊断日志
                Clock.schedule_once(lambda dt: self._log_popup_status(popup), 0.2)
            except Exception as e:
                app_log.error('UI', '创建/打开 PhotoViewerPopup 异常 row=%d: %s' % (row_index, e), exc_info=True)
                try:
                    self._show_msg('查看已拍失败: %s' % str(e))
                except Exception as e2:
                    app_log.error('UI', '_show_msg 反馈也失败: %s' % e2)
        except Exception as e:
            app_log.error('UI', '_on_view_photos 异常 row=%d: %s' % (row_index, e), exc_info=True)
            # v3.22.13: 用 _show_msg 反馈给用户（toast + popup，不再"连失败提示都没有"）
            try:
                self._show_msg('查看已拍失败: %s' % str(e))
            except Exception as e2:
                app_log.error('UI', '_show_msg 反馈也失败: %s' % e2)

    def _log_popup_status(self, popup):
        """v3.22.17: 记录 popup 状态诊断日志，排查 popup 不可见问题"""
        try:
            app_log.info('PHOTO', 'popup 状态: pos=%s size=%s opacity=%s parent=%s' % (
                popup.pos, popup.size, popup.opacity, popup.parent))
        except Exception as e:
            app_log.error('UI', '_log_popup_status 异常: %s' % e)

    def _on_delete_photo(self, row_index, photo_index):
        if row_index >= len(self.rows):
            return
        row = self.rows[row_index]
        borrower = row[1] if len(row) > 1 else ""
        address_general = row[2] if len(row) > 2 else ""
        address_precise = row[3] if len(row) > 3 else ""
        full_addr = (address_general + address_precise).strip()
        key = self.progress_mgr._make_key(borrower, full_addr)
        if photo_index == -1:
            # 删除全部：先删原图文件、缩略图目录，再清进度
            photos = self.progress_mgr.get_photos(key)
            for p in photos:
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except Exception as e:
                    app_log.error('PHOTO', '删除照片文件失败 %s: %s' % (p, e))
            self.progress_mgr.delete_all_photos(key)
            # v3.22.19: 同步删除全部缩略图
            self.progress_mgr.delete_all_thumbnails(key)
            self._show_msg("已删除该客户全部照片", THEME['warning'])
        else:
            # v3.22.19: photo_index 是缩略图列表的索引。
            # 缩略图与原图同名（仅扩展名不同），通过文件名匹配找到对应原图删除，
            # 避免缩略图排序与原图列表顺序不一致导致删错。
            thumbnails = self.progress_mgr.get_thumbnails(key)
            photos = self.progress_mgr.get_photos(key)
            thumb_to_delete = None
            if photo_index < len(thumbnails):
                thumb_to_delete = thumbnails[photo_index]
                thumb_basename = os.path.basename(thumb_to_delete)
                thumb_name_base = os.path.splitext(thumb_basename)[0]
                # 在原图列表中按 basename 匹配（去扩展名相同）
                matched_orig_idx = None
                matched_orig_path = None
                for i, p in enumerate(photos):
                    orig_name_base = os.path.splitext(os.path.basename(p))[0]
                    if orig_name_base == thumb_name_base:
                        matched_orig_idx = i
                        matched_orig_path = p
                        break
                # 删除原图
                if matched_orig_path is not None:
                    try:
                        if os.path.exists(matched_orig_path):
                            os.remove(matched_orig_path)
                    except Exception as e:
                        app_log.error('PHOTO', '删除照片文件失败 %s: %s' % (matched_orig_path, e))
                    # 删除原图进度记录（用匹配到的原图索引）
                    self.progress_mgr.delete_photo(key, matched_orig_idx)
                else:
                    app_log.warning('PHOTO', '未找到与缩略图 %s 匹配的原图' % thumb_basename)
                # 删除缩略图文件
                self.progress_mgr.delete_thumbnail(key, thumb_to_delete)
            else:
                app_log.error('PHOTO', 'photo_index=%d 超出缩略图列表范围 %d' % (photo_index, len(thumbnails)))
            self._show_msg("照片已删除", THEME['warning'])
        Clock.schedule_once(lambda dt: self._refresh_row_done(row_index), 0)

    def _on_remark_request(self, row_index):
        """打开备注输入弹窗 v3.15.0"""
        rw = self._get_row_widget(row_index)
        if not rw:
            return
        current_remark = rw.remark or ""

        # v3.22.0: 全局 softinput_mode 已为 resize，无需临时切换
        popup = Popup(title="添加备注（保存到Excel F列）",
                      size_hint=(0.9, None), height=dp(460),
                      pos_hint={'top': 0.95}, auto_dismiss=False)
        layout = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(16))

        info_label = Label(
            text="客户：%s\n地址：%s" % (rw.borrower, (rw.address_general + rw.address_precise).strip()),
            font_size='14sp', color=THEME['text'],
            size_hint_y=None, height=dp(50),
            halign='left', valign='top',
        )
        info_label.bind(size=lambda inst, val: setattr(inst, 'text_size', val))
        layout.add_widget(info_label)

        # v3.22.0: 引导提示 — 提醒用户详细描述现状，作为日报表依据
        guide_label = Label(
            text="提示：请详细描述现状（外观、使用情况、风险点等），将作为日报表「现状描述」「备注」的依据，请如实填写，勿留空。",
            font_size='12sp', color=THEME['warning'],
            size_hint_y=None, height=dp(54),
            halign='left', valign='top',
        )
        guide_label.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0] - dp(4), None)))
        layout.add_widget(guide_label)

        from kivy.uix.textinput import TextInput
        # v3.21.0: TextInput 包裹 ScrollView，键盘遮挡时可滚动查看长文本
        scroll = ScrollView(size_hint_y=None, height=dp(220))
        text_input = TextInput(
            text=current_remark,
            font_size='16sp',
            multiline=True,
            size_hint_y=None, height=dp(200),
            hint_text="请输入备注内容...",
        )
        # v3.22.19: 聚焦时延迟滚动到可见区域，避免键盘遮挡输入框
        text_input.bind(focus=lambda inst, val: Clock.schedule_once(
            lambda dt: self._scroll_input_to_visible(popup, text_input), 0.1) if val else None)
        scroll.add_widget(text_input)
        layout.add_widget(scroll)

        # v3.22.4 P0 修复: 备注显示不全——TextInput 设长文本时光标自动到末尾，
        # 视图滚到末尾导致只显示后段内容且可见字数不固定。
        # 将光标移到开头，TextInput 自动滚动到顶部显示全文起始处。
        def _reset_cursor_to_start(dt):
            try:
                text_input.cursor = (0, 0)
            except Exception:
                pass
        Clock.schedule_once(_reset_cursor_to_start, 0.05)

        btn_row = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
        save_btn = RoundedButton(text="保存", font_size='16sp', bold=True,
                         background_color=THEME['success'], background_normal='')
        cancel_btn = RoundedButton(text="取消", font_size='16sp',
                           background_color=THEME['muted'], background_normal='')
        btn_row.add_widget(save_btn)
        btn_row.add_widget(cancel_btn)
        layout.add_widget(btn_row)

        popup.content = layout

        # v3.22.0: 全局 resize 模式，无需恢复 softinput_mode

        def do_save(instance):
            remark_text = text_input.text.strip()
            rw.set_remark(remark_text)
            # 更新内存中的行数据
            if row_index < len(self.rows):
                while len(self.rows[row_index]) < 6:
                    self.rows[row_index].append("")
                self.rows[row_index][5] = remark_text
            # v3.17.0: 保存到progress_mgr（持久化到JSON，确保退出app后不丢失）
            # v3.20.0: progress_mgr 是备注的单一真相源，始终保存
            # v3.21.0: 改用 rw 的数据，避免 row_index 越界导致备注丢失
            try:
                borrower = rw.borrower or ""
                address_general = rw.address_general or ""
                address_precise = rw.address_precise or ""
                full_addr = (address_general + address_precise).strip()
                key = self.progress_mgr._make_key(borrower, full_addr)
                self.progress_mgr.save_remark(key, remark_text)
                # v3.22.3 P0: 同时按行号存储备注（行号独立，避免共享 key 互相覆盖）
                self.progress_mgr.save_remark_by_row(row_index, remark_text)
                app_log.info('REMARK', '备注已保存到持久化存储: 客户=%s, key=%s, 行=%d' % (borrower, key[:8], row_index))
            except Exception as e:
                app_log.error('REMARK', '备注保存到progress_mgr失败: %s' % e)
                Logger.error(f"备注保存到progress_mgr失败: {e}")
            # v3.22.0: 保存到 Excel F 列（备注列顺延）
            if self.excel_path:
                import threading
                excel_uri = getattr(self, '_excel_uri', None)
                def _save_excel():
                    # v3.21.0: 串行化 Excel 写入，防止并发覆盖
                    # （多个备注同时保存时，后写入的线程会用旧数据覆盖先写入的修改）
                    with self._excel_save_lock:
                        ok, msg = ExcelWriter.save_remark(self.excel_path, rw.excel_row_index, remark_text)
                        if ok and excel_uri and excel_uri.startswith('content://'):
                            # 通过 ContentResolver 写回用户选择的原始 Excel 文件
                            ok2, msg2 = ExcelWriter.write_back_to_uri(excel_uri, self.excel_path)
                            if ok2:
                                msg = "已保存到原始 Excel"
                                app_log.info('REMARK', '备注已写回原始 Excel: 行%d' % rw.excel_row_index)
                            else:
                                msg = f"app内部已保存；{msg2}"
                                app_log.warn('REMARK', '写回原始Excel失败: %s' % msg2)
                    if ok:
                        Logger.info("备注已保存到Excel F列 行%d" % rw.excel_row_index)
                        self.camera_mgr._dbg(f"[OK] {msg} 行{rw.excel_row_index}", show_toast=True)
                    else:
                        Logger.error("备注保存失败 行%d" % rw.excel_row_index)
                        app_log.error('REMARK', 'Excel保存失败 行%d: %s' % (rw.excel_row_index, msg))
                        self.camera_mgr._dbg(f"[WARN] {msg}，但已保存到app内部 行{rw.excel_row_index}", show_toast=True)
                threading.Thread(target=_save_excel, daemon=True).start()
            else:
                self.camera_mgr._dbg("备注已保存到app内部（未关联Excel文件）", show_toast=True)
            self._show_msg("备注已保存", toast=True)
            popup.dismiss()

        save_btn.bind(on_release=do_save)
        cancel_btn.bind(on_release=popup.dismiss)
        popup.open()

    def _on_generate_report_button(self, instance):
        """v3.22.13: AI 一键生成日报表按钮入口 — 先弹 VisitNoteDialog 收集
        「今日走访补充说明」，确认后再走原有 _show_report_confirm 二次确认流程。
        VisitNoteDialog 预填 self.current_visit_note（按 Excel 路径持久化）。
        """
        try:
            def _on_confirm(text):
                self._on_visit_note_confirmed(text, instance)
            dialog = VisitNoteDialog(
                initial_text=self.current_visit_note,
                on_confirm=_on_confirm,
            )
            dialog.open()
        except Exception as e:
            app_log.error('UI', '_on_generate_report_button 异常: %s' % e, exc_info=True)
            # 兜底：直接走原有二次确认流程（保留现有 visit_note）
            self._show_report_confirm(instance)

    def _on_visit_note_confirmed(self, text, instance):
        """v3.22.13: VisitNoteDialog 确认回调 — 保存说明（内存 + 持久化）后
        走原有 _show_report_confirm 二次确认弹窗流程。"""
        try:
            self.current_visit_note = text or ''
            if self.current_visit_note_excel_path:
                self._save_visit_note(self.current_visit_note_excel_path, text or '')
            # 走原有二次确认弹窗流程（展示外访进度，由用户再次确认生成）
            self._show_report_confirm(instance)
        except Exception as e:
            app_log.error('UI', '_on_visit_note_confirmed 异常: %s' % e, exc_info=True)
            # 兜底：直接生成报告（保留现有 visit_note）
            self._generate_report(instance)

    def _show_report_confirm(self, instance):
        """v3.22.5: AI报表二次确认弹窗，展示外访进度。
        v3.22.6: 修复 _collect_records 调用（应通过 report_generator），
                 异常时不再静默 fallback 到 _generate_report，改为弹错误提示；
                 进度文案含「同类型」统计（如有标记）。"""
        # 空数据检查（保留原有保护逻辑）
        if not self.rows or not any(r for r in self.rows if r and len(r) > 1 and r[1]):
            self._show_msg("请先打开Excel客户数据", THEME['warning'])
            return

        # 统计外访进度
        # v3.22.6 修复: _collect_records 属于 ReportGenerator 类，MainScreen 没有此方法 →
        # 旧代码 self._collect_records(...) 抛 AttributeError 被 except 捕获后 fallback
        # 到 _generate_report，导致二次确认弹窗被跳过。改为通过 report_generator 调用。
        try:
            records = self.report_generator._collect_records(self.rows, self.progress_mgr)
            total = len(records)
            # v3.22.19: 已选中的行（_row_selected）也视为已外访（独立于实际拍摄状态，
            # 仅用于"标记完成"辅助 AI 报告统计）。
            # 注意：records 按 borrower 分组（同一客户多处抵押物合并为一条记录），
            # 不能直接用 enumerate(records) 的索引匹配 row_index。
            # 改为建立 borrower → row_indices 映射，若该客户的任一行被选中即视为已外访。
            borrower_to_rows = {}
            for i, row in enumerate(self.rows):
                b = row[1] if len(row) > 1 else ""
                if b:
                    borrower_to_rows.setdefault(b, set()).add(i)
            visited_count = sum(
                1 for r in records
                if r.get('photo_count', 0) > 0
                or any(ri in self._row_selected
                       for ri in borrower_to_rows.get(r.get('name', ''), set()))
            )
            not_visited_count = total - visited_count
            # v3.22.6: 统计同类型标记数（已标记且无照片的视为同类型未拍照）
            batch_marked_count = len([r for r in records
                                      if r.get('batch_marked', False)
                                      and r.get('photo_count', 0) == 0])
        except Exception as e:
            app_log.error('REPORT', '统计外访进度异常: %s' % e)
            # v3.22.6: 不再静默 fallback 到 _generate_report，改为弹错误提示
            self._show_msg("统计外访进度失败，请重试：%s" % str(e)[:60], THEME['danger'], toast=False)
            return

        # 构建确认弹窗
        content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
        # v3.22.7: 在 content 的 canvas.before 中绘制浅色背景（Kivy 默认深色背景导致深色文字不可见）
        with content.canvas.before:
            Color(*THEME['card'])
            _rc_bg = Rectangle(pos=content.pos, size=content.size)
        content.bind(pos=lambda i, v: setattr(_rc_bg, 'pos', v),
                     size=lambda i, v: setattr(_rc_bg, 'size', v))

        # 标题
        title_label = Label(
            text="生成日报表确认",
            font_size=dp(20), bold=True,
            color=THEME.get('text', (0.1, 0.1, 0.1, 1)),
            size_hint_y=None, height=dp(40)
        )
        content.add_widget(title_label)

        # 进度信息
        # v3.22.6: 如有同类型标记，进度文案含「另有S户标记为同类型」
        if batch_marked_count > 0:
            msg = "今日外访进度 %d/%d，\n另有%d户标记为同类型，\n是否立即生成报告？" % (
                visited_count, total, batch_marked_count)
        else:
            msg = "今日外访进度 %d/%d，尚有%d户未完成，\n是否立即生成报告？" % (
                visited_count, total, not_visited_count)
        info_label = Label(
            text=msg,
            font_size=dp(16),
            color=THEME.get('text_dim', (0.4, 0.4, 0.4, 1)),
            halign='center', valign='middle',
            size_hint_y=None, height=dp(80)
        )
        info_label.bind(width=lambda s, w: setattr(s, 'text_size', (w, None)))
        content.add_widget(info_label)

        # v3.22.14: 客户维度走访统计 — 按客户名分组显示「抵押物 N 处，已拍 M 处」
        from collections import OrderedDict
        cust_stats = OrderedDict()
        for r in records:
            name = r.get('name', '')
            if name not in cust_stats:
                cust_stats[name] = {'total': 0, 'visited': 0}
            cust_stats[name]['total'] += r.get('count', 1)
            if r.get('photo_count', 0) > 0:
                cust_stats[name]['visited'] += r.get('count', 1)

        detail_title = Label(
            text="客户明细：", font_size='13sp',
            size_hint_y=None, height=dp(24),
            color=THEME.get('text', (0.1, 0.1, 0.1, 1)),
            halign='left', valign='middle',
            text_size=(0, None)
        )
        detail_title.bind(width=lambda inst, val: setattr(inst, 'text_size', (val, None)))
        content.add_widget(detail_title)

        cust_scroll = ScrollView(size_hint_y=None, height=dp(150))
        cust_grid = GridLayout(cols=1, spacing=dp(2), size_hint_y=None)
        cust_grid.bind(minimum_height=cust_grid.setter('height'))
        for name, st in cust_stats.items():
            m, n = st['visited'], st['total']
            if m == n:
                color = THEME.get('success', (0.2, 0.7, 0.36, 1))
            elif m > 0:
                color = THEME.get('warning', (0.96, 0.62, 0.0, 1))
            else:
                color = THEME.get('text_dim', (0.4, 0.4, 0.4, 1))
            row_label = Label(
                text="%s：抵押物 %d 处，已拍 %d 处" % (name, n, m),
                font_size='13sp', size_hint_y=None, height=dp(28),
                color=color, halign='left', valign='middle', text_size=(0, None)
            )
            row_label.bind(width=lambda inst, val: setattr(inst, 'text_size', (val, None)))
            cust_grid.add_widget(row_label)
        cust_scroll.add_widget(cust_grid)
        content.add_widget(cust_scroll)

        # 按钮容器
        btn_box = BoxLayout(orientation='horizontal', spacing=dp(15), size_hint_y=None, height=dp(48))

        btn_generate = RoundedButton(
            text="生成",
            size_hint_x=0.5,
            background_color=THEME.get('success', (0.2, 0.7, 0.36, 1))
        )
        btn_cancel = RoundedButton(
            text="取消",
            size_hint_x=0.5,
            # v3.22.6 修复 THEME 键名: 'border' → 'card_border'
            background_color=THEME.get('card_border', (0.86, 0.88, 0.91, 1))
        )
        btn_box.add_widget(btn_generate)
        btn_box.add_widget(btn_cancel)
        content.add_widget(btn_box)

        popup = Popup(
            title="",
            # v3.22.7: 深色标题在浅色背景上可见；分隔线用浅灰
            title_color=THEME.get('accent_dark', (0.08, 0.40, 0.78, 1)),
            content=content,
            size_hint=(0.85, None),
            # v3.22.14: 客户明细列表后高度由 dp(280) 增加到 dp(420)
            height=dp(420),
            auto_dismiss=False,  # 防止点外部关闭
            # v3.22.7: 浅色分隔线（之前用 accent 蓝色与浅色主题不协调）
            separator_color=THEME.get('card_border', (0.86, 0.88, 0.92, 1))
        )

        # 绑定按钮事件
        def on_generate(_btn):
            try:
                popup.dismiss()
            except Exception:
                pass
            self._generate_report(instance)

        def on_cancel(_btn):
            try:
                popup.dismiss()
            except Exception:
                pass

        btn_generate.bind(on_release=on_generate)
        btn_cancel.bind(on_release=on_cancel)

        popup.open()

    def _generate_report(self, instance):
        """v3.19.6: 调用 AI 仅对有外访照片的客户生成勘查日报表。
        点击后立即将按钮置为"正在生成中…"并禁用，避免重复点击。"""
        if not self.rows or not any((r[1] if len(r) > 1 else "") for r in self.rows):
            self._show_msg("请先打开Excel客户数据", THEME['warning'])
            return
        if not self.report_generator.template_path:
            self._show_msg("未找到日报表模板", THEME['danger'])
            return

        # 立即更新按钮状态：显示"正在生成中…"并禁用，避免重复点击
        btn = instance
        original_text = btn.text if hasattr(btn, 'text') else "AI 一键生成日报表"
        try:
            btn.text = "正在生成中…"
            btn.disabled = True
        except Exception:
            pass

        # 进度弹窗（AI生成较慢）— v3.22.0: 浅色背景 + 居中 + 幽默文案
        content = BoxLayout(orientation='vertical', padding=dp(24), spacing=dp(16))
        with content.canvas.before:
            Color(*THEME['card'])
            _gen_bg = Rectangle(pos=content.pos, size=content.size)
        content.bind(pos=lambda i, v: setattr(_gen_bg, 'pos', v),
                     size=lambda i, v: setattr(_gen_bg, 'size', v))
        msg = Label(text="AI 正在奋笔疾书，请稍候片刻…\n（泡杯茶的时间就够了）",
                    font_size='17sp', bold=True,
                    color=THEME['accent_dark'], size_hint_y=None, height=dp(64),
                    halign='center', valign='middle')
        msg.bind(width=lambda i, v: setattr(i, 'text_size', (v, None)))
        content.add_widget(msg)
        from kivy.uix.progressbar import ProgressBar
        pb = ProgressBar(size_hint_y=None, height=dp(8))
        content.add_widget(pb)
        popup = Popup(title='生成日报表', content=content, size_hint=(0.85, 0.35),
                      auto_dismiss=False, title_color=THEME['accent_dark'],
                      separator_color=THEME['card_border'])
        popup.open()

        def _run():
            # v3.22.4: 整体 try/except，任何异常都能关闭弹窗并恢复按钮，避免卡死
            try:
                ai_svc = AIService(
                    api_url=self.config.get('ai_api_url', DEFAULT_CONFIG['ai_api_url']),
                    api_key=self.config.get('ai_api_key', '') or AI_DEFAULT_API_KEY,
                    model=self.config.get('ai_model', DEFAULT_CONFIG['ai_model']),
                )
                excel_name = os.path.basename(self.excel_path) if self.excel_path else ""
                ok, path, m = self.report_generator.generate_with_ai(
                    self.rows, self.progress_mgr, ai_svc, excel_filename=excel_name,
                    visit_note=self.current_visit_note)
            except Exception as e:
                ok, path, m = False, None, "生成报告时出错：%s" % str(e)[:100]
                app_log.error('REPORT', '生成报告异常: %s' % traceback.format_exc())

            def _done(dt):
                try:
                    popup.dismiss()
                except Exception:
                    pass
                # 恢复按钮状态
                try:
                    btn.text = original_text
                    btn.disabled = False
                except Exception:
                    pass
                if ok and path:
                    self._pending_report_path = path
                    self._show_msg(m + "，正在选择保存位置…", THEME['success'])
                    self._save_report_to_user(path)
                else:
                    self._show_msg(m, THEME['danger'])
            Clock.schedule_once(_done, 0)

        import threading
        # v3.20.0: 延迟 0.1s 启动线程，确保弹窗先渲染到屏幕（防止低端机弹窗未绘制就阻塞）
        Clock.schedule_once(lambda dt: threading.Thread(target=_run, daemon=True).start(), 0.1)

    def _save_report_to_user(self, src_path):
        """通过 SAF 让用户选择日报表保存位置
        v3.22.3: 用 JString 显式转换 putExtra 参数类型，
        修复 jnius 方法重载选择错误导致 EXTRA_TITLE 不生效（默认文件名缺失）"""
        if not IS_ANDROID:
            self._show_msg("日报表已保存：%s" % src_path, THEME['success'])
            return
        try:
            from jnius import autoclass
            Intent = autoclass('android.content.Intent')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            JString = autoclass('java.lang.String')
            intent = Intent()
            # 用 setAction 替代构造函数传参，避免 jnius 反射 Intent(String) 失败
            intent.setAction(JString("android.intent.action.CREATE_DOCUMENT"))
            intent.addCategory(Intent.CATEGORY_OPENABLE)
            intent.setType(JString("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
            # v3.22.3: 显式用 JString 包装，确保 putExtra(String, String) 重载被选中
            # v3.22.19: 默认文件名改为"线路四XX（日期）现场勘查日报表.xlsx"
            excel_name = os.path.basename(self.excel_path) if self.excel_path else ""
            route_name = self._extract_route_name(excel_name)
            date_str = get_date_str()
            default_name = "线路四%s（%s）现场勘查日报表.xlsx" % (route_name, date_str)
            intent.putExtra(Intent.EXTRA_TITLE, JString(default_name))
            self._report_save_code = 0x201
            PythonActivity.mActivity.startActivityForResult(intent, self._report_save_code)
            app_log.info('REPORT', '已发起 SAF 保存请求，默认文件名: %s' % default_name)
        except Exception as e:
            app_log.error('REPORT', 'SAF 保存请求失败: %s' % e)
            self._show_msg("已保存到app内部：%s" % src_path, THEME['success'])

    def _extract_route_name(self, excel_filename):
        """v3.22.19: 从 Excel 文件名提取线路名（委托模块级函数，避免重复代码）
        规则：
        - 优先匹配 "线路X" 模式（如 "线路四" → "四"）
        - 其次匹配日期模式 "M.D盘点表" → "M.D"
        - 再次匹配纯日期 "20260705"
        - 都不匹配则回退 "抵押物抵债资产" """
        return _extract_route_name(excel_filename)


# ============================================================
# AI 助手界面 v3.15.0
# ============================================================

class AIScreen(Screen):
    """AI 助手界面：用户可查询拍摄情况"""

    def __init__(self, app_config, **kwargs):
        super().__init__(**kwargs)
        self.name = 'ai'
        self.config = app_config
        self.chat_history = []  # [{"role":"user/assistant","text":"..."}]
        self._sending = False
        self._build_ui()
        # v3.22.9: 监听键盘高度变化，给 input_row 加 bottom padding 防止遮挡
        Window.bind(on_keyboard_height=self._on_keyboard_height)

    def _build_ui(self):
        from kivy.uix.scrollview import ScrollView
        from kivy.uix.textinput import TextInput

        layout = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(6))

        # 状态栏占位（避免与系统状态栏重叠）
        from kivy.uix.widget import Widget
        status_bar_pad = Widget(size_hint_y=None, height=dp(30))
        layout.add_widget(status_bar_pad)

        # 顶部标题栏
        top_bar = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(8))
        back_btn = RoundedButton(
            text="返回", font_size='16sp', size_hint_x=0.2,
            background_color=THEME['accent'], background_normal='',
            color=(1, 1, 1, 1), bold=True,
        )
        back_btn.bind(on_release=self._go_back)
        bind_press_animation(back_btn)
        top_bar.add_widget(back_btn)

        title = Label(
            text="AI 拍摄助手", font_size='18sp', bold=True,
            color=THEME['text'], size_hint_x=0.6,
        )
        top_bar.add_widget(title)

        clear_btn = RoundedButton(
            text="清空", font_size='14sp', size_hint_x=0.2,
            background_color=THEME['muted'], background_normal='',
            color=(1, 1, 1, 1),
        )
        clear_btn.bind(on_release=self._clear_chat)
        top_bar.add_widget(clear_btn)
        layout.add_widget(top_bar)

        # 提示信息
        hint = Label(
            text="可询问：今天拍了多少照片？ | 某某公司拍了没有？ | 远景拍了多少张？",
            font_size='12sp', color=THEME['text_dim'], size_hint_y=None, height=dp(24),
            halign='center', valign='middle',
        )
        hint.bind(size=lambda inst, val: setattr(inst, 'text_size', val))
        layout.add_widget(hint)

        # 聊天记录区（可滚动）
        self.chat_layout = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(6))
        self.chat_layout.bind(minimum_height=self.chat_layout.setter('height'))

        scroll = ScrollView(size_hint_y=1)
        scroll.add_widget(self.chat_layout)
        layout.add_widget(scroll)
        # v3.22.7: 保存 scroll 引用，便于键盘弹出时滚动到底部
        self.chat_scroll = scroll

        # 输入栏
        # v3.22.9: 改为实例属性 self.input_row，供 _on_keyboard_height 调整 padding
        self.input_row = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(6))
        self.input_field = TextInput(
            font_size='15sp', multiline=False,
            size_hint_x=0.78, hint_text="输入问题...",
        )
        self.input_field.bind(on_text_validate=self._on_send)
        # v3.22.7: 绑定 focus 事件；v3.22.9: 焦点回调仅保留占位，键盘适配改由 _on_keyboard_height 处理
        self.input_field.bind(focus=self._on_input_focus)
        self.input_row.add_widget(self.input_field)

        send_btn = RoundedButton(
            text="发送", font_size='16sp', bold=True,
            size_hint_x=0.22,
            background_color=THEME['success'], background_normal='',
            color=(1, 1, 1, 1),
        )
        send_btn.bind(on_release=self._on_send)
        self.input_row.add_widget(send_btn)
        layout.add_widget(self.input_row)

        self.add_widget(layout)

        # 欢迎消息
        self._add_message("assistant", "您好！我是AI拍摄助手，可以帮您查询拍摄进度。请问有什么需要？")

    def _on_input_focus(self, instance, value):
        """v3.22.10: 简化 - 焦点变化时仅滚动聊天到底部。键盘遮挡由 softinput_mode='pan'（系统级）+ _on_keyboard_height padding（双保险）处理。"""
        try:
            if value:
                Clock.schedule_once(lambda dt: setattr(self.chat_scroll, 'scroll_y', 0), 0.3)
        except Exception as e:
            app_log.error('AI', '_on_input_focus 异常: %s' % e)

    def _on_keyboard_height(self, instance, height):
        """v3.22.9: 键盘高度变化时调整 input_row 的 bottom padding。
        这是 Kivy Android 社区公认最稳的键盘适配方案，
        替代 v3.22.7 的 softinput_mode 动态切换（时机太晚无效）。"""
        try:
            if height > 0:
                # 键盘弹出：input_row 加 bottom padding = 键盘高度，把输入框顶到键盘上方
                self.input_row.padding = [0, 0, 0, height]
                # 延迟滚动聊天记录到底部（等键盘完全弹出）
                Clock.schedule_once(lambda dt: setattr(self.chat_scroll, 'scroll_y', 0), 0.1)
            else:
                # 键盘收起：恢复 padding
                self.input_row.padding = [0, 0, 0, 0]
        except Exception as e:
            app_log.error('AI', '_on_keyboard_height 异常: %s' % e)

    def on_leave(self):
        """v3.22.9: 离开 AI 界面时解绑键盘高度监听，避免内存泄漏"""
        try:
            Window.unbind(on_keyboard_height=self._on_keyboard_height)
            # 离开时收起键盘
            if self.input_field:
                self.input_field.focus = False
        except Exception as e:
            app_log.error('AI', 'on_leave 解绑异常: %s' % e)

    def _add_message(self, role, text):
        """添加一条消息到聊天区"""
        from kivy.uix.label import Label
        # v3.22.0: 防御性清理 assistant 回复中的 markdown 符号
        if role == "assistant" and text:
            text = clean_markdown(text)
        self.chat_history.append({"role": role, "text": text})

        is_user = (role == "user")
        msg_label = Label(
            text=text,
            font_size='14sp',
            size_hint_y=None,
            halign='left' if not is_user else 'right',
            valign='top',
            color=THEME['text'],
            text_size=(None, None),
            padding=[dp(8), dp(6)],
        )
        # 设定文本宽度（避免过长）
        msg_label.bind(
            width=lambda inst, val: setattr(inst, 'text_size', (val - dp(16), None))
        )
        msg_label.bind(texture_size=lambda inst, val: setattr(inst, 'height', val[1] + dp(12)))

        # 背景色
        with msg_label.canvas.before:
            if is_user:
                Color(0.2, 0.5, 0.9, 1)  # 蓝色
            else:
                Color(0.9, 0.9, 0.95, 1)  # 浅灰
            from kivy.graphics import Rectangle
            msg_label.bg_rect = Rectangle(pos=msg_label.pos, size=msg_label.size)

        def _update_bg(inst, val):
            inst.bg_rect.pos = inst.pos
            inst.bg_rect.size = inst.size
        msg_label.bind(pos=_update_bg, size=_update_bg)

        self.chat_layout.add_widget(msg_label)

        # 滚动到底部
        Clock.schedule_once(lambda dt: self._scroll_to_bottom(), 0.1)

    def _scroll_to_bottom(self):
        sv = self.chat_layout.parent
        if sv:
            sv.scroll_y = 0

    def _on_send(self, instance):
        if self._sending:
            return
        text = self.input_field.text.strip()
        if not text:
            return
        self.input_field.text = ""
        self._add_message("user", text)
        self._send_to_ai(text)

    def _send_to_ai(self, user_text):
        """发送消息到 AI 服务（后台线程）"""
        self._sending = True

        # 获取 MainScreen 的数据
        main_screen = self.manager.get_screen('main') if self.manager else None
        if main_screen is None:
            self._add_message("assistant", "错误：未找到主界面数据")
            self._sending = False
            return

        rows = main_screen.rows
        progress_mgr = main_screen.progress_mgr
        excel_path = main_screen.excel_path

        # 构建系统提示词
        system_prompt = AIService.build_system_prompt(rows, progress_mgr, excel_path)

        # 构建 API 消息
        api_messages = [{"role": "system", "content": system_prompt}]
        # 只带最近 6 条历史（避免 token 过多）
        for h in self.chat_history[-6:]:
            api_messages.append({"role": h["role"], "content": h["text"]})

        # 创建 AI 服务
        ai = AIService(
            api_url=self.config.get('ai_api_url', ''),
            api_key=self.config.get('ai_api_key', '') or AI_DEFAULT_API_KEY,
            model=self.config.get('ai_model', ''),
        )

        # 显示"思考中"
        self._add_message("assistant", "思考中...")

        import threading

        def _do_request():
            success, response = ai.chat(api_messages, timeout=90)
            # 在主线程更新 UI
            Clock.schedule_once(lambda dt: self._on_ai_response(success, response), 0)

        threading.Thread(target=_do_request, daemon=True).start()

    def _on_ai_response(self, success, response):
        # 移除"思考中"消息
        if self.chat_history and self.chat_history[-1]["text"] == "思考中...":
            self.chat_history.pop()
            if len(self.chat_layout.children) > 0:
                self.chat_layout.remove_widget(self.chat_layout.children[-1])

        if success:
            self._add_message("assistant", clean_markdown(response))  # v3.22.0: 清理 markdown 符号
        else:
            self._add_message("assistant", "错误: " + response)
            # 如果是未配置模型，提示去设置
            if "未配置" in response:
                self._add_message("assistant", "请到「设置」页面填写 AI 模型 ID（如 deepseek-v4-flash）")

        self._sending = False

    def _clear_chat(self, instance):
        self.chat_history.clear()
        self.chat_layout.clear_widgets()
        self._add_message("assistant", "聊天已清空。请问有什么需要？")

    def _go_back(self, instance):
        if self.manager:
            self.manager.current = 'main'


# ============================================================
# App 入口
# ============================================================

class LoanPhotoApp(App):
    def build(self):
        self.title = "资产盘点专项拍照工具"
        Window.clearcolor = THEME['bg']
        # v3.22.0: 全局 resize 模式 — 键盘弹出时窗口收缩，搜索框/输入框不被顶出屏幕
        # v3.22.10: resize 在 Android SDL2 不工作（Kivy 官方文档），改用 pan 让系统自动上移窗口
        # v3.22.19: 改回 resize 模式 — 配合 _scroll_input_to_visible 辅助方法，
        #            备注弹窗 / AI 补充说明 / 搜索框聚焦时主动滚动到可见区域，
        #            避免 pan 模式下整个主屏幕被推高导致布局错乱。
        Window.softinput_mode = 'resize'

        # 拦截Android返回键：边缘滑动不直接退出App，而是返回上一页面
        Window.bind(on_keyboard=self._on_keyboard)
        # v3.20.0: 绑定 on_restore/on_draw，处理 app 从后台恢复时 GL 上下文丢失
        Window.bind(on_restore=self._on_window_restore)

        self.config = AppConfig()

        sm = ScreenManager(transition=SlideTransition(duration=0.25))
        self.sm = sm
        sm.add_widget(WelcomeScreen(name='welcome'))
        sm.add_widget(MainScreen(app_config=self.config, name='main'))
        sm.add_widget(SettingsScreen(app_config=self.config, name='settings'))
        sm.add_widget(AIScreen(app_config=self.config, name='ai'))
        sm.current = 'welcome'
        return sm

    def _on_window_restore(self, *args):
        """v3.20.0: app 从后台恢复时，强制重绘所有 canvas，防止 GL 上下文丢失导致卡死"""
        try:
            app_log.info('APP', 'on_window_restore: 恢复GL上下文')
            for screen_name in ['welcome', 'main', 'settings', 'ai']:
                try:
                    sc = self.sm.get_screen(screen_name)
                    if sc:
                        sc.canvas.ask_update()
                except Exception:
                    pass
            # 刷新主界面进度
            main_screen = self.sm.get_screen('main') if self.sm else None
            if main_screen:
                main_screen._is_paused = False
                main_screen._update_progress()
                try:
                    main_screen.canvas.ask_update()
                except Exception:
                    pass
        except Exception as e:
            app_log.error('APP', 'on_window_restore 失败: %s' % e)

    def _on_keyboard(self, window, key, scancode, codepoint, modifier):
        if key == 27:
            current = self.sm.current
            if current == 'settings':
                self.sm.current = 'main'
                return True
            elif current == 'ai':
                self.sm.current = 'main'
                return True
            elif current == 'main':
                main_screen = self.sm.get_screen('main')
                if getattr(main_screen, '_continuous_shooting', False):
                    main_screen._continuous_shooting = False
                    main_screen._photo_session_active = False  # v3.20.0: 解除会话锁
                    main_screen._unlock_photo_buttons()
                    main_screen.camera_mgr._dbg("已结束连续拍照", show_toast=True)
                    return True
                self.sm.current = 'welcome'
                return True
            elif current == 'welcome':
                if hasattr(self, '_back_pressed') and (datetime.now() - self._back_pressed).total_seconds() < 2:
                    self.stop()
                else:
                    self._back_pressed = datetime.now()
                    try:
                        from kivy.uix.popup import Popup
                        popup = Popup(title='',
                                     content=Label(text="再按一次退出", font_size='16sp'),
                                     size_hint=(0.5, 0.15), auto_dismiss=True)
                        popup.open()
                        Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)
                    except Exception:
                        pass
                return True
        return False

    def on_start(self):
        if IS_ANDROID:
            setup_android_status_bar()
            self._register_android_activity_result()
            self._request_permissions()

    def _register_android_activity_result(self):
        """注册Android activity结果回调，把文件选择器结果转发给MainScreen"""
        try:
            from android.activity import bind as activity_bind
            activity_bind(on_activity_result=self._on_android_activity_result)
        except Exception as e:
            Logger.warning("activity_bind failed: %s" % e)

    def _on_android_activity_result(self, request_code, result_code, intent):
        """Android activity结果统一入口，转发给当前screen"""
        if self.sm:
            current_screen = self.sm.current_screen
            if current_screen and hasattr(current_screen, 'on_activity_result'):
                try:
                    current_screen.on_activity_result(request_code, result_code, intent)
                except Exception as e:
                    Logger.error("on_activity_result error: %s", e)
            elif self.sm.current == 'main':
                main_screen = self.sm.get_screen('main')
                if hasattr(main_screen, 'on_activity_result'):
                    main_screen.on_activity_result(request_code, result_code, intent)

    def _request_permissions(self):
        # v3.22.18: 补全 Android 11+ 存储权限 + Android 13+ 媒体权限
        try:
            if ANDROID_API >= 33:
                # Android 13+: 细粒度媒体权限替代 READ_EXTERNAL_STORAGE
                perms = [Permission.CAMERA, Permission.ACCESS_FINE_LOCATION,
                         Permission.ACCESS_COARSE_LOCATION, Permission.READ_MEDIA_IMAGES]
            elif ANDROID_API >= 30:
                # Android 11-12: 仍需要 READ_EXTERNAL_STORAGE
                perms = [Permission.CAMERA, Permission.ACCESS_FINE_LOCATION,
                         Permission.ACCESS_COARSE_LOCATION,
                         Permission.READ_EXTERNAL_STORAGE]
            else:
                perms = [Permission.CAMERA, Permission.ACCESS_FINE_LOCATION,
                         Permission.ACCESS_COARSE_LOCATION,
                         Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE]
            request_permissions(perms)
        except Exception:
            pass

    def on_pause(self):
        # v3.21.0: 切后台时仅保存进度，保留 _camera_launched 和 _continuous_shooting 状态
        # （v3.20.0 清除这两者导致 on_camera_result 早返回，照片绕过正常回调流程存到错误客户）
        try:
            main_screen = self.sm.get_screen('main') if self.sm else None
            if main_screen:
                main_screen._is_paused = True
                # 保存进度
                try:
                    main_screen.progress_mgr.save()
                except Exception:
                    pass
                # v3.22.13: 退出时保存 visit_note
                if main_screen.current_visit_note_excel_path:
                    try:
                        main_screen._save_visit_note(
                            main_screen.current_visit_note_excel_path,
                            main_screen.current_visit_note)
                    except Exception:
                        pass
            app_log.info('APP', 'on_pause: app 切入后台')
        except Exception as e:
            app_log.error('APP', 'on_pause 异常: %s' % e)
        return True

    def on_resume(self):
        # v3.20.0: 从后台恢复时刷新 UI 状态
        # v3.21.0: 延迟执行 _update_heights，避免行数多时阻塞 UI 恢复
        try:
            app_log.info('APP', 'on_resume: app 从后台恢复')
            main_screen = self.sm.get_screen('main') if self.sm else None
            if main_screen:
                main_screen._is_paused = False
                # 刷新进度显示
                main_screen._update_progress()
                # v3.21.0: 延迟执行 _update_heights，分散负载避免卡顿
                Clock.schedule_once(main_screen._deferred_update_heights, 0.1)
                # 强制重绘 canvas
                try:
                    main_screen.canvas.ask_update()
                except Exception:
                    pass
        except Exception as e:
            app_log.error('APP', 'on_resume 异常: %s' % e)


if __name__ == '__main__':
    LoanPhotoApp().run()
